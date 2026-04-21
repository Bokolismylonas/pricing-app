import os
import io
import base64
import json
import re
import uuid
import shutil
import time
import gc
from difflib import SequenceMatcher
from pathlib import Path
from datetime import date, datetime, timedelta, timezone

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
import stripe
from supabase import create_client, Client
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pdfplumber
from pypdf import PdfReader, PdfWriter
try:
    from openai import OpenAI
except Exception:
    OpenAI = None
from storage import (
    list_comparisons,
    save_new_comparison,
    update_comparison,
    get_comparison,
    delete_comparison,
    build_display_label,
)
from central_engine import CentralMatchEngine, suggest_product


BASE_DIR = Path(__file__).parent
ASSETS_DIR = BASE_DIR / "assets"


def resolve_asset_path(filename: str) -> Path:
    candidates = [
        ASSETS_DIR / filename,
        BASE_DIR / filename,
        Path("/mnt/data") / filename,
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return ASSETS_DIR / filename


FULL_LOGO_PATH = resolve_asset_path("pricingtool-final-logo-dark.svg")
ICON_LOGO_PATH = resolve_asset_path("pricingtool-icon-dark.svg")
PAGE_ICON = str(ICON_LOGO_PATH) if ICON_LOGO_PATH.exists() else "💎"

# -------------------------------------------------
# APP CONFIG
# -------------------------------------------------
st.set_page_config(
    page_title="Pricing Tool – Compare Products & Export Excel Reports",
    page_icon=PAGE_ICON,
    layout="wide",
)

st.markdown(
    """
    <meta name="description" content="Upload supplier data, compare products and export polished Excel reports instantly with Pricing Tool.">
    <meta name="robots" content="index,follow">
    <link rel="canonical" href="https://www.pricingtool.gr/">
    """,
    unsafe_allow_html=True,
)

stripe.api_key = os.getenv("STRIPE_SECRET_KEY")


# -------------------------------------------------
# RENDER -> CREATE .streamlit/secrets.toml FROM ENV
# -------------------------------------------------
STREAMLIT_DIR = BASE_DIR / ".streamlit"
STREAMLIT_DIR.mkdir(parents=True, exist_ok=True)
SECRETS_FILE = STREAMLIT_DIR / "secrets.toml"


def _escape_toml(value: str) -> str:
    return str(value).replace("\\", "\\\\").replace('"', '\\"')


def ensure_render_secrets_file():
    auth_redirect_uri = os.getenv("AUTH_REDIRECT_URI", "")
    auth_cookie_secret = os.getenv("AUTH_COOKIE_SECRET", "")
    google_client_id = os.getenv("AUTH_CLIENT_ID", "")
    google_client_secret = os.getenv("AUTH_CLIENT_SECRET", "")
    google_server_metadata_url = os.getenv("AUTH_SERVER_METADATA_URL", "")
    ms_client_id = os.getenv("MS_CLIENT_ID", "")
    ms_client_secret = os.getenv("MS_CLIENT_SECRET", "")
    ms_server_metadata_url = os.getenv("MS_SERVER_METADATA_URL", "")
    supabase_url = os.getenv("SUPABASE_URL", "")
    supabase_key = os.getenv("SUPABASE_KEY", "")

    if not all(
        [
            auth_redirect_uri,
            auth_cookie_secret,
            google_client_id,
            google_client_secret,
            google_server_metadata_url,
        ]
    ):
        return

    content = f'''SUPABASE_URL = "{_escape_toml(supabase_url)}"
SUPABASE_KEY = "{_escape_toml(supabase_key)}"

[auth]
redirect_uri = "{_escape_toml(auth_redirect_uri)}"
cookie_secret = "{_escape_toml(auth_cookie_secret)}"

[auth.google]
client_id = "{_escape_toml(google_client_id)}"
client_secret = "{_escape_toml(google_client_secret)}"
server_metadata_url = "{_escape_toml(google_server_metadata_url)}"
'''

    if ms_client_id and ms_client_secret and ms_server_metadata_url:
        content += f'''
[auth.microsoft]
client_id = "{_escape_toml(ms_client_id)}"
client_secret = "{_escape_toml(ms_client_secret)}"
server_metadata_url = "{_escape_toml(ms_server_metadata_url)}"
'''
        # Keep compatibility if env var typo not present
        content = content.replace(
            os.getenv("MS_SERVER_METADATA_URL", ""),
            _escape_toml(ms_server_metadata_url),
        )

    SECRETS_FILE.write_text(content, encoding="utf-8")


ensure_render_secrets_file()


# -------------------------------------------------
# UI STYLE
# -------------------------------------------------
st.markdown(
    """
    <style>
    .main .block-container {
        padding-top: 1.2rem;
        padding-bottom: 2rem;
        max-width: 1420px;
    }

    .app-hero {
        padding: 28px 32px;
        border-radius: 22px;
        background: linear-gradient(135deg, #2563eb 0%, #1d4ed8 48%, #1e3a8a 100%);
        color: white;
        border: 1px solid rgba(255,255,255,0.10);
        margin-bottom: 1.2rem;
        box-shadow: 0 14px 34px rgba(0,0,0,0.18);
    }

    .app-card {
        background: linear-gradient(180deg, #0f172a 0%, #111827 100%);
        border: 1px solid rgba(255,255,255,0.08);
        border-radius: 18px;
        padding: 22px 24px;
        color: white;
        box-shadow: 0 8px 22px rgba(0,0,0,0.12);
        margin-bottom: 14px;
    }

    .locked-wrap {
        max-width: 760px;
        margin: 30px auto 0 auto;
        padding: 32px;
        border-radius: 22px;
        background: linear-gradient(180deg, #0f172a 0%, #111827 100%);
        border: 1px solid rgba(255,255,255,0.10);
        box-shadow: 0 20px 50px rgba(0,0,0,0.22);
        text-align: center;
        color: white;
    }

    .locked-title {
        font-size: 32px;
        font-weight: 800;
        margin-bottom: 10px;
        letter-spacing: -0.03em;
    }

    .locked-subtitle {
        font-size: 16px;
        color: #cbd5e1;
        line-height: 1.65;
        margin-bottom: 22px;
    }

    .locked-badge {
        display: inline-block;
        padding: 6px 12px;
        border-radius: 999px;
        background: rgba(96,165,250,0.12);
        border: 1px solid rgba(96,165,250,0.25);
        color: #bfdbfe;
        font-size: 13px;
        margin-bottom: 18px;
    }

    
.login-shell {
        max-width: 620px;
        margin: 56px auto 18px auto;
        padding: 44px 42px 34px 42px;
        border-radius: 28px;
        background: linear-gradient(180deg, #020617 0%, #0f172a 100%);
        border: 1px solid rgba(255,255,255,0.08);
        box-shadow: 0 24px 60px rgba(0,0,0,0.30);
        text-align: center;
        color: white;
    }

    .login-shell .login-header-logo {
        margin: 0 0 14px 0;
    }

    .login-shell .login-header-logo img {
        height: 46px;
        width: auto;
        max-width: 100%;
        display: inline-block;
    }

    .login-shell p {
        margin: 0;
        font-size: 18px;
        color: #cbd5e1;
    }

    .login-badge {
        display: inline-block;
        padding: 7px 14px;
        border-radius: 999px;
        background: rgba(96,165,250,0.16);
        border: 1px solid rgba(147,197,253,0.28);
        color: #dbeafe;
        font-size: 13px;
        font-weight: 700;
        margin-bottom: 16px;
    }

    .login-shell-premium {
        background:
            radial-gradient(circle at top right, rgba(59,130,246,0.22), transparent 28%),
            linear-gradient(180deg, #020617 0%, #0f172a 100%);
    }

    .login-feature-row {
        display: grid;
        grid-template-columns: 1fr 1fr;
        gap: 12px;
        margin-top: 22px;
        text-align: left;
    }

    .login-feature-card {
        padding: 14px 16px;
        border-radius: 16px;
        background: rgba(255,255,255,0.05);
        border: 1px solid rgba(255,255,255,0.08);
    }

    .login-feature-title {
        font-size: 14px;
        font-weight: 800;
        color: #ffffff;
        margin-bottom: 6px;
    }

    .login-feature-text {
        font-size: 13px;
        line-height: 1.5;
        color: #cbd5e1;
    }

    .login-provider-preview {
        display: grid;
        grid-template-columns: 1fr 1fr;
        gap: 12px;
        margin-bottom: 16px;
    }

    .login-provider-card {
        display: flex;
        align-items: center;
        gap: 12px;
        padding: 14px 16px;
        border-radius: 16px;
        background: #ffffff;
        border: 1px solid #e5e7eb;
        color: #111827;
        text-align: left;
        box-shadow: 0 10px 22px rgba(15,23,42,0.10);
    }

    .login-provider-card-ms {
        background: #f8fafc;
    }

    .provider-logo-box {
        width: 38px;
        height: 38px;
        border-radius: 10px;
        background: #ffffff;
        border: 1px solid #e5e7eb;
        display: inline-flex;
        align-items: center;
        justify-content: center;
        flex-shrink: 0;
    }

    .provider-logo-box.ms-box {
        background: #ffffff;
    }

    .provider-logo-google {
        width: 20px;
        height: 20px;
        border-radius: 999px;
        display: inline-flex;
        align-items: center;
        justify-content: center;
        font-size: 15px;
        font-weight: 800;
        color: #4285f4;
        background: #ffffff;
    }

    .provider-logo-microsoft {
        width: 18px;
        height: 18px;
        display: grid;
        grid-template-columns: 1fr 1fr;
        gap: 2px;
    }

    .provider-logo-microsoft span:nth-child(1) { background: #f25022; }
    .provider-logo-microsoft span:nth-child(2) { background: #7fba00; }
    .provider-logo-microsoft span:nth-child(3) { background: #00a4ef; }
    .provider-logo-microsoft span:nth-child(4) { background: #ffb900; }

    .provider-logo-microsoft span {
        display: block;
        width: 8px;
        height: 8px;
        border-radius: 1px;
    }

    .provider-meta {
        display: flex;
        flex-direction: column;
        gap: 2px;
    }

    .provider-title {
        font-size: 14px;
        font-weight: 800;
        color: #111827;
        line-height: 1.2;
    }

    .provider-subtitle {
        font-size: 12px;
        color: #6b7280;
        line-height: 1.3;
    }

    .login-actions-label {
        text-align: center;
        margin: 4px 0 12px 0;
        font-size: 13px;
        color: #94a3b8;
        font-weight: 600;
    }

    .login-note {
        text-align: center;
        margin-top: 16px;
        font-size: 13px;
        color: #94a3b8;
    }

    .provider-chip {
        display: inline-block;
        padding: 7px 12px;
        border-radius: 999px;
        background: rgba(255,255,255,0.06);
        border: 1px solid rgba(255,255,255,0.08);
        color: #cbd5e1;
        font-size: 13px;
        margin-bottom: 18px;
    }

    div[data-testid="stSidebar"] {
        border-right: 1px solid rgba(255,255,255,0.08);
    }

    div[data-testid="stButton"] > button {
        border-radius: 12px;
        font-size: 15px;
        font-weight: 700;
        border: 1px solid rgba(255,255,255,0.10);
        box-shadow: 0 4px 12px rgba(0,0,0,0.08);
    }

    div[data-testid="stButton"] > button:hover {
        border: 1px solid rgba(59,130,246,0.55);
        box-shadow: 0 6px 14px rgba(59,130,246,0.16);
    }

    div[data-testid="stDownloadButton"] > button {
        border-radius: 12px;
        font-size: 15px;
        font-weight: 700;
        border: 1px solid rgba(255,255,255,0.10);
        box-shadow: 0 4px 12px rgba(0,0,0,0.08);
    }

    div[data-testid="stDownloadButton"] > button:hover {
        border: 1px solid rgba(16,185,129,0.55);
        box-shadow: 0 6px 14px rgba(16,185,129,0.16);
    }

    div[data-testid="stMetric"] {
        background: rgba(255,255,255,0.03);
        border: 1px solid rgba(255,255,255,0.08);
        padding: 10px 12px;
        border-radius: 16px;
    }

    .stDataFrame {
        border: 1px solid rgba(255,255,255,0.10);
        border-radius: 14px;
        overflow: hidden;
    }

    .stAlert {
        border-radius: 12px;
        border: 1px solid rgba(255,255,255,0.08);
    }

    h2, h3 {
        letter-spacing: -0.02em;
    }

    hr {
        border-color: rgba(255,255,255,0.08);
    }

    .hero-logo-wrap {
        margin-bottom: 14px;
        animation: fadeInLogo 0.8s ease forwards;
    }

    .hero-logo-wrap img {
        height: 88px;
        width: auto;
        max-width: 100%;
        display: inline-block;
        transition: transform 0.25s ease, filter 0.25s ease;
    }

    .hero-logo-wrap img:hover {
        transform: scale(1.04);
        filter: drop-shadow(0 6px 18px rgba(59,130,246,0.35));
    }

    @keyframes fadeInLogo {
        from {
            opacity: 0;
            transform: translateY(6px);
        }
        to {
            opacity: 1;
            transform: translateY(0);
        }
    }

    @media (max-width: 768px) {
        .hero-logo-wrap img {
            height: 58px;
        }
    }

    .row-nav-wrap {
        display: flex;
        justify-content: center;
        align-items: center;
        gap: 8px;
        margin: 10px 0 4px 0;
        flex-wrap: nowrap;
    }

    .row-nav-btn {
        width: 40px;
        height: 40px;
        min-width: 40px;
        border-radius: 10px;
        border: 1px solid rgba(148,163,184,0.35);
        background: #ffffff;
        color: #111827;
        font-size: 18px;
        font-weight: 700;
        line-height: 1;
        display: inline-flex;
        align-items: center;
        justify-content: center;
        box-shadow: 0 4px 12px rgba(0,0,0,0.08);
        cursor: pointer;
        transition: transform 0.15s ease, box-shadow 0.15s ease, opacity 0.15s ease;
    }

    .row-nav-btn:hover {
        transform: translateY(-1px);
        box-shadow: 0 6px 14px rgba(59,130,246,0.16);
    }

    .row-nav-btn:disabled {
        opacity: 0.38;
        cursor: default;
        box-shadow: none;
    }

    @media (max-width: 768px) {
        .row-nav-wrap {
            justify-content: center;
            gap: 10px;
            margin: 8px 0 0 0;
        }

        .row-nav-btn {
            width: 42px;
            height: 42px;
            min-width: 42px;
            border-radius: 12px;
            font-size: 18px;
        }
    }

    .save-panel-card {
        padding: 16px 18px;
        border-radius: 16px;
        background: rgba(255,255,255,0.03);
        border: 1px solid rgba(255,255,255,0.08);
        margin-top: 10px;
        margin-bottom: 10px;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


# -------------------------------------------------
# AUTH / USER HELPERS
# -------------------------------------------------
def is_logged_in():
    try:
        return bool(st.user.is_logged_in)
    except Exception:
        return False


def get_current_user_id():
    try:
        return st.user.get("sub") or st.user.get("email") or "anonymous"
    except Exception:
        return "anonymous"


def get_current_user_email():
    try:
        return st.user.get("email", "").strip().lower()
    except Exception:
        return ""


def get_current_user_name():
    try:
        return st.user.get("name", "").strip()
    except Exception:
        return ""


def render_logo_if_available(path, width=None):
    try:
        if Path(path).exists():
            st.image(str(path), width=width)
            return True
    except Exception:
        pass
    return False


def show_login_screen():
    top_left, top_mid, top_right = st.columns([1.0, 2.5, 1.0])

    with top_mid:
        login_header_logo_html = ""
        if FULL_LOGO_PATH.exists():
            login_logo_b64 = base64.b64encode(FULL_LOGO_PATH.read_bytes()).decode("utf-8")
            login_header_logo_html = (
                '<div class="login-header-logo">'
                f'<img src="data:image/svg+xml;base64,{login_logo_b64}" alt="Pricing Tool logo" />'
                '</div>'
            )

        st.markdown(
            f"""
            <div class="login-shell login-shell-premium">
                <div class="login-badge">Secure workspace access</div>
                {login_header_logo_html}
                <p>Upload supplier sources, compare products and export polished Excel reports.</p>
                <div class="login-feature-row">
                    <div class="login-feature-card">
                        <div class="login-feature-title">Fast comparison</div>
                        <div class="login-feature-text">Build, save and reload pricing scenarios in seconds.</div>
                    </div>
                    <div class="login-feature-card">
                        <div class="login-feature-title">Protected access</div>
                        <div class="login-feature-text">Sign in securely with your Google or Microsoft account.</div>
                    </div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    btn_left, btn_mid, btn_right = st.columns([1.1, 2.2, 1.1])

    with btn_mid:
        st.markdown(
            """
            <div class="provider-chip">Choose a sign-in provider</div>
            <div class="login-provider-preview">
                <div class="login-provider-card">
                    <span class="provider-logo-box">
                        <span class="provider-logo-google">G</span>
                    </span>
                    <span class="provider-meta">
                        <span class="provider-title">Continue with Google</span>
                        <span class="provider-subtitle">Use your Google account</span>
                    </span>
                </div>
                <div class="login-provider-card login-provider-card-ms">
                    <span class="provider-logo-box ms-box">
                        <span class="provider-logo-microsoft">
                            <span></span><span></span><span></span><span></span>
                        </span>
                    </span>
                    <span class="provider-meta">
                        <span class="provider-title">Sign in with Microsoft</span>
                        <span class="provider-subtitle">Use your Microsoft account</span>
                    </span>
                </div>
            </div>
            <div class="login-actions-label">Use the buttons below to continue</div>
            """,
            unsafe_allow_html=True,
        )

        if st.button(
            "Continue with Google",
            use_container_width=True,
            key="login_google_button",
        ):
            st.login("google")

        st.write("")

        if st.button(
            "Sign in with Microsoft",
            use_container_width=True,
            key="login_microsoft_button",
        ):
            st.login("microsoft")

        st.markdown(
            '<div class="login-note">Secure login • No passwords stored • Existing app flow unchanged</div>',
            unsafe_allow_html=True,
        )


# -------------------------------------------------
# SUPABASE
# -------------------------------------------------
@st.cache_resource
def get_supabase() -> Client:
    return create_client(
        st.secrets["SUPABASE_URL"],
        st.secrets["SUPABASE_KEY"],
    )


# -------------------------------------------------
# FILE STORAGE
# -------------------------------------------------
PERSIST_ROOT = Path(os.getenv("PERSIST_ROOT", "/var/data"))
PERSIST_ROOT.mkdir(parents=True, exist_ok=True)

ROOT_STORAGE = PERSIST_ROOT
ADMIN_DIR = ROOT_STORAGE / "_admin"
ADMIN_DIR.mkdir(parents=True, exist_ok=True)

USERS_REGISTRY_FILE = ADMIN_DIR / "users_registry.json"
COMPANIES_REGISTRY_FILE = ADMIN_DIR / "companies_registry.json"

MAIN_CODES = ["SINIAT", "KNAUF", "SAINT_GOBAIN"]
ADMIN_EMAILS = ["gmyl13@gmail.com"]

TEMPLATE_FILE = BASE_DIR / "templates" / "source_template_english.xlsx"


# -------------------------------------------------
# JSON HELPERS
# -------------------------------------------------
def now_iso():
    return datetime.utcnow().isoformat()


def now_utc():
    return datetime.utcnow()


def parse_iso(value):
    try:
        return datetime.fromisoformat(value)
    except Exception:
        return None



def short_session_id(sid):
    try:
        return str(sid)[:6]
    except:
        return "-"
def format_duration_from_iso(iso_str):
    dt = parse_iso(iso_str)
    if dt is None:
        return "-"
    delta = now_utc() - dt
    total_seconds = int(max(0, delta.total_seconds()))
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    if hours > 0:
        return f"{hours}h {minutes}m"
    return f"{minutes}m"


def summarize_active_sessions(sessions):
    items = []
    for idx, s in enumerate(sessions or [], start=1):
        sid = short_session_id(s.get("session_id"))
        items.append(
            f"S{idx} ({sid}): {format_duration_from_iso(s.get('started_at', ''))} / {format_duration_from_iso(s.get('last_seen', ''))} ago"
        )
    return " | ".join(items)



def get_comparison_state_signature():
    relevant = {}
    prefixes = ("row_", "select_", "carry_forward_")
    direct_keys = {
        "row_ids",
        "next_row_id",
        "comparison_company_selection",
        "comparison_name_input",
        "current_comparison_id",
        "selected_export_fields",
    }
    excluded_keys = {
        "sidebar_navigation",
        "comparison_dirty",
        "comparison_saved_signature",
        "comparison_last_active_view",
        "show_leave_comparison_prompt",
        "pending_leave_target_view",
        "save_as_exit_name",
        "active_comparison_label",
        "leave_prompt_step",
        "pending_focus_row_id",
        "bulk_discount_success_message",
        "comparison_loaded_success_message",
        "pending_load_payload",
        "pending_loaded_comparison_id",
        "pending_loaded_comparison_name",
        "pending_clear_comparison",
        "checkout_email",
        "checkout_url",
        "checkout_created_at",
        "app_session_id",
    }

    for key in list(st.session_state.keys()):
        if key in excluded_keys:
            continue
        if key in direct_keys or key.startswith(prefixes):
            try:
                value = st.session_state.get(key)
                if isinstance(value, Path):
                    value = str(value)
                relevant[key] = value
            except Exception:
                pass

    return json.dumps(relevant, sort_keys=True, ensure_ascii=False, default=str)



def get_current_selected_codes_from_state():
    company_options_for_state = {
        f"{row['name']} ({row['code']})": row["code"] for _, row in companies_df.iterrows()
    }

    selected_codes = [
        company_options_for_state[x]
        for x in st.session_state.get("comparison_company_selection", [])
        if x in company_options_for_state
    ]

    if selected_codes:
        return selected_codes

    known_codes = set(companies_df["code"].astype(str).tolist())
    inferred_codes = []

    for code in known_codes:
        has_select_key = f"select_{code}" in st.session_state
        has_carry_key = f"carry_forward_{code}" in st.session_state
        has_row_key = any(
            key.startswith("row_") and f"_{code}_" in key
            for key in st.session_state.keys()
        )

        if has_select_key or has_carry_key or has_row_key:
            inferred_codes.append(code)

    return inferred_codes


def get_comparison_payload_signature_from_state():
    selected_displays = list(st.session_state.get("comparison_company_selection", []))
    selected_codes = get_current_selected_codes_from_state()
    row_ids = list(st.session_state.get("row_ids", []))

    stable_payload = {
        "row_ids": row_ids,
        "comparison_company_selection": selected_displays,
        "comparison_name_input": st.session_state.get("comparison_name_input", ""),
        "current_comparison_id": st.session_state.get("current_comparison_id"),
        "selected_export_fields": list(st.session_state.get("selected_export_fields", [])),
    }

    for code in selected_codes:
        stable_payload[f"select_{code}"] = st.session_state.get(f"select_{code}", "")
        stable_payload[f"carry_forward_{code}"] = bool(st.session_state.get(f"carry_forward_{code}", False))

    for row_id in row_ids:
        for code in selected_codes:
            stable_payload[f"row_{row_id}_{code}_product"] = st.session_state.get(
                f"row_{row_id}_{code}_product", ""
            )
            stable_payload[get_manual_final_price_data_key(row_id, code)] = str(
                st.session_state.get(get_manual_final_price_data_key(row_id, code), "") or ""
            ).strip()
            for j in range(1, 6):
                disc_key = f"row_{row_id}_{code}_disc_{j}"
                try:
                    stable_payload[disc_key] = float(st.session_state.get(disc_key, 0.0) or 0.0)
                except Exception:
                    stable_payload[disc_key] = 0.0

    return json.dumps(stable_payload, sort_keys=True, ensure_ascii=False, default=str)


def mark_comparison_clean():
    st.session_state["comparison_saved_signature"] = get_comparison_state_signature()
    st.session_state["comparison_baseline_state_json"] = get_comparison_payload_signature_from_state()
    st.session_state["comparison_clean_generation"] = st.session_state.get("comparison_edit_generation", 0)
    st.session_state["comparison_dirty"] = False
    st.session_state["comparison_user_modified"] = False


def comparison_has_meaningful_content():
    if st.session_state.get("current_comparison_id"):
        return True

    if len(st.session_state.get("row_ids", [])) > 1:
        return True

    if st.session_state.get("comparison_company_selection", []):
        return True

    if st.session_state.get("comparison_name_input", "").strip():
        return True

    for key, value in st.session_state.items():
        if key.startswith("row_") and key.endswith("_product") and str(value).strip():
            return True

        if key.startswith("row_") and "_disc_" in key:
            try:
                if float(value or 0) != 0:
                    return True
            except Exception:
                pass

    return False


def refresh_comparison_dirty_state():
    if not comparison_has_meaningful_content():
        st.session_state["comparison_dirty"] = False
        return

    baseline = st.session_state.get("comparison_baseline_state_json", "")
    if not baseline:
        st.session_state["comparison_dirty"] = False
        return

    current = get_comparison_payload_signature_from_state()
    st.session_state["comparison_dirty"] = (current != baseline)


def has_real_changes_against_loaded_baseline():
    return bool(
        st.session_state.get("comparison_user_modified", False)
        and comparison_has_meaningful_content()
    )


def has_unsaved_comparison_changes():
    return bool(
        st.session_state.get("comparison_user_modified", False)
        and comparison_has_meaningful_content()
    )


def mark_comparison_dirty():
    if not comparison_has_meaningful_content():
        st.session_state["comparison_dirty"] = False
        st.session_state["comparison_user_modified"] = False
        return
    st.session_state["comparison_edit_generation"] = st.session_state.get("comparison_edit_generation", 0) + 1
    st.session_state["comparison_dirty"] = True
    st.session_state["comparison_user_modified"] = True


@st.cache_data(show_spinner=False)
def load_prepared_catalog_from_file(file_path_str: str, modified_ts: float):
    file_path = Path(file_path_str)
    df = pd.read_excel(file_path, sheet_name="PRICELIST")
    df.columns = [str(c).strip() for c in df.columns]

    def _find_col(local_df, names):
        cols = {str(c).strip().lower(): c for c in local_df.columns}
        for n in names:
            if n.lower() in cols:
                return cols[n.lower()]
        return None

    sap = _find_col(df, ["SAP", "Κωδικός SAP"])
    prod = _find_col(df, ["Product", "Προϊόν"])
    price = _find_col(df, ["Price", "Τιμή €/ΜΜ", "Τιμή", "Price €/MM"])
    mm = _find_col(df, ["ΜΜ πώλησης", "MM", "Unit", "ΜΜ"])
    pack = _find_col(df, ["Συσκευασία", "Package", "pack"])
    category = _find_col(df, ["Κατηγορία", "Category", "category"])

    if not sap or not prod or not price:
        return None

    out = pd.DataFrame()
    out["SAP"] = df[sap].astype(str).str.strip()
    out["Product"] = df[prod].astype(str).str.strip()
    out["Price"] = pd.to_numeric(df[price], errors="coerce")
    out["MM"] = df[mm].astype(str).str.strip() if mm else ""
    out["Package"] = df[pack].astype(str).str.strip() if pack else ""
    out["Category"] = df[category].astype(str).str.strip() if category else ""
    out = out.dropna(subset=["Price"])
    out = out[out["Price"] > 0]
    out = out.reset_index(drop=True)
    out["DISPLAY"] = out["Product"] + " | SAP " + out["SAP"]
    return out


def load_users_registry():
    if not USERS_REGISTRY_FILE.exists():
        return []
    try:
        return json.loads(USERS_REGISTRY_FILE.read_text(encoding="utf-8"))
    except Exception:
        return []


def save_users_registry(data):
    USERS_REGISTRY_FILE.write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def load_companies_registry():
    if not COMPANIES_REGISTRY_FILE.exists():
        return []
    try:
        return json.loads(COMPANIES_REGISTRY_FILE.read_text(encoding="utf-8"))
    except Exception:
        return []


def save_companies_registry(data):
    COMPANIES_REGISTRY_FILE.write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def load_pdf_ai_usage_log():
    if not PDF_AI_USAGE_FILE.exists():
        return []
    try:
        data = json.loads(PDF_AI_USAGE_FILE.read_text(encoding="utf-8"))
        return data if isinstance(data, list) else []
    except Exception:
        return []


def append_pdf_ai_usage_log(entry: dict):
    rows = load_pdf_ai_usage_log()
    rows.append(entry)
    PDF_AI_USAGE_FILE.write_text(json.dumps(rows[-500:], ensure_ascii=False, indent=2, default=str), encoding="utf-8")


def get_openai_client_for_pdf_extraction():
    api_key = os.getenv("OPENAI_API_KEY", "")
    if not api_key:
        try:
            api_key = st.secrets.get("OPENAI_API_KEY", "")
        except Exception:
            api_key = ""
    if not api_key or OpenAI is None:
        return None
    try:
        return OpenAI(api_key=api_key)
    except Exception:
        return None


def current_month_pdf_ai_usage_summary():
    rows = load_pdf_ai_usage_log()
    prefix = datetime.now().strftime("%Y-%m")
    month_rows = [r for r in rows if str(r.get("timestamp", "")).startswith(prefix)]
    return {
        "calls": len(month_rows),
        "pages": sum(int(r.get("pages", 0) or 0) for r in month_rows),
        "prompt_tokens": sum(int(r.get("prompt_tokens", 0) or 0) for r in month_rows),
        "completion_tokens": sum(int(r.get("completion_tokens", 0) or 0) for r in month_rows),
        "rows": sum(int(r.get("rows", 0) or 0) for r in month_rows),
        "entries": month_rows[-20:],
    }


# -------------------------------------------------
# BILLING / PLANS
# -------------------------------------------------
TRIAL_DAYS = 2
MAX_ACTIVE_SESSIONS = 2
STALE_SESSION_TIMEOUT_HOURS = 12


def get_stripe_subscription_row(email: str):
    if not email:
        return None

    try:
        customers = stripe.Customer.list(email=email.strip().lower(), limit=1)
        if not customers.data:
            return None

        customer = customers.data[0]
        subs = stripe.Subscription.list(customer=customer.id, status="all", limit=20)

        if not subs.data:
            return {
                "email": email.strip().lower(),
                "is_active": False,
                "billing_status": "free",
                "stripe_customer_id": customer.id,
                "stripe_subscription_id": None,
            }

        preferred = None

        for sub in subs.data:
            if sub.status == "active":
                preferred = sub
                break

        if preferred is None:
            for sub in subs.data:
                if sub.status in [
                    "trialing",
                    "past_due",
                    "unpaid",
                    "canceled",
                    "incomplete",
                    "incomplete_expired",
                ]:
                    preferred = sub
                    break

        if preferred is None:
            preferred = subs.data[0]

        return {
            "email": email.strip().lower(),
            "is_active": preferred.status == "active",
            "billing_status": preferred.status,
            "stripe_customer_id": customer.id,
            "stripe_subscription_id": preferred.id,
        }

    except Exception as e:
        print("STRIPE BILLING CHECK ERROR:", repr(e))
        return None


from billing import create_checkout_session


def get_checkout_url(user_email: str):
    if not user_email:
        return None

    cache_email = st.session_state.get("checkout_email")
    cache_url = st.session_state.get("checkout_url")
    cache_ts = st.session_state.get("checkout_created_at")

    if cache_email == user_email and cache_url and cache_ts:
        try:
            created_at = datetime.fromisoformat(cache_ts)
            if datetime.utcnow() - created_at < timedelta(minutes=20):
                return cache_url
        except Exception:
            pass

    try:
        url = create_checkout_session(user_email)
        st.session_state["checkout_email"] = user_email
        st.session_state["checkout_url"] = url
        st.session_state["checkout_created_at"] = datetime.utcnow().isoformat()
        return url
    except Exception as e:
        st.error(f"Stripe error: {e}")
        return None


# -------------------------------------------------
# USER / COMPANY HELPERS
# -------------------------------------------------
def normalize_domain(domain: str) -> str:
    return str(domain).strip().lower().replace("@", "")


def normalize_company_key(text: str) -> str:
    text = str(text).strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text


def get_email_domain(email: str) -> str:
    email = str(email).strip().lower()
    if "@" not in email:
        return ""
    return email.split("@", 1)[1]


def find_user_index(users, email, sub):
    for i, row in enumerate(users):
        if row.get("email") == email and row.get("sub") == sub:
            return i
    return None


def find_company_index(companies, company_key):
    for i, row in enumerate(companies):
        if row.get("key") == company_key:
            return i
    return None


def find_company_by_key(companies, company_key):
    for row in companies:
        if row.get("key") == company_key:
            return row
    return None


def find_company_by_domain(companies, domain):
    domain = normalize_domain(domain)
    for row in companies:
        if normalize_domain(row.get("domain", "")) == domain:
            return row
    return None


def ensure_user_fields(user_row):
    if "trial_start" not in user_row:
        user_row["trial_start"] = now_iso()
    if "trial_end" not in user_row:
        user_row["trial_end"] = (now_utc() + timedelta(days=TRIAL_DAYS)).isoformat()
    if "billing_status" not in user_row:
        user_row["billing_status"] = "trialing"
    if "is_premium" not in user_row:
        user_row["is_premium"] = False
    if "active_sessions" not in user_row:
        user_row["active_sessions"] = []
    if "company_key" not in user_row:
        user_row["company_key"] = None
    if "company_name" not in user_row:
        user_row["company_name"] = None
    if "role" not in user_row:
        user_row["role"] = "member"
    if "max_active_sessions_override" not in user_row:
        user_row["max_active_sessions_override"] = None
    return user_row


def ensure_company_fields(company_row):
    if "billing_status" not in company_row:
        company_row["billing_status"] = "trialing"
    if "max_seats" not in company_row:
        company_row["max_seats"] = 0
    if "is_active" not in company_row:
        company_row["is_active"] = True
    if "shared_workspace_enabled" not in company_row:
        company_row["shared_workspace_enabled"] = False
    if "plan_start" not in company_row:
        company_row["plan_start"] = None
    if "plan_end" not in company_row:
        company_row["plan_end"] = None
    if "trial_start" not in company_row:
        company_row["trial_start"] = now_iso()
    if "trial_end" not in company_row:
        company_row["trial_end"] = (now_utc() + timedelta(days=7)).isoformat()
    if "stripe_customer_id" not in company_row:
        company_row["stripe_customer_id"] = None
    if "stripe_subscription_id" not in company_row:
        company_row["stripe_subscription_id"] = None
    if "owner_email" not in company_row:
        company_row["owner_email"] = ""
    return company_row


def get_current_user_registry_row():
    user = {
        "email": get_current_user_email(),
        "sub": get_current_user_id() if get_current_user_id() != "anonymous" else "",
    }
    users = load_users_registry()
    idx = find_user_index(users, user["email"], user["sub"])

    if idx is None:
        return None, None, users

    users[idx] = ensure_user_fields(users[idx])
    save_users_registry(users)
    return idx, users[idx], users


def get_current_user_company():
    idx, row, users = get_current_user_registry_row()
    if row is None:
        return None
    company_key = row.get("company_key")
    if not company_key:
        return None
    companies = load_companies_registry()
    company = find_company_by_key(companies, company_key)
    if company:
        company = ensure_company_fields(company)
    return company


def trial_days_left(trial_end_value):
    dt = parse_iso(trial_end_value)
    if dt is None:
        return 0
    remaining = dt - now_utc()
    if remaining.total_seconds() <= 0:
        return 0
    return max(1, remaining.days + (1 if remaining.seconds > 0 else 0))


def company_user_count(company_key):
    users = load_users_registry()
    count = 0
    for row in users:
        if row.get("company_key") == company_key and row.get("status") != "blocked":
            count += 1
    return count


def company_has_available_seat(company):
    if not company:
        return True
    max_seats = int(company.get("max_seats", 0) or 0)
    if max_seats <= 0:
        return True
    used = company_user_count(company["key"])
    return used < max_seats


def format_company_seats(company):
    if not company:
        return "-"
    used = company_user_count(company["key"])
    max_seats = int(company.get("max_seats", 0) or 0)
    return f"{used}/{max_seats}"


def is_admin_user():
    return get_current_user_email() in ADMIN_EMAILS


def current_user_status():
    if is_admin_user():
        return "approved"
    idx, row, users = get_current_user_registry_row()
    if row is None:
        return "pending"
    return row.get("status", "pending")


def current_user_is_blocked():
    return current_user_status() == "blocked"


def current_user_is_approved():
    return current_user_status() == "approved"


def online_status_from_last_seen(last_seen_value):
    dt = parse_iso(last_seen_value)
    if dt is None:
        return "Offline"
    if datetime.utcnow() - dt <= timedelta(minutes=2):
        return "Online"
    return "Offline"


def set_user_status(email, sub, new_status):
    users = load_users_registry()
    idx = find_user_index(users, email, sub)
    if idx is not None:
        users[idx]["status"] = new_status
        save_users_registry(users)


def set_user_premium(email, sub, is_premium=True):
    users = load_users_registry()
    idx = find_user_index(users, email, sub)
    if idx is not None:
        users[idx]["is_premium"] = bool(is_premium)
        users[idx]["billing_status"] = "active" if is_premium else "expired"
        if is_premium:
            users[idx]["status"] = "approved"
        save_users_registry(users)


def reset_user_to_trial(email, sub):
    users = load_users_registry()
    idx = find_user_index(users, email, sub)
    if idx is not None:
        users[idx]["is_premium"] = False
        users[idx]["billing_status"] = "trialing"
        users[idx]["trial_start"] = now_iso()
        users[idx]["trial_end"] = (now_utc() + timedelta(days=TRIAL_DAYS)).isoformat()
        users[idx]["status"] = "approved"
        save_users_registry(users)


def reset_user_sessions(email, sub):
    users = load_users_registry()
    idx = find_user_index(users, email, sub)
    if idx is not None:
        users[idx]["active_sessions"] = []
        save_users_registry(users)


def set_user_max_active_sessions_override(email, sub, max_sessions_override):
    users = load_users_registry()
    idx = find_user_index(users, email, sub)
    if idx is not None:
        if max_sessions_override is None or int(max_sessions_override) <= 0:
            users[idx]["max_active_sessions_override"] = None
        else:
            users[idx]["max_active_sessions_override"] = int(max_sessions_override)
        save_users_registry(users)


def get_user_max_active_sessions(user_row=None, email=None, sub=None):
    row = user_row
    if row is None and email and sub:
        users = load_users_registry()
        idx = find_user_index(users, email, sub)
        if idx is not None:
            row = ensure_user_fields(users[idx])

    if row is None:
        return MAX_ACTIVE_SESSIONS

    override = row.get("max_active_sessions_override")
    try:
        override_int = int(override)
        if override_int > 0:
            return override_int
    except Exception:
        pass

    return MAX_ACTIVE_SESSIONS


def remove_user_from_company(email, sub):
    users = load_users_registry()
    idx = find_user_index(users, email, sub)
    if idx is not None:
        users[idx]["company_key"] = None
        users[idx]["company_name"] = None
        users[idx]["role"] = "member"
        save_users_registry(users)


def get_current_session_id():
    if "app_session_id" not in st.session_state:
        st.session_state["app_session_id"] = str(uuid.uuid4())
    return st.session_state["app_session_id"]


def cleanup_stale_active_sessions(sessions):
    cleaned = []
    now_dt = now_utc()

    for s in sessions or []:
        last_seen_raw = s.get("last_seen")
        last_seen_dt = parse_iso(last_seen_raw)

        if last_seen_dt is None:
            continue

        if now_dt - last_seen_dt < timedelta(hours=STALE_SESSION_TIMEOUT_HOURS):
            cleaned.append(s)

    return cleaned


def register_current_session():
    idx, row, users = get_current_user_registry_row()
    if row is None:
        return True, 0

    current_session_id = get_current_session_id()
    force_replace = st.session_state.get("force_replace_session", False)
    sessions = cleanup_stale_active_sessions(row.get("active_sessions", []))
    allowed_sessions = get_user_max_active_sessions(row)

    for s in sessions:
        if s.get("session_id") == current_session_id:
            s["last_seen"] = now_iso()
            users[idx]["active_sessions"] = sessions
            save_users_registry(users)
            return True, len(sessions)

    if force_replace:
        sessions = [{
            "session_id": current_session_id,
            "last_seen": now_iso(),
            "started_at": now_iso(),
        }]
        users[idx]["active_sessions"] = sessions
        save_users_registry(users)
        st.session_state["force_replace_session"] = False
        return True, 1

    if len(sessions) >= allowed_sessions:
        users[idx]["active_sessions"] = sessions
        save_users_registry(users)
        return False, len(sessions)

    sessions.append({"session_id": current_session_id, "last_seen": now_iso(), "started_at": now_iso()})
    users[idx]["active_sessions"] = sessions
    save_users_registry(users)
    return True, len(sessions)


def unregister_current_session():
    idx, row, users = get_current_user_registry_row()
    if row is None:
        return
    current_session_id = get_current_session_id()
    sessions = row.get("active_sessions", [])
    sessions = [s for s in sessions if s.get("session_id") != current_session_id]
    users[idx]["active_sessions"] = sessions
    save_users_registry(users)


def replace_with_current_session_only():
    idx, row, users = get_current_user_registry_row()
    if row is None:
        return False
    current_session_id = get_current_session_id()
    users[idx]["active_sessions"] = [{
        "session_id": current_session_id,
        "last_seen": now_iso(),
        "started_at": now_iso(),
    }]
    save_users_registry(users)
    return True


def touch_current_user():
    idx, row, users = get_current_user_registry_row()
    if row is not None:
        users[idx]["last_seen"] = now_iso()
        save_users_registry(users)


def touch_current_session():
    idx, row, users = get_current_user_registry_row()
    if row is None:
        return
    current_session_id = get_current_session_id()
    original_sessions = row.get("active_sessions", [])
    sessions = cleanup_stale_active_sessions(original_sessions)
    changed = len(sessions) != len(original_sessions)

    for s in sessions:
        if s.get("session_id") == current_session_id:
            s["last_seen"] = now_iso()
            changed = True
            break

    if changed:
        users[idx]["active_sessions"] = sessions
        save_users_registry(users)


def logout_current_user():
    unregister_current_session()
    st.logout()


def ensure_current_user_in_registry():
    email = get_current_user_email()
    sub = get_current_user_id()
    name = get_current_user_name()

    users = load_users_registry()
    idx = find_user_index(users, email, sub)

    if idx is None:
        users.append(
            {
                "email": email,
                "sub": sub,
                "name": name,
                "status": "approved",
                "first_seen": now_iso(),
                "last_login": now_iso(),
                "last_seen": now_iso(),
                "trial_start": now_iso(),
                "trial_end": (now_utc() + timedelta(days=TRIAL_DAYS)).isoformat(),
                "billing_status": "trialing",
                "is_premium": False,
                "active_sessions": [],
                "max_active_sessions_override": None,
                "company_key": None,
                "company_name": None,
                "role": "member",
            }
        )
    else:
        users[idx]["name"] = name
        users[idx]["last_login"] = now_iso()
        users[idx]["last_seen"] = now_iso()
        if email in ADMIN_EMAILS:
            users[idx]["status"] = "approved"
        users[idx] = ensure_user_fields(users[idx])

    save_users_registry(users)


def sync_company_assignment_from_domain():
    idx, row, users = get_current_user_registry_row()
    if row is None:
        return {"status": "none", "company": None}

    email = row.get("email", "")
    domain = get_email_domain(email)
    if not domain:
        return {"status": "none", "company": None}

    companies = load_companies_registry()
    company = find_company_by_domain(companies, domain)
    if not company:
        return {"status": "none", "company": None}

    company = ensure_company_fields(company)
    if not company.get("is_active", True):
        return {"status": "inactive", "company": company}

    existing_company_key = row.get("company_key")
    if existing_company_key == company.get("key"):
        return {"status": "assigned", "company": company}

    if not company_has_available_seat(company):
        return {"status": "full", "company": company}

    users[idx]["company_key"] = company.get("key")
    users[idx]["company_name"] = company.get("name", company.get("key"))
    users[idx]["status"] = "approved"

    if company.get("owner_email") == email:
        users[idx]["role"] = "company_admin"

    save_users_registry(users)
    return {"status": "assigned", "company": company}


def sync_individual_status_from_stripe(email: str):
    idx, row, users = get_current_user_registry_row()
    if row is None or not email:
        return False

    stripe_row = get_stripe_subscription_row(email)
    if not stripe_row:
        return False

    if stripe_row.get("billing_status") == "active":
        users[idx]["billing_status"] = "active"
        users[idx]["is_premium"] = True
        save_users_registry(users)
        return True

    if row.get("billing_status") == "active":
        users[idx]["billing_status"] = "expired"
        users[idx]["is_premium"] = False
        save_users_registry(users)

    return False


def get_company_plan_window(company):
    if not company:
        return None, None

    plan_start = company.get("plan_start")
    plan_end = company.get("plan_end")
    start_dt = None
    end_dt = None

    if plan_start:
        try:
            start_dt = datetime.combine(date.fromisoformat(str(plan_start)[:10]), datetime.min.time())
        except Exception:
            start_dt = parse_iso(str(plan_start))

    if plan_end:
        try:
            end_dt = datetime.combine(date.fromisoformat(str(plan_end)[:10]), datetime.max.time())
        except Exception:
            end_dt = parse_iso(str(plan_end))

    return start_dt, end_dt


def get_company_plan_days_left(company):
    _, end_dt = get_company_plan_window(company)
    if end_dt is None:
        return None
    remaining = end_dt - now_utc()
    if remaining.total_seconds() <= 0:
        return 0
    return max(1, remaining.days + (1 if remaining.seconds > 0 else 0))


def company_has_access(company):
    if not company:
        return False

    if not company.get("is_active", True):
        return False

    plan_start_dt, plan_end_dt = get_company_plan_window(company)
    if plan_start_dt or plan_end_dt:
        now_dt = now_utc()
        if plan_start_dt and now_dt < plan_start_dt:
            return False
        if plan_end_dt and now_dt > plan_end_dt:
            return False
        return True

    if company.get("billing_status") == "active":
        return True

    trial_end = parse_iso(company.get("trial_end", ""))
    if trial_end and now_utc() <= trial_end:
        return True

    return False


def current_user_has_access():
    idx, row, users = get_current_user_registry_row()
    if row is None:
        return False

    company = get_current_user_company()
    if company and company_has_access(company):
        return True

    if row.get("billing_status") == "active" or row.get("is_premium") is True:
        return True

    trial_end = parse_iso(row.get("trial_end", ""))
    if trial_end and now_utc() <= trial_end:
        return True

    user_email = get_current_user_email()
    if user_email and sync_individual_status_from_stripe(user_email):
        return True

    if idx is not None:
        users[idx]["billing_status"] = "expired"
        users[idx]["is_premium"] = False
        save_users_registry(users)

    return False


def upsert_company(
    company_key,
    name,
    domain,
    max_seats=0,
    is_active=True,
    billing_status="trialing",
    owner_email="",
    shared_workspace_enabled=False,
    plan_start=None,
    plan_end=None,
):
    companies = load_companies_registry()
    normalized_key = normalize_company_key(company_key)
    idx = find_company_index(companies, normalized_key)

    payload = {
        "key": normalized_key,
        "name": str(name).strip(),
        "domain": normalize_domain(domain),
        "max_seats": int(max_seats),
        "is_active": bool(is_active),
        "billing_status": billing_status,
        "trial_start": now_iso(),
        "trial_end": (now_utc() + timedelta(days=7)).isoformat(),
        "stripe_customer_id": None,
        "stripe_subscription_id": None,
        "owner_email": owner_email.strip().lower(),
        "shared_workspace_enabled": bool(shared_workspace_enabled),
        "plan_start": plan_start.isoformat() if hasattr(plan_start, "isoformat") else (str(plan_start) if plan_start else None),
        "plan_end": plan_end.isoformat() if hasattr(plan_end, "isoformat") else (str(plan_end) if plan_end else None),
        "updated_at": now_iso(),
    }

    if idx is None:
        payload["created_at"] = now_iso()
        companies.append(payload)
    else:
        existing = companies[idx]
        payload["created_at"] = existing.get("created_at", now_iso())
        payload["stripe_customer_id"] = existing.get("stripe_customer_id")
        payload["stripe_subscription_id"] = existing.get("stripe_subscription_id")
        payload["trial_start"] = existing.get("trial_start", payload["trial_start"])
        payload["trial_end"] = existing.get("trial_end", payload["trial_end"])
        if not owner_email:
            payload["owner_email"] = existing.get("owner_email", "")
        payload["shared_workspace_enabled"] = bool(shared_workspace_enabled)
        if plan_start in [None, ""]:
            payload["plan_start"] = existing.get("plan_start")
        if plan_end in [None, ""]:
            payload["plan_end"] = existing.get("plan_end")
        companies[idx] = payload

    save_companies_registry(companies)


def remove_company(company_key):
    companies = load_companies_registry()
    companies = [c for c in companies if c.get("key") != company_key]
    save_companies_registry(companies)


# -------------------------------------------------
# COMPANY / USER FILES
# -------------------------------------------------
def get_storage_slug_for_current_user():
    company = get_current_user_company()
    if company:
        slug = normalize_company_key(company.get("key", "workspace"))
    else:
        slug = (
            get_current_user_id()
            .replace("@", "_")
            .replace(".", "_")
            .replace("/", "_")
            .replace("\\", "_")
        )
    return slug


WORKSPACE_SLUG = get_storage_slug_for_current_user()
WORKSPACE_DIR = PERSIST_ROOT / WORKSPACE_SLUG
UPLOADS_DIR = WORKSPACE_DIR / "uploads"
COMPANIES_FILE = WORKSPACE_DIR / "companies.csv"
COMPARISONS_DIR = WORKSPACE_DIR / "_saved_comparisons"

WORKSPACE_DIR.mkdir(parents=True, exist_ok=True)
UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
COMPARISONS_DIR.mkdir(parents=True, exist_ok=True)

TRASH_DIR = WORKSPACE_DIR / "_trash"
TRASH_SOURCES_DIR = TRASH_DIR / "sources"
TRASH_COMPARISONS_DIR = TRASH_DIR / "comparisons"
TRASH_SOURCES_DIR.mkdir(parents=True, exist_ok=True)
TRASH_COMPARISONS_DIR.mkdir(parents=True, exist_ok=True)


def get_current_user_comparisons_file():
    company = get_current_user_company()
    if company and bool(company.get("shared_workspace_enabled", False)):
        company_key = normalize_company_key(company.get("key", "workspace"))
        return COMPARISONS_DIR / f"company_{company_key}.json"

    raw_user = get_current_user_id() or get_current_user_email() or "anonymous"
    safe_user = normalize_company_key(raw_user)
    return COMPARISONS_DIR / f"{safe_user}.json"

COMPARISON_LOCK_STALE_MINUTES = 120


def get_comparison_lock_file(comparison_file: Path) -> Path:
    comparison_file = Path(comparison_file)
    return comparison_file.with_name(f"{comparison_file.stem}__locks.json")


def load_comparison_locks(comparison_file: Path) -> dict:
    lock_file = get_comparison_lock_file(comparison_file)
    if not lock_file.exists():
        return {}
    try:
        data = json.loads(lock_file.read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def save_comparison_locks(comparison_file: Path, data: dict):
    lock_file = get_comparison_lock_file(comparison_file)
    lock_file.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def _comparison_lock_is_stale(lock_payload: dict) -> bool:
    if not isinstance(lock_payload, dict):
        return True
    last_seen = parse_iso(lock_payload.get("last_seen", ""))
    if last_seen is None:
        return True
    return (now_utc() - last_seen) > timedelta(minutes=COMPARISON_LOCK_STALE_MINUTES)


def get_comparison_lock_info(comparison_file: Path, comparison_id: str):
    locks = load_comparison_locks(comparison_file)
    lock_payload = locks.get(str(comparison_id))
    if not lock_payload:
        return None
    if _comparison_lock_is_stale(lock_payload):
        locks.pop(str(comparison_id), None)
        save_comparison_locks(comparison_file, locks)
        return None
    return lock_payload


def comparison_lock_owned_by_current_session(lock_payload: dict) -> bool:
    if not lock_payload:
        return False

    lock_email = str(lock_payload.get("owner_email", "") or "").strip().lower()
    current_email = str(get_current_user_email() or "").strip().lower()

    if lock_email and current_email and lock_email == current_email:
        return True

    return str(lock_payload.get("session_id", "")) == str(get_current_session_id())


def acquire_comparison_lock(comparison_file: Path, comparison_id: str, comparison_name: str = ""):
    comparison_id = str(comparison_id or "").strip()
    if not comparison_id:
        return False, "Invalid comparison."

    locks = load_comparison_locks(comparison_file)
    existing = locks.get(comparison_id)

    if existing and _comparison_lock_is_stale(existing):
        locks.pop(comparison_id, None)
        existing = None

    if existing and not comparison_lock_owned_by_current_session(existing):
        holder = existing.get("owner_name") or existing.get("owner_email") or "another user"
        return False, f"This comparison is currently being edited by {holder}."

    lock_payload = {
        "comparison_id": comparison_id,
        "comparison_name": comparison_name or "",
        "owner_email": get_current_user_email(),
        "owner_name": get_current_user_name(),
        "session_id": get_current_session_id(),
        "locked_at": existing.get("locked_at", now_iso()) if existing else now_iso(),
        "last_seen": now_iso(),
    }
    locks[comparison_id] = lock_payload
    save_comparison_locks(comparison_file, locks)
    st.session_state["locked_comparison_id"] = comparison_id
    st.session_state["locked_comparison_file"] = str(comparison_file)
    return True, ""


def touch_current_comparison_lock():
    comparison_id = str(st.session_state.get("locked_comparison_id", "") or "").strip()
    comparison_file_str = str(st.session_state.get("locked_comparison_file", "") or "").strip()
    if not comparison_id or not comparison_file_str:
        return

    comparison_file = Path(comparison_file_str)
    locks = load_comparison_locks(comparison_file)
    existing = locks.get(comparison_id)
    if not existing or not comparison_lock_owned_by_current_session(existing):
        return

    existing["last_seen"] = now_iso()
    locks[comparison_id] = existing
    save_comparison_locks(comparison_file, locks)


def release_comparison_lock(comparison_file: Path | None = None, comparison_id: str | None = None):
    cid = str(comparison_id or st.session_state.get("locked_comparison_id", "") or "").strip()
    cfile = str(comparison_file or st.session_state.get("locked_comparison_file", "") or "").strip()
    if not cid or not cfile:
        st.session_state["locked_comparison_id"] = None
        st.session_state["locked_comparison_file"] = None
        return

    lock_file_target = Path(cfile)
    locks = load_comparison_locks(lock_file_target)
    existing = locks.get(cid)
    if existing and comparison_lock_owned_by_current_session(existing):
        locks.pop(cid, None)
        save_comparison_locks(lock_file_target, locks)

    st.session_state["locked_comparison_id"] = None
    st.session_state["locked_comparison_file"] = None




def duplicate_saved_comparison(comparison_file: Path, comparison_id: str):
    try:
        comparison_file = Path(comparison_file)
        records = list_comparisons(comparison_file)

        for rec in records:
            if str(rec.get("id", "")) == str(comparison_id):
                new_rec = json.loads(json.dumps(rec, ensure_ascii=False, default=str))
                new_rec["id"] = str(uuid.uuid4())
                original_name = str(rec.get("name", "") or "Comparison").strip()
                new_rec["name"] = f"Copy of {original_name}"
                new_rec["created_at"] = now_iso()
                new_rec["updated_at"] = now_iso()

                all_records = list(records)
                all_records.append(new_rec)
                comparison_file.write_text(
                    json.dumps(all_records, ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )
                return True, "Comparison duplicated.", new_rec

        return False, "Comparison not found.", None
    except Exception as e:
        return False, f"Could not duplicate comparison: {e}", None


def duplicate_saved_comparison_with_name(comparison_file: Path, comparison_id: str, new_name: str):
    try:
        comparison_file = Path(comparison_file)
        target_name = str(new_name or "").strip()
        if not target_name:
            return False, "Please enter a name.", None

        records = list_comparisons(comparison_file)

        for rec in records:
            if str(rec.get("id", "")) == str(comparison_id):
                new_rec = json.loads(json.dumps(rec, ensure_ascii=False, default=str))
                new_rec["id"] = str(uuid.uuid4())
                new_rec["name"] = target_name
                new_rec["created_at"] = now_iso()
                new_rec["updated_at"] = now_iso()

                all_records = list(records)
                all_records.append(new_rec)
                comparison_file.write_text(
                    json.dumps(all_records, ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )
                return True, "Comparison saved as new copy.", new_rec

        return False, "Comparison not found.", None
    except Exception as e:
        return False, f"Could not save as new comparison: {e}", None


def get_all_active_comparison_locks():
    rows = []
    for lock_file in sorted(COMPARISONS_DIR.glob("*__locks.json")):
        try:
            lock_map = json.loads(lock_file.read_text(encoding="utf-8"))
        except Exception:
            continue
        if not isinstance(lock_map, dict):
            continue
        changed = False
        for comparison_id, lock_payload in list(lock_map.items()):
            if _comparison_lock_is_stale(lock_payload):
                lock_map.pop(comparison_id, None)
                changed = True
                continue
            rows.append({
                "Comparison ID": comparison_id,
                "Comparison": lock_payload.get("comparison_name", ""),
                "Locked By": lock_payload.get("owner_name") or lock_payload.get("owner_email", ""),
                "Email": lock_payload.get("owner_email", ""),
                "Locked At": lock_payload.get("locked_at", ""),
                "Last Seen": lock_payload.get("last_seen", ""),
                "Lock File": lock_file.name,
            })
        if changed:
            lock_file.write_text(json.dumps(lock_map, ensure_ascii=False, indent=2), encoding="utf-8")
    return rows




# -------------------------------------------------
# SMART PRODUCT MATCHING (BETA)
# -------------------------------------------------
MATCH_HISTORY_FILE = ADMIN_DIR / "match_history.json"

def _normalize_match_text(text):
    text = str(text or "").lower().strip()
    text = re.sub(r"[^a-z0-9α-ωάέήίόύώϊϋΐΰ\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text

def _extract_match_mm(text):
    text = str(text or "").lower().replace(",", ".")
    m = re.search(r"(\d+(?:\.\d+)?)\s*mm", text)
    if not m:
        return None
    try:
        return float(m.group(1))
    except Exception:
        return None

def _extract_match_dimensions(text):
    text = str(text or "").lower().replace("×", "x")
    m = re.search(r"(\d{3,4})\s*x\s*(\d{3,4})", text)
    if not m:
        return None
    try:
        return (int(m.group(1)), int(m.group(2)))
    except Exception:
        return None

def _strip_dimension_tails(text):
    text = str(text or "").strip()
    text = text.replace(",", ".")
    text = re.sub(r"\b\d{2,4}\s*[x×]\s*\d{2,4}(?:\s*[x×]\s*\d{2,4})?\s*mm\b", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\b\d{2,4}\s*[x×]\s*\d{2,4}(?:\s*[x×]\s*\d{2,4})?\b", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+", " ", text).strip()
    return text

def _extract_core_product_name(text):
    text = str(text or "").replace(",", ".")
    text = _strip_dimension_tails(text)
    text = re.sub(r"\b(?:mm|m2|m²|kg|gr|g|lt|l|ml|cm|tmx|τεμ|τεμ\.|pcs|pc)\b", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+", " ", text).strip()
    return text

def _core_history_keys_for_product(product_text):
    full_key = _normalize_match_text(product_text)
    core_key = _normalize_match_text(_extract_core_product_name(product_text))
    keys = []
    if full_key:
        keys.append(full_key)
    if core_key and core_key not in keys:
        keys.append(core_key)
    return keys

def _core_variant_keys(core_text: str):
    core_text = _normalize_match_text(core_text)
    if not core_text:
        return []
    parts = core_text.split()
    variants = [core_text]

    # prefix variants to let "nida flam plus" inherit from "nida flam"
    if len(parts) >= 2:
        variants.append(" ".join(parts[:2]))
    if len(parts) >= 3:
        variants.append(" ".join(parts[:3]))

    # remove common commercial suffixes
    removable = {"plus", "pro", "ultra", "max", "smart", "expert", "premium", "basic"}
    trimmed_parts = list(parts)
    while len(trimmed_parts) > 1 and trimmed_parts[-1] in removable:
        trimmed_parts = trimmed_parts[:-1]
        variants.append(" ".join(trimmed_parts))

    out = []
    for v in variants:
        v = _normalize_match_text(v)
        if v and v not in out:
            out.append(v)
    return out

def _infer_product_family(product_text: str, category_text: str = "", mm_text: str = ""):
    text = _normalize_match_text(f"{product_text} {category_text}")
    mm_norm = _normalize_match_text(mm_text)

    board_tokens = ["γυψο", "γυψοσαν", "gypsum", "plasterboard", "drywall", "habito", "vidiwall", "massivbauplatte"]
    profile_tokens = ["profil", "profile", "προφιλ", "ορθοστατ", "στρωτηρ", "uw", "cw", "ud", "cd"]
    waterproof_tokens = ["aquamat", "sikaelastic", "mapelastic", "στεγαν", "waterproof", "hydro", "aqua"]
    adhesive_tokens = ["adhes", "glue", "κολλα", "tilefix", "fix", "τσιμεντοκολλα"]
    insulation_tokens = ["insulation", "μονω", "xps", "eps", "πετροβαμβ", "ορυκτοβαμβ"]
    accessory_tokens = ["ντιζ", "anker", "anchor", "βιδ", "screw", "washer", "γωνιοκρανο", "tape", "joint"]

    def has_any(tokens):
        return any(tok in text for tok in tokens)

    if has_any(profile_tokens):
        return "profile"
    if has_any(accessory_tokens):
        return "accessory"
    if has_any(board_tokens):
        return "board"
    if has_any(waterproof_tokens):
        return "waterproofing"
    if has_any(adhesive_tokens):
        return "adhesive"
    if has_any(insulation_tokens):
        return "insulation"

    if mm_norm in {"m", "meter", "μέτρο"}:
        return "profile"
    if mm_norm in {"τεμαχιο", "τεμαχια", "piece", "pcs", "pc"} and has_any(accessory_tokens):
        return "accessory"
    if mm_norm in {"m2", "m²"} and has_any(board_tokens):
        return "board"

    return "unknown"

def _normalize_mm_unit(mm_text: str):
    t = _normalize_match_text(mm_text)
    if not t:
        return ""
    if t in {"m2", "m²", "sqm", "sq m", "τετρ μετρο", "τετραγωνικο μετρο"}:
        return "m2"
    if t in {"m", "meter", "metre", "μετρο", "μέτρο"}:
        return "m"
    if t in {"tmx", "temaxio", "temachio", "temaxia", "temachia", "τεμ", "τεμ.", "τεμαχιο", "τεμαχια", "piece", "pieces", "pc", "pcs"}:
        return "piece"
    return t

def _mm_units_compatible(source_mm: str, target_mm: str):
    a = _normalize_mm_unit(source_mm)
    b = _normalize_mm_unit(target_mm)
    if not a or not b:
        return True
    return a == b

def _products_are_compatible(source_product: str, target_product: str, source_category: str = "", target_category: str = "", source_mm: str = "", target_mm: str = ""):
    if not _mm_units_compatible(source_mm, target_mm):
        return False

    fam_a = _infer_product_family(source_product, source_category, source_mm)
    fam_b = _infer_product_family(target_product, target_category, target_mm)

    if fam_a != "unknown" and fam_b != "unknown" and fam_a != fam_b:
        return False

    # Extra hard stop for board/profile accidental mixes even when one side is unknown
    text_a = _normalize_match_text(f"{source_product} {source_category}")
    text_b = _normalize_match_text(f"{target_product} {target_category}")
    board_markers = ["γυψο", "γυψοσαν", "gypsum", "plasterboard", "drywall", "habito", "vidiwall", "massivbauplatte"]
    profile_markers = ["profil", "profile", "προφιλ", "uw", "cw", "ud", "cd", "ορθοστατ", "στρωτηρ"]

    a_board = any(t in text_a for t in board_markers)
    b_board = any(t in text_b for t in board_markers)
    a_profile = any(t in text_a for t in profile_markers)
    b_profile = any(t in text_b for t in profile_markers)

    if (a_board and b_profile) or (a_profile and b_board):
        return False

    return True

def _history_hits_for_candidate(history, user_email, source_product, target_company, target_product):
    source_keys = _core_history_keys_for_product(source_product)
    target_keys = _core_history_keys_for_product(target_product)
    personal_hits = 0
    global_hits = 0

    for s_key in source_keys:
        for t_key in target_keys:
            personal_hits = max(
                personal_hits,
                int(
                    history.get("personal", {})
                    .get(str(user_email or "").strip().lower(), {})
                    .get(s_key, {})
                    .get(target_company, {})
                    .get(t_key, 0)
                ),
            )
            global_hits = max(
                global_hits,
                int(
                    history.get("global", {})
                    .get(s_key, {})
                    .get(target_company, {})
                    .get(t_key, 0)
                ),
            )

    return personal_hits, global_hits

def _learned_core_hits_for_candidate(history, source_product, target_company, target_product):
    source_core = _normalize_match_text(_extract_core_product_name(source_product))
    target_core = _normalize_match_text(_extract_core_product_name(target_product))
    if not source_core or not target_core:
        return 0

    company_key = str(target_company or "").strip().upper()
    best_hits = 0
    source_variants = _core_variant_keys(source_core)
    target_variants = _core_variant_keys(target_core)

    for s_var in source_variants:
        for t_var in target_variants:
            best_hits = max(
                best_hits,
                int(
                    history.get("learned_core", {})
                    .get(s_var, {})
                    .get(company_key, {})
                    .get(t_var, 0)
                ),
            )

    return best_hits

def _text_similarity(a, b):
    return SequenceMatcher(None, _normalize_match_text(a), _normalize_match_text(b)).ratio()

def _load_match_history():
    if not MATCH_HISTORY_FILE.exists():
        return {"personal": {}, "global": {}, "learned_core": {}}
    try:
        data = json.loads(MATCH_HISTORY_FILE.read_text(encoding="utf-8"))
        if isinstance(data, dict):
            data.setdefault("personal", {})
            data.setdefault("global", {})
            data.setdefault("learned_core", {})
            return data
    except Exception:
        pass
    return {"personal": {}, "global": {}, "learned_core": {}}

def _save_match_history(data):
    MATCH_HISTORY_FILE.write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

def _record_match_pair_to_history(
    history: dict,
    user_email: str,
    source_product: str,
    target_company: str,
    target_product: str,
    source_category: str = "",
    target_category: str = "",
    source_mm: str = "",
    target_mm: str = "",
):
    user_email = str(user_email or "").strip().lower()
    target_company = str(target_company or "").strip().upper()

    source_keys = _core_history_keys_for_product(source_product)
    target_keys = _core_history_keys_for_product(target_product)

    if not user_email or not source_keys or not target_keys or not target_company:
        return

    if not _products_are_compatible(source_product, target_product, source_category, target_category, source_mm, target_mm):
        return

    for source_key in source_keys:
        personal_targets = (
            history.setdefault("personal", {})
            .setdefault(user_email, {})
            .setdefault(source_key, {})
            .setdefault(target_company, {})
        )
        global_targets = (
            history.setdefault("global", {})
            .setdefault(source_key, {})
            .setdefault(target_company, {})
        )

        for target_key in target_keys:
            personal_targets[target_key] = int(personal_targets.get(target_key, 0)) + 1
            global_targets[target_key] = int(global_targets.get(target_key, 0)) + 1

    source_core = _normalize_match_text(_extract_core_product_name(source_product))
    target_core = _normalize_match_text(_extract_core_product_name(target_product))
    if source_core and target_core:
        for s_var in _core_variant_keys(source_core):
            learned_targets = (
                history.setdefault("learned_core", {})
                .setdefault(s_var, {})
                .setdefault(target_company, {})
            )
            for t_var in _core_variant_keys(target_core):
                learned_targets[t_var] = int(learned_targets.get(t_var, 0)) + 1

def _backfill_match_history_from_saved_comparisons():
    if st.session_state.get("_smart_match_backfill_done", False):
        return

    history = _load_match_history()
    touched = False
    display_to_code = {
        f"{row['name']} ({row['code']})": row["code"]
        for _, row in companies_df.iterrows()
    }

    for comparison_file in sorted(COMPARISONS_DIR.glob("*.json")):
        try:
            records = list_comparisons(comparison_file)
        except Exception:
            continue

        for record in records:
            state = record.get("state", {}) or {}
            row_ids = state.get("row_ids", []) or []
            if not isinstance(row_ids, list) or not row_ids:
                continue

            selected_codes = []
            for display in state.get("comparison_company_selection", []) or []:
                if display in display_to_code:
                    selected_codes.append(display_to_code[display])

            if len(selected_codes) < 2:
                continue

            owner_email = str(record.get("owner_email", "") or "").strip().lower() or get_current_user_email()

            for row_id in row_ids:
                row_products = []
                for code in selected_codes:
                    display_value = str(state.get(f"row_{row_id}_{code}_product", "") or "").strip()
                    if display_value:
                        row_products.append((code, display_value.split("| SAP")[0].strip()))

                if len(row_products) < 2:
                    continue

                for source_code, source_product in row_products:
                    for target_code, target_product in row_products:
                        if source_code == target_code:
                            continue
                        _record_match_pair_to_history(
                            history,
                            owner_email,
                            source_product,
                            target_code,
                            target_product,
                        )
                        touched = True

    if touched:
        _save_match_history(history)

    st.session_state["_smart_match_backfill_done"] = True

def rebuild_match_history_from_scratch():
    history = {"personal": {}, "global": {}, "learned_core": {}}
    display_to_code = {
        f"{row['name']} ({row['code']})": row["code"]
        for _, row in companies_df.iterrows()
    }

    for comparison_file in sorted(COMPARISONS_DIR.glob("*.json")):
        try:
            records = list_comparisons(comparison_file)
        except Exception:
            continue

        for record in records:
            state = record.get("state", {}) or {}
            row_ids = state.get("row_ids", []) or []
            if not isinstance(row_ids, list) or not row_ids:
                continue

            selected_codes = []
            for display in state.get("comparison_company_selection", []) or []:
                if display in display_to_code:
                    selected_codes.append(display_to_code[display])

            if len(selected_codes) < 2:
                continue

            owner_email = str(record.get("owner_email", "") or "").strip().lower() or get_current_user_email()

            for row_id in row_ids:
                row_products = []
                for code in selected_codes:
                    display_value = str(state.get(f"row_{row_id}_{code}_product", "") or "").strip()
                    if display_value:
                        row_products.append((code, display_value.split("| SAP")[0].strip()))

                if len(row_products) < 2:
                    continue

                for source_code, source_product in row_products:
                    for target_code, target_product in row_products:
                        if source_code == target_code:
                            continue
                        _record_match_pair_to_history(
                            history,
                            owner_email,
                            source_product,
                            target_code,
                            target_product,
                        )

    _save_match_history(history)
    st.session_state["_smart_match_backfill_done"] = True


def _infer_board_functional_family(product_text: str, category_text: str = ""):
    text = _normalize_match_text(f"{product_text} {category_text}")

    is_fire = any(tok in text for tok in [
        "flam", "fire", "πυραντ", " df ", " dfh", "rf", "type f", "f1"
    ])
    is_moisture = any(tok in text for tok in [
        "hydro", "h2", "ανθυγρ", "moist", "aqua"
    ])
    is_acoustic = any(tok in text for tok in [
        "acoustic", "sound", "phon", "silent", "ηχο"
    ])

    tags = []
    if is_fire:
        tags.append("fire")
    if is_moisture:
        tags.append("moisture")
    if is_acoustic:
        tags.append("acoustic")

    if not tags:
        return "standard"

    tags.sort()
    return "+".join(tags)

def _score_product_match_history_aware(user_email: str, source_row: dict, target_row: dict, target_company: str) -> float:
    product_a = str(source_row.get("Product", "") or "")
    product_b = str(target_row.get("Product", "") or "")
    category_a = str(source_row.get("Category", "") or "")
    category_b = str(target_row.get("Category", "") or "")
    mm_a = str(source_row.get("MM", "") or "")
    mm_b = str(target_row.get("MM", "") or "")
    package_a = str(source_row.get("Package", "") or "")
    package_b = str(target_row.get("Package", "") or "")

    history = _load_match_history()
    compatible_for_history = _products_are_compatible(
        product_a, product_b, category_a, category_b, mm_a, mm_b
    )

    if compatible_for_history:
        personal_hits, global_hits = _history_hits_for_candidate(
            history=history,
            user_email=user_email,
            source_product=product_a,
            target_company=target_company,
            target_product=product_b,
        )
        learned_core_hits = _learned_core_hits_for_candidate(
            history=history,
            source_product=product_a,
            target_company=target_company,
            target_product=product_b,
        )
    else:
        personal_hits, global_hits, learned_core_hits = 0, 0, 0

    # HISTORY FIRST, but still combined with names + characteristics
    personal_history_score = min(personal_hits * 30.0, 60.0)
    global_history_score = min(global_hits * 15.0, 50.0)
    learned_core_score = min(learned_core_hits * 18.0, 55.0)

    full_name_score = _text_similarity(product_a, product_b) * 6.0
    core_name_score = _text_similarity(_extract_core_product_name(product_a), _extract_core_product_name(product_b)) * 14.0
    name_score = full_name_score + core_name_score

    dim_score = 0.0
    dim_a = _extract_match_dimensions(product_a)
    dim_b = _extract_match_dimensions(product_b)
    if dim_a and dim_b:
        if dim_a == dim_b:
            dim_score += 6.0
        elif sorted(dim_a) == sorted(dim_b):
            dim_score += 3.0

    thickness_score = 0.0
    thick_a = _extract_match_mm(product_a + " " + mm_a)
    thick_b = _extract_match_mm(product_b + " " + mm_b)
    if thick_a is not None and thick_b is not None:
        if abs(thick_a - thick_b) < 0.01:
            thickness_score += 8.0
        elif abs(thick_a - thick_b) <= 0.5:
            thickness_score += 3.0
        else:
            thickness_score -= 5.0

    category_score = _text_similarity(category_a, category_b) * 6.0

    mm_score = 0.0
    if str(mm_a).strip() and str(mm_b).strip():
        if str(mm_a).strip().lower() == str(mm_b).strip().lower():
            mm_score += 3.0

    # Package remains very low priority
    package_score = _text_similarity(package_a, package_b) * 1.5

    # General rule: dimensions and packaging are weak signals, core name is stronger.
    dim_score *= 0.35
    package_score *= 0.8

    # Special strengthening for gypsum/plasterboards
    product_context = f"{category_a} {category_b} {product_a} {product_b}".lower()
    is_gypsum_board = any(token in product_context for token in [
        "gypsum", "plasterboard", "drywall", "board", "γυψο", "γυψοσαν"
    ])

    functional_family_score = 0.0
    if is_gypsum_board:
        # For gypsum boards, dimensions should not affect matching.
        dim_score = 0.0
        thickness_score *= 2.4
        category_score *= 1.4
        full_name_score *= 0.6
        core_name_score *= 2.4
        name_score = full_name_score + core_name_score

        fam_a = _infer_board_functional_family(product_a, category_a)
        fam_b = _infer_board_functional_family(product_b, category_b)

        if fam_a == fam_b:
            functional_family_score += 30.0
        elif fam_a == "standard" or fam_b == "standard":
            functional_family_score -= 8.0
        else:
            functional_family_score -= 28.0

        # If the board functional family clearly disagrees, history should not dominate.
        if fam_a != fam_b and fam_a != "standard" and fam_b != "standard":
            personal_history_score *= 0.2
            global_history_score *= 0.2
            learned_core_score *= 0.2

    score = (
        personal_history_score
        + global_history_score
        + learned_core_score
        + name_score
        + dim_score
        + thickness_score
        + category_score
        + mm_score
        + package_score
        + functional_family_score
    )

    return round(score, 2)


def _generate_smart_product_suggestion(user_email: str, source_row: dict, target_df: pd.DataFrame, target_company: str):
    if source_row is None or target_df is None or target_df.empty:
        return None, 0.0, "", ""

    source_engine_row = {
        "Product": str(source_row.get("Product", "") or ""),
        "Category": str(source_row.get("Category", "") or ""),
        "MM": str(source_row.get("MM", "") or ""),
    }

    target_rows = []
    for _, row in target_df.iterrows():
        row_dict = row.to_dict()
        target_rows.append({
            "Product": str(row_dict.get("Product", "") or ""),
            "Category": str(row_dict.get("Category", "") or ""),
            "MM": str(row_dict.get("MM", "") or ""),
            "_full_row": row_dict,
        })

    match, score, mode, confidence = suggest_product(CENTRAL_ENGINE, source_engine_row, target_company, target_rows)
    if match is None:
        return None, 0.0, "", ""

    full_row = match.get("_full_row") if isinstance(match, dict) else None
    if isinstance(full_row, dict):
        return full_row, float(score), str(mode or ""), str(confidence or "")
    return None, 0.0, "", ""

def _record_match_history_once_per_session(user_email: str, source_product: str, target_company: str, target_product: str):
    user_email = str(user_email or "").strip().lower()
    source_key = _normalize_match_text(source_product)
    target_key = _normalize_match_text(target_product)
    target_company = str(target_company or "").strip().upper()

    if not user_email or not source_key or not target_key or not target_company:
        return

    event_sig = f"{user_email}|{source_key}|{target_company}|{target_key}"
    seen = st.session_state.setdefault("_smart_match_recorded_events", set())
    if event_sig in seen:
        return

    history = _load_match_history()

    personal_targets = (
        history.setdefault("personal", {})
        .setdefault(user_email, {})
        .setdefault(source_key, {})
        .setdefault(target_company, {})
    )
    personal_targets[target_key] = int(personal_targets.get(target_key, 0)) + 1

    global_targets = (
        history.setdefault("global", {})
        .setdefault(source_key, {})
        .setdefault(target_company, {})
    )
    global_targets[target_key] = int(global_targets.get(target_key, 0)) + 1

    _save_match_history(history)
    seen.add(event_sig)

def _apply_smart_product_suggestions_for_row(row_id: int, selected_codes: list, catalogs: dict):
    if not st.session_state.get("smart_matching_enabled", False):
        return

    if len(selected_codes) < 2:
        return

    source_code = selected_codes[0]
    source_display = str(st.session_state.get(f"row_{row_id}_{source_code}_product", "") or "").strip()
    if not source_display:
        return

    source_row = get_catalog_row(catalogs.get(source_code), source_display)
    if source_row is None:
        return

    user_email = get_current_user_email()
    suggestion_notes = st.session_state.setdefault("smart_match_notes", {})
    score_notes = st.session_state.setdefault("smart_match_scores", {})
    confidence_notes = st.session_state.setdefault("smart_match_confidence", {})
    mode_notes = st.session_state.setdefault("smart_match_mode", {})
    suggestion_targets = st.session_state.setdefault("smart_match_target_display", {})
    auto_source_map = st.session_state.setdefault("smart_match_auto_source_display", {})
    auto_target_map = st.session_state.setdefault("smart_match_auto_target_display", {})

    for target_code in selected_codes[1:]:
        note_key = f"{row_id}|{target_code}"
        previous_auto_source = auto_source_map.get(note_key, "")
        previous_auto_target = auto_target_map.get(note_key, "")
        previous_suggested_display = suggestion_targets.get(note_key, "")

        suggestion_notes.pop(note_key, None)
        score_notes.pop(note_key, None)
        confidence_notes.pop(note_key, None)
        mode_notes.pop(note_key, None)
        suggestion_targets.pop(note_key, None)

        target_key = f"row_{row_id}_{target_code}_product"
        widget_key = get_product_widget_key(row_id, target_code)
        existing_target = str(st.session_state.get(target_key, "") or "").strip()

        best_row, best_score, best_mode, best_confidence = _generate_smart_product_suggestion(
            user_email=user_email,
            source_row=source_row.to_dict() if hasattr(source_row, "to_dict") else dict(source_row),
            target_df=catalogs.get(target_code),
            target_company=target_code,
        )

        if best_row is None or best_score < 2:
            # No valid learned match for this company: clear only previous auto-filled values,
            # but do not overwrite a user-selected manual value.
            if previous_auto_target and existing_target == previous_auto_target:
                st.session_state[target_key] = ""
                st.session_state[widget_key] = ""
                existing_target = ""
                auto_target_map.pop(note_key, None)
                auto_source_map.pop(note_key, None)

            if not existing_target:
                suggestion_notes[note_key] = "No match found"
                score_notes.pop(note_key, None)
                confidence_notes.pop(note_key, None)
                mode_notes.pop(note_key, None)
                suggestion_targets.pop(note_key, None)
            continue

        suggested_display = str(best_row.get("DISPLAY", "") or "").strip()
        if not suggested_display:
            if previous_auto_target and existing_target == previous_auto_target:
                st.session_state[target_key] = ""
                st.session_state[widget_key] = ""
                existing_target = ""
                auto_target_map.pop(note_key, None)
                auto_source_map.pop(note_key, None)

            if not existing_target:
                suggestion_notes[note_key] = "No match found"
                score_notes.pop(note_key, None)
                confidence_notes.pop(note_key, None)
                mode_notes.pop(note_key, None)
                suggestion_targets.pop(note_key, None)
            continue

        source_changed_from_previous_auto = previous_auto_source and previous_auto_source != source_display

        should_replace_existing = (
            not existing_target
            or (previous_auto_target and existing_target == previous_auto_target)
            or (
                source_changed_from_previous_auto
                and previous_suggested_display
                and existing_target == previous_suggested_display
            )
        )

        if should_replace_existing:
            st.session_state[target_key] = suggested_display
            st.session_state[widget_key] = suggested_display
            suggestion_notes[note_key] = "Suggested"
            score_notes[note_key] = best_score
            confidence_notes[note_key] = best_confidence or ("medium" if best_score >= 2 else "low")
            mode_notes[note_key] = best_mode or "table"
            suggestion_targets[note_key] = suggested_display
            auto_source_map[note_key] = source_display
            auto_target_map[note_key] = suggested_display
        else:
            if existing_target != suggested_display:
                suggestion_notes[note_key] = "Better match available"
                score_notes[note_key] = best_score
                confidence_notes[note_key] = best_confidence or ("medium" if best_score >= 2 else "low")
                mode_notes[note_key] = best_mode or "table"
                suggestion_targets[note_key] = suggested_display
            else:
                suggestion_notes[note_key] = "Suggested"
                score_notes[note_key] = best_score
                confidence_notes[note_key] = best_confidence or ("medium" if best_score >= 2 else "low")
                mode_notes[note_key] = best_mode or "table"
                suggestion_targets[note_key] = suggested_display


def _run_render_loop_suggestions(selected_codes: list, catalogs: dict):
    if not st.session_state.get("smart_matching_enabled", False):
        return

    if len(selected_codes) < 2:
        return

    for row_id in st.session_state.get("row_ids", []) or []:
        _apply_smart_product_suggestions_for_row(row_id, selected_codes, catalogs)


# -------------------------------------------------
# SAFE COMPANIES LOADING
# -------------------------------------------------
def load_companies_safe():
    default = pd.DataFrame(
        [
            {"code": "SINIAT", "name": "Siniat"},
            {"code": "KNAUF", "name": "Knauf"},
            {"code": "SAINT_GOBAIN", "name": "Saint-Gobain"},
        ]
    )

    if not COMPANIES_FILE.exists():
        default.to_csv(COMPANIES_FILE, index=False)
        return default

    try:
        df = pd.read_csv(COMPANIES_FILE)

        if "code" not in df.columns or "name" not in df.columns:
            default.to_csv(COMPANIES_FILE, index=False)
            return default

        df = df[["code", "name"]].copy()
        df["code"] = df["code"].astype(str).str.strip().str.upper()
        df["name"] = df["name"].astype(str).str.strip()

        if df.empty:
            default.to_csv(COMPANIES_FILE, index=False)
            return default

        return df

    except Exception:
        default.to_csv(COMPANIES_FILE, index=False)
        return default


def save_companies(df):
    df.to_csv(COMPANIES_FILE, index=False)


def normalize_code(text):
    text = str(text).strip().upper()
    text = re.sub(r"[^A-Z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text


def workspace_slug_from_user_row(user_row):
    raw_user = user_row.get("sub") or user_row.get("email") or ""
    return (
        str(raw_user)
        .replace("@", "_")
        .replace(".", "_")
        .replace("/", "_")
        .replace("\\", "_")
    )


def resolve_workspace_label(workspace_name, users_registry):
    for row in users_registry or []:
        try:
            if workspace_slug_from_user_row(row) == workspace_name:
                return row.get("email", workspace_name)
        except Exception:
            pass
    return workspace_name


companies_df = load_companies_safe()
for _, row in companies_df.iterrows():
    (UPLOADS_DIR / row["code"]).mkdir(parents=True, exist_ok=True)



def refresh_source_file_views():
    return None


def _ensure_preview_row_id(df: pd.DataFrame) -> pd.DataFrame:
    working = pd.DataFrame(df).copy().reset_index(drop=True)
    if "__row_id" not in working.columns:
        working.insert(0, "__row_id", range(1, len(working) + 1))
    else:
        row_ids = pd.to_numeric(working["__row_id"], errors="coerce")
        fallback_ids = pd.Series(range(1, len(working) + 1), index=working.index, dtype="int64")
        row_ids = row_ids.where(row_ids.notna(), fallback_ids)
        working["__row_id"] = row_ids.astype(int)
    return working


def _preview_display_df(df: pd.DataFrame) -> pd.DataFrame:
    working = _ensure_preview_row_id(df)
    out = working.copy()
    out = out.rename(columns={"__row_id": "Row"})
    return out


def _normalize_preview_editor_output(edited_df: pd.DataFrame) -> pd.DataFrame:
    working = pd.DataFrame(edited_df).copy()
    if "Row" in working.columns and "__row_id" not in working.columns:
        working = working.rename(columns={"Row": "__row_id"})
    working = _ensure_preview_row_id(working)

    expected_cols = _source_generator_output_columns()
    for col in expected_cols:
        if col not in working.columns:
            working[col] = ""

    working = working[["__row_id"] + expected_cols].copy()
    working["Base Price"] = pd.to_numeric(working["Base Price"], errors="coerce")
    working["Price"] = working["Base Price"]
    for text_col in ["SAP", "Product", "MM", "Package", "Category"]:
        working[text_col] = working[text_col].fillna("").astype(str).str.strip()

    working = working[
        ~(
            working["SAP"].eq("")
            & working["Product"].eq("")
            & working["Base Price"].isna()
        )
    ].reset_index(drop=True)
    working = _ensure_preview_row_id(working.drop(columns=["__row_id"], errors="ignore"))
    return working


def _delete_selected_preview_rows_from_state(state_key: str, selected_row_ids) -> int:
    df = st.session_state.get(state_key)
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return 0

    selected_set = {int(x) for x in selected_row_ids if str(x).strip()}
    if not selected_set:
        return 0

    working = _ensure_preview_row_id(df)
    before = len(working)
    working = working[~working["__row_id"].isin(selected_set)].copy().reset_index(drop=True)
    working = _ensure_preview_row_id(working.drop(columns=["__row_id"], errors="ignore"))
    deleted = before - len(working)
    st.session_state[state_key] = working
    return deleted


def _make_soft_delete_name(path_obj: Path) -> str:
    stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    return f"{stamp}__{path_obj.name}"


def _comparison_uses_source(record: dict, source_filename: str) -> bool:
    if not isinstance(record, dict):
        return False

    source_filename = str(source_filename).strip()
    if not source_filename:
        return False

    source_files = record.get("source_files", {}) or {}
    for _, value in source_files.items():
        if str(value).strip() == source_filename:
            return True

    state = record.get("state", {}) or {}
    if isinstance(state, dict):
        for key, value in state.items():
            if str(key).startswith("select_") and str(value).strip() == source_filename:
                return True

    blob = json.dumps(record, ensure_ascii=False, default=str)
    return source_filename in blob


def _find_comparisons_using_source(source_filename: str):
    matches = []
    for comparison_file in sorted(COMPARISONS_DIR.glob("*.json")):
        try:
            records = list_comparisons(comparison_file)
        except Exception:
            continue

        for record in records:
            if _comparison_uses_source(record, source_filename):
                matches.append({"comparison_file": comparison_file, "record": record})
    return matches


def _soft_delete_source_and_related_comparisons(source_path: Path):
    source_path = Path(source_path)
    source_filename = source_path.name

    matches = _find_comparisons_using_source(source_filename)
    archived_comparisons = []

    for item in matches:
        comparison_file = item["comparison_file"]
        record = item["record"]
        archived_record = {
            "deleted_at": now_iso(),
            "reason": f"Source soft-deleted: {source_filename}",
            "source_filename": source_filename,
            "from_comparison_file": comparison_file.name,
            "record": record,
        }
        archived_comparisons.append(archived_record)
        try:
            delete_comparison(comparison_file, record.get("id"))
        except Exception:
            pass

    archive_name = _make_soft_delete_name(source_path).replace(".xlsx", "").replace(".xlsm", "") + "__comparisons.json"
    archive_path = TRASH_COMPARISONS_DIR / archive_name
    archive_path.write_text(json.dumps(archived_comparisons, ensure_ascii=False, indent=2, default=str), encoding="utf-8")

    trashed_source_path = TRASH_SOURCES_DIR / _make_soft_delete_name(source_path)
    shutil.move(str(source_path), str(trashed_source_path))

    return {
        "deleted_comparisons_count": len(archived_comparisons),
        "trashed_source_path": str(trashed_source_path),
        "comparisons_archive_path": str(archive_path),
    }

# -------------------------------------------------
# FILE / CATALOG HELPERS
# -------------------------------------------------
def get_company_folder(code):
    folder = UPLOADS_DIR / code
    folder.mkdir(parents=True, exist_ok=True)
    return folder


def get_next_version_filename(company, dt, original_name):
    folder = get_company_folder(company)
    yyyy = dt.strftime("%Y")
    mm = dt.strftime("%m")
    dd = dt.strftime("%d")

    ext = Path(original_name).suffix.lower()
    if ext not in [".xlsx", ".xlsm"]:
        ext = ".xlsx"

    existing = list(folder.glob(f"{company}_{yyyy}_{mm}_{dd}_v*{ext}"))
    max_v = 0

    for f in existing:
        try:
            v = int(f.stem.split("_v")[-1])
            max_v = max(max_v, v)
        except Exception:
            pass

    return f"{company}_{yyyy}_{mm}_{dd}_v{max_v + 1}{ext}"


def sanitize_source_filename(value: str) -> str:
    cleaned = re.sub(r'[\/:*?"<>|]+', ' ', str(value or '').strip())
    cleaned = re.sub(r'\s+', ' ', cleaned).strip().strip('.')
    return cleaned


def build_custom_source_filename(custom_name: str, original_name: str) -> str:
    raw_name = str(custom_name or '').strip()
    requested_path = Path(raw_name)
    ext = requested_path.suffix.lower()
    if ext not in [".xlsx", ".xlsm"]:
        ext = Path(original_name).suffix.lower()
        if ext not in [".xlsx", ".xlsm"]:
            ext = ".xlsx"
        stem = raw_name
    else:
        stem = requested_path.stem

    stem = sanitize_source_filename(stem)
    if not stem:
        return ""
    return f"{stem}{ext}"


def get_company_files(code):
    folder = get_company_folder(code)
    files = []
    for f in sorted(folder.glob("*.*"), reverse=True):
        if f.suffix.lower() in [".xlsx", ".xlsm"]:
            files.append(f.name)
    return files


def list_saved_sources():
    rows = []

    for _, row in companies_df.iterrows():
        code = row["code"]
        name = row["name"]
        folder = get_company_folder(code)

        for f in sorted(folder.glob("*.*"), reverse=True):
            if f.suffix.lower() in [".xlsx", ".xlsm"]:
                rows.append(
                    {
                        "Company Code": code,
                        "Company Name": name,
                        "Filename": f.name,
                        "Folder": str(folder),
                        "Full Path": str(f),
                        "Modified": pd.to_datetime(f.stat().st_mtime, unit="s"),
                    }
                )

    if rows:
        return (
            pd.DataFrame(rows)
            .sort_values("Modified", ascending=False)
            .reset_index(drop=True)
        )

    return pd.DataFrame(
        columns=[
            "Company Code",
            "Company Name",
            "Filename",
            "Folder",
            "Full Path",
            "Modified",
        ]
    )


def company_has_files(code):
    folder = get_company_folder(code)
    for f in folder.glob("*.*"):
        if f.suffix.lower() in [".xlsx", ".xlsm"]:
            return True
    return False


def load_data(file):
    try:
        df = pd.read_excel(file, sheet_name="PRICELIST", engine="openpyxl")
        df.columns = [str(c).strip() for c in df.columns]
        return df
    except Exception as e:
        st.error(f"Error loading file: {e}")
        return None



def load_excel_file_any(uploaded_file):
    file_bytes = uploaded_file.getvalue()
    attempts = [
        ("openpyxl", io.BytesIO(file_bytes)),
    ]

    last_error = None
    for engine, buffer in attempts:
        try:
            return pd.ExcelFile(buffer, engine=engine), file_bytes
        except Exception as e:
            last_error = e

    try:
        import xlrd  # noqa: F401
        try:
            return pd.ExcelFile(io.BytesIO(file_bytes), engine="xlrd"), file_bytes
        except Exception as e:
            last_error = e
    except Exception:
        pass

    raise ValueError(
        "The uploaded file could not be read as a valid Excel workbook. "
        "If it was saved as .xls or exported with the wrong extension, please re-save it as a real .xlsx file and upload it again."
    ) from last_error


def read_excel_any(file_bytes, sheet_name, header=None):
    try:
        return pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, header=header, engine="openpyxl")
    except Exception as first_error:
        try:
            import xlrd  # noqa: F401
            return pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, header=header, engine="xlrd")
        except Exception:
            raise ValueError(
                "This workbook is not a standard .xlsx file that can be opened safely. "
                "Please open it in Excel and use Save As -> Excel Workbook (.xlsx), then upload it again."
            ) from first_error


def find_col(df, names):
    cols = {str(c).strip().lower(): c for c in df.columns}
    for n in names:
        if n.lower() in cols:
            return cols[n.lower()]
    return None


def prepare_catalog(df):
    if df is None:
        return None

    sap = find_col(df, ["SAP", "Κωδικός SAP"])
    prod = find_col(df, ["Product", "Προϊόν"])
    price = find_col(df, ["Price", "Τιμή €/ΜΜ", "Τιμή", "Price €/MM"])
    mm = find_col(df, ["ΜΜ πώλησης", "MM", "Unit", "ΜΜ"])
    pack = find_col(df, ["Συσκευασία", "Package", "pack"])
    category = find_col(df, ["Κατηγορία", "Category", "category"])

    if not sap or not prod or not price:
        return None

    out = pd.DataFrame()
    out["SAP"] = df[sap].astype(str).str.strip()
    out["Product"] = df[prod].astype(str).str.strip()
    out["Price"] = pd.to_numeric(df[price], errors="coerce")
    out["MM"] = df[mm].astype(str).str.strip() if mm else ""
    out["Package"] = df[pack].astype(str).str.strip() if pack else ""
    out["Category"] = df[category].astype(str).str.strip() if category else ""

    out = out.dropna(subset=["Price"])
    out = out[out["Price"] > 0]
    out = out.reset_index(drop=True)
    out["DISPLAY"] = out["Product"] + " | SAP " + out["SAP"]
    return out


def apply_discounts(price, discs):
    if price is None or pd.isna(price):
        return None

    p = float(price)
    for d in discs:
        if d is not None and d != 0:
            p *= 1 - d / 100
    return round(p, 2)


def format_total_discounts(discs):
    valid_discounts = []

    for d in discs:
        try:
            d = float(d)
        except Exception:
            continue

        if d > 0:
            if d.is_integer():
                valid_discounts.append(f"{int(d)}%")
            else:
                valid_discounts.append(f"{d:.1f}%")

    return ", ".join(valid_discounts)


def get_company_label(code):
    company_name_row = companies_df[companies_df["code"] == code]
    if company_name_row.empty:
        return code
    return company_name_row.iloc[0]["name"]


def compare_percent_text(a_name, a_price, b_name, b_price):
    if a_price is None or b_price is None:
        return ""

    try:
        a = float(a_price)
        b = float(b_price)
    except Exception:
        return ""

    if b == 0:
        return ""

    if round(a, 2) == round(b, 2):
        return f"Same price as {b_name}"

    if a < b:
        pct = ((b - a) / b) * 100
        return f"{pct:.1f}% cheaper than {b_name}"

    pct = ((a - b) / b) * 100
    return f"{pct:.1f}% more expensive than {b_name}"


def build_comparison_summary(final_prices, selected_codes):
    summaries = {}

    for code in selected_codes:
        label = get_company_label(code)
        a_price = final_prices.get(code)
        notes = []

        for other_code in selected_codes:
            if other_code == code:
                continue

            other_label = get_company_label(other_code)
            other_price = final_prices.get(other_code)

            note = compare_percent_text(label, a_price, other_label, other_price)
            if note:
                notes.append(note)

        summaries[code] = " | ".join(notes)

    return summaries



def _extract_sap_from_display_value(display_value: str) -> str:
    text = str(display_value or "").strip()
    if " | SAP " in text:
        return text.split(" | SAP ", 1)[1].strip()
    return ""


def _extract_product_from_display_value(display_value: str) -> str:
    text = str(display_value or "").strip()
    if " | SAP " in text:
        return text.split(" | SAP ", 1)[0].strip()
    return text


def _reconcile_selected_products_for_source_change(code, df):
    if df is None or df.empty:
        return

    valid_displays = set(df["DISPLAY"].astype(str).tolist())
    sap_to_display = {}
    product_to_display = {}

    for _, row in df.iterrows():
        display = str(row.get("DISPLAY", "")).strip()
        sap = str(row.get("SAP", "")).strip()
        product = str(row.get("Product", "")).strip()

        if display:
            if sap and sap not in sap_to_display:
                sap_to_display[sap] = display
            if product and product not in product_to_display:
                product_to_display[product] = display

    for row_id in list(st.session_state.get("row_ids", [])):
        data_key = f"row_{row_id}_{code}_product"
        widget_key = get_product_widget_key(row_id, code)
        selected_value = str(st.session_state.get(data_key, "") or "").strip()

        if not selected_value:
            continue

        if selected_value in valid_displays:
            st.session_state[widget_key] = selected_value
            continue

        remapped_value = ""
        sap = _extract_sap_from_display_value(selected_value)
        product = _extract_product_from_display_value(selected_value)

        if sap and sap in sap_to_display:
            remapped_value = sap_to_display[sap]
        elif product and product in product_to_display:
            remapped_value = product_to_display[product]

        st.session_state[data_key] = remapped_value
        st.session_state[widget_key] = remapped_value


def _handle_source_selection_change(code):
    selected_file = str(st.session_state.get(f"select_{code}", "") or "").strip()
    tracking_key = f"_last_source_selection_{code}"

    previous_file = str(st.session_state.get(tracking_key, "") or "").strip()
    st.session_state[tracking_key] = selected_file

    if not selected_file or selected_file == previous_file:
        return

    try:
        source_path = get_company_folder(code) / selected_file
        if not source_path.exists():
            return
        df = load_prepared_catalog_from_file(str(source_path), source_path.stat().st_mtime)
        _reconcile_selected_products_for_source_change(code, df)
    except Exception:
        return


def get_catalog_row(df, display_value):
    if df is None or df.empty or not display_value:
        return None

    rows = df[df["DISPLAY"] == display_value]
    if rows.empty:
        return None

    return rows.iloc[0]


CENTRAL_ENGINE = CentralMatchEngine(ADMIN_DIR / "central_match_table.json", 10)

def _engine_row_from_catalog_row(row):
    if row is None:
        return None
    return {
        "Product": str(row.get("Product", "") or ""),
        "Category": str(row.get("Category", "") or ""),
        "MM": str(row.get("MM", "") or ""),
    }

def _catalogs_for_saved_state(selected_codes, payload_state: dict, source_files_map: dict | None = None):
    catalogs = {}
    for code in selected_codes:
        selected_file = str(payload_state.get(f"select_{code}", "") or "").strip()
        if not selected_file and source_files_map:
            selected_file = str(source_files_map.get(get_company_label(code), "") or "").strip()

        if selected_file:
            source_path = get_company_folder(code) / selected_file
            try:
                catalogs[code] = load_prepared_catalog_from_file(str(source_path), source_path.stat().st_mtime)
            except Exception:
                catalogs[code] = None
        else:
            catalogs[code] = None
    return catalogs

def _register_payload_to_central_table(payload_state: dict, selected_codes: list, comparison_id: str = "", source_files_map: dict | None = None):
    if not isinstance(payload_state, dict) or not selected_codes or len(selected_codes) < 2:
        return

    engine_data = CENTRAL_ENGINE.load()
    catalogs = _catalogs_for_saved_state(selected_codes, payload_state, source_files_map)
    row_ids = payload_state.get("row_ids", []) or []

    for row_id in row_ids:
        row_choices = []
        for code in selected_codes:
            display_value = str(payload_state.get(f"row_{row_id}_{code}_product", "") or "").strip()
            if not display_value:
                continue
            row = get_catalog_row(catalogs.get(code), display_value)
            if row is None:
                continue
            engine_row = _engine_row_from_catalog_row(row)
            if engine_row:
                row_choices.append((code, engine_row))

        if len(row_choices) >= 2:
            CENTRAL_ENGINE.register_row_choices(engine_data, row_choices, comparison_id=str(comparison_id or ""))

    engine_data["meta"]["total_saved_events"] = int(engine_data.get("meta", {}).get("total_saved_events", 0)) + 1
    engine_data = CENTRAL_ENGINE.maybe_reevaluate(engine_data)
    CENTRAL_ENGINE.save(engine_data)

def rebuild_central_match_table_from_comparisons():
    existing = CENTRAL_ENGINE.load()
    empty = {
        "register": {},
        "stable": {},
        "quarantine": {},
        "seed": existing.get("seed", {}) if isinstance(existing, dict) else {},
        "meta": {
            "version": 3,
            "pair_mode": "symmetric_all_pairs",
            "reeval_threshold": 10,
            "total_saved_events": 0,
            "last_reeval_at_saved_event": 0,
            "has_seed_layer": True,
        },
    }
    CENTRAL_ENGINE.save(empty)

    display_to_code = {
        f"{row['name']} ({row['code']})": row["code"]
        for _, row in companies_df.iterrows()
    }

    for comparison_file in sorted(COMPARISONS_DIR.glob("*.json")):
        try:
            records = list_comparisons(comparison_file)
        except Exception:
            continue

        for rec in records:
            state = rec.get("state", {}) or {}
            selected_codes = []
            for display in state.get("comparison_company_selection", []) or []:
                if display in display_to_code:
                    selected_codes.append(display_to_code[display])

            if len(selected_codes) < 2:
                continue

            _register_payload_to_central_table(
                state,
                selected_codes,
                comparison_id=str(rec.get("id", "") or ""),
                source_files_map=rec.get("source_files", {}) or {},
            )


def ensure_central_table_schema_current():
    try:
        data = CENTRAL_ENGINE.load()
    except Exception:
        return

    meta = data.get("meta", {}) if isinstance(data, dict) else {}
    version = int(meta.get("version", 1) or 1)
    pair_mode = str(meta.get("pair_mode", "") or "").strip().lower()

    needs_rebuild = version < 3 or pair_mode != "symmetric_all_pairs"
    if not needs_rebuild:
        return

    rebuild_central_match_table_from_comparisons()


ensure_central_table_schema_current()


def get_row_summary_text(row_id, selected_codes, catalogs):
    parts = []
    for code in selected_codes:
        selected_product = str(st.session_state.get(f"row_{row_id}_{code}_product", "") or "").strip()
        if selected_product:
            row = get_catalog_row(catalogs.get(code), selected_product)
            if row is not None:
                parts.append(f"{get_company_label(code)}: {row['Product']}")
            else:
                parts.append(f"{get_company_label(code)}: selected")
    if parts:
        summary = " | ".join(parts[:2])
        if len(parts) > 2:
            summary += f" +{len(parts)-2} more"
        return summary
    return "Empty row"

def row_result_dict(visible_index, row_id, catalogs, selected_codes):
    result = {"Row": visible_index + 1}
    final_prices = {}

    for code in selected_codes:
        label = get_company_label(code)

        df = catalogs.get(code)
        selected_product = st.session_state.get(f"row_{row_id}_{code}_product", "")
        row = get_catalog_row(df, selected_product)

        discs = []
        for d in range(1, 6):
            discs.append(st.session_state.get(f"row_{row_id}_{code}_disc_{d}", 0.0))

        if row is not None:
            base_price = round(float(row["Price"]), 2)
            final_price, final_is_manual = get_effective_final_price(row_id, code, row["Price"], discs)
            final_prices[code] = final_price
            total_discounts_text = "Manual Price" if final_is_manual else format_total_discounts(discs)

            result[f"{label} Product"] = row["Product"]
            result[f"{label} SAP"] = row["SAP"]
            result[f"{label} MM"] = row["MM"]
            result[f"{label} Package"] = row["Package"]
            result[f"{label} Base Price"] = base_price
            result[f"{label} Total Discounts"] = total_discounts_text

            for i, disc in enumerate(discs, start=1):
                result[f"{label} Disc{i}"] = disc

            result[f"{label} Final Price"] = final_price
        else:
            final_prices[code] = None
            result[f"{label} Product"] = ""
            result[f"{label} SAP"] = ""
            result[f"{label} MM"] = ""
            result[f"{label} Package"] = ""
            result[f"{label} Base Price"] = ""
            result[f"{label} Total Discounts"] = format_total_discounts(discs)

            for i in range(1, 6):
                result[f"{label} Disc{i}"] = st.session_state.get(
                    f"row_{row_id}_{code}_disc_{i}", 0.0
                )

            result[f"{label} Final Price"] = ""

    comparison_summaries = build_comparison_summary(final_prices, selected_codes)

    combined_comparisons = []
    for code in selected_codes:
        label = get_company_label(code)
        summary = comparison_summaries.get(code, "")
        if summary:
            combined_comparisons.append(f"{label}: {summary}")

    result["Comparisons"] = " || ".join(combined_comparisons)

    valid = {k: v for k, v in final_prices.items() if v is not None}
    if valid:
        best_code = min(valid, key=valid.get)
        result["Best Price"] = get_company_label(best_code)
    else:
        result["Best Price"] = ""

    return result


def build_export_dataframe(row_ids, catalogs, selected_codes):
    rows = []
    for visible_index, row_id in enumerate(row_ids):
        rows.append(row_result_dict(visible_index, row_id, catalogs, selected_codes))
    return pd.DataFrame(rows)


# -------------------------------------------------
# EXPORT FIELD SELECTION
# -------------------------------------------------
EXPORT_FIELD_OPTIONS = [
    "Product",
    "SAP",
    "MM",
    "Package",
    "Base Price",
    "Total Discounts",
    "Final Price",
    "Comparison %",
    "Best Price",
]


def filter_export_dataframe(export_df, selected_codes, selected_fields, companies_df):
    if export_df.empty:
        return export_df

    columns_to_keep = ["Row"]

    for code in selected_codes:
        company_name_row = companies_df[companies_df["code"] == code]
        label = code if company_name_row.empty else company_name_row.iloc[0]["name"]

        for field in selected_fields:
            if field in ["Best Price", "Comparison %"]:
                continue

            col_name = f"{label} {field}"
            if col_name in export_df.columns:
                columns_to_keep.append(col_name)

    if "Comparison %" in selected_fields and "Comparisons" in export_df.columns:
        columns_to_keep.append("Comparisons")

    if "Best Price" in selected_fields and "Best Price" in export_df.columns:
        columns_to_keep.append("Best Price")

    columns_to_keep = [c for c in columns_to_keep if c in export_df.columns]
    return export_df[columns_to_keep]


def style_excel_worksheet(ws):
    title_fill = PatternFill(fill_type="solid", fgColor="0F172A")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    header_fill = PatternFill(fill_type="solid", fgColor="1E3A8A")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    subheader_fill = PatternFill(fill_type="solid", fgColor="DBEAFE")
    zebra_fill = PatternFill(fill_type="solid", fgColor="F8FAFC")
    siniat_fill = PatternFill(fill_type="solid", fgColor="E0F2FE")
    knauf_fill = PatternFill(fill_type="solid", fgColor="ECFCCB")
    sg_fill = PatternFill(fill_type="solid", fgColor="FCE7F3")
    result_fill = PatternFill(fill_type="solid", fgColor="FEF3C7")
    best_fill = PatternFill(fill_type="solid", fgColor="DCFCE7")

    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="CBD5E1")
    medium = Side(style="medium", color="94A3B8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_border = Border(left=medium, right=medium, top=medium, bottom=medium)

    max_col = ws.max_column

    ws.insert_rows(1, amount=2)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws["A1"] = "Pricing Comparison Report"
    ws["A1"].fill = title_fill
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align
    ws["A1"].border = header_border

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    ws["A2"] = f"Generated on {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    ws["A2"].fill = subheader_fill
    ws["A2"].font = Font(color="0F172A", italic=True, size=10)
    ws["A2"].alignment = left_align
    ws["A2"].border = border

    header_row = 3
    data_start_row = 4
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col)}{ws.max_row}"

    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = header_border

    headers = [cell.value for cell in ws[header_row]]

    for row_idx in range(data_start_row, ws.max_row + 1):
        is_zebra = (row_idx - data_start_row) % 2 == 1
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            header = headers[col_idx - 1]
            cell.border = border

            if isinstance(cell.value, (int, float)) and header != "Row":
                cell.number_format = '#,##0.00'
                cell.alignment = right_align
            else:
                cell.alignment = left_align

            if is_zebra and cell.fill.fill_type is None:
                cell.fill = zebra_fill

            if isinstance(header, str):
                if header.startswith("Siniat "):
                    cell.fill = siniat_fill
                elif header.startswith("Knauf "):
                    cell.fill = knauf_fill
                elif header.startswith("Saint-Gobain "):
                    cell.fill = sg_fill
                elif header in ["Best Price", "Comparisons"]:
                    cell.fill = result_fill

        best_price_col = None
        for idx, header in enumerate(headers, start=1):
            if header == "Best Price":
                best_price_col = idx
                break
        if best_price_col:
            ws.cell(row=row_idx, column=best_price_col).fill = best_fill
            ws.cell(row=row_idx, column=best_price_col).font = Font(bold=True, color="166534")

    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header)) if header is not None else 0
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            val_len = len(str(val)) if val is not None else 0
            if val_len > max_len:
                max_len = val_len

        if header == "Comparisons":
            adjusted_width = 70
        elif header == "Row":
            adjusted_width = 10
        elif isinstance(header, str) and ("Price" in header or "Disc" in header):
            adjusted_width = 14
        else:
            adjusted_width = min(max(max_len + 3, 14), 38)

        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 28

def to_excel_bytes(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Comparison Report")
        ws = writer.book["Comparison Report"]
        style_excel_worksheet(ws)

    output.seek(0)
    return output.getvalue()



# -------------------------------------------------
# COMPARISON STORAGE HELPERS
# -------------------------------------------------
def auto_comparison_name(selected_codes):
    if not selected_codes:
        return f"Comparison {datetime.utcnow().strftime('%Y-%m-%d %H:%M')}"
    labels = [get_company_label(code) for code in selected_codes]
    return " vs ".join(labels)


def build_source_files_map(selected_codes):
    source_files = {}
    for code in selected_codes:
        label = get_company_label(code)
        source_files[label] = st.session_state.get(f"select_{code}", "")
    return source_files


def collect_comparison_state_payload(selected_codes):
    payload = {}

    static_keys = [
        "row_ids",
        "next_row_id",
        "comparison_company_selection",
        "selected_export_fields",
    ]

    for key in static_keys:
        if key in st.session_state:
            payload[key] = st.session_state[key]

    for code in selected_codes:
        select_key = f"select_{code}"
        if select_key in st.session_state:
            payload[select_key] = st.session_state[select_key]

        carry_key = f"carry_forward_{code}"
        if carry_key in st.session_state:
            payload[carry_key] = st.session_state[carry_key]

    for row_id in st.session_state.get("row_ids", []):
        for code in selected_codes:
            product_key = f"row_{row_id}_{code}_product"
            if product_key in st.session_state:
                payload[product_key] = st.session_state[product_key]

            manual_final_key = get_manual_final_price_data_key(row_id, code)
            if manual_final_key in st.session_state:
                payload[manual_final_key] = st.session_state[manual_final_key]

            for j in range(1, 6):
                disc_key = f"row_{row_id}_{code}_disc_{j}"
                if disc_key in st.session_state:
                    payload[disc_key] = st.session_state[disc_key]

    return payload


def collect_merged_comparison_state_payload(selected_codes):
    payload = dict(st.session_state.get("active_loaded_state_payload", {}) or {})
    current_payload = collect_comparison_state_payload(selected_codes)
    payload.update(current_payload)

    # remove stale row keys for rows that no longer exist in current state
    current_row_ids = set(current_payload.get("row_ids", st.session_state.get("row_ids", [])))
    keys_to_remove = []
    for key in payload.keys():
        if key.startswith("row_"):
            parts = key.split("_")
            if len(parts) >= 2:
                try:
                    row_id = int(parts[1])
                    if row_id not in current_row_ids:
                        keys_to_remove.append(key)
                except Exception:
                    pass
    for key in keys_to_remove:
        payload.pop(key, None)

    return payload


def restore_comparison_state_payload(payload: dict):
    keys_to_clear = [
        key for key in list(st.session_state.keys())
        if key.startswith("row_") or key.startswith("select_") or key.startswith("carry_forward_") or key.startswith("widget_row_")
    ]

    for key in keys_to_clear:
        if key in st.session_state:
            del st.session_state[key]

    protected_keys = {
        "current_comparison_id",
        "comparison_name_input",
        "show_saved_comparisons",
        "show_new_comparison_confirm",
        "pending_load_payload",
        "pending_loaded_comparison_id",
        "pending_loaded_comparison_name",
        "pending_clear_comparison",
    }

    for key, value in payload.items():
        if key not in protected_keys:
            st.session_state[key] = value


def clear_current_comparison_state():
    release_comparison_lock()
    keys_to_clear = [
        key for key in list(st.session_state.keys())
        if key.startswith("row_") or key.startswith("select_") or key.startswith("carry_forward_") or key.startswith("widget_row_")
    ]

    for key in keys_to_clear:
        if key in st.session_state:
            del st.session_state[key]

    st.session_state["row_ids"] = [1]
    st.session_state["active_row_id"] = None
    st.session_state["next_row_id"] = 2
    st.session_state["comparison_company_selection"] = []
    st.session_state["comparison_name_input"] = ""
    st.session_state["pending_comparison_name_input"] = None
    st.session_state["active_comparison_label"] = ""
    st.session_state["active_loaded_state_payload"] = {}
    st.session_state["comparison_loaded_from_record"] = False
    st.session_state["comparison_baseline_state_json"] = ""
    st.session_state["comparison_dirty"] = False
    st.session_state["comparison_user_modified"] = False
    st.session_state["comparison_edit_generation"] = 0
    st.session_state["comparison_clean_generation"] = 0
    st.session_state["current_comparison_id"] = None
    st.session_state["show_saved_comparisons"] = False
    st.session_state["show_new_comparison_confirm"] = False
    st.session_state["comparison_mode"] = "menu"
    st.session_state["selected_export_fields"] = [
        "Product",
        "Total Discounts",
        "Final Price",
        "Comparison %",
        "Best Price",
    ]


def load_selected_comparison_record(selected_record):
    if not selected_record:
        return False, "Could not load comparison."

    comparison_file = get_current_user_comparisons_file()
    comparison_id = selected_record.get("id")
    comparison_name = selected_record.get("name", "")
    ok_lock, lock_msg = acquire_comparison_lock(comparison_file, comparison_id, comparison_name)
    if not ok_lock:
        return False, lock_msg

    state_payload = selected_record.get("state", {}) or {}
    st.session_state["active_loaded_state_payload"] = dict(state_payload)
    st.session_state["pending_load_payload"] = state_payload
    st.session_state["pending_loaded_comparison_id"] = selected_record.get("id")
    st.session_state["pending_loaded_comparison_name"] = selected_record.get("name", "")
    st.session_state["active_comparison_label"] = selected_record.get("name", "")
    st.session_state["comparison_loaded_from_record"] = True
    st.session_state["comparison_mode"] = "edit"
    st.session_state["show_saved_comparisons"] = False
    return True, ""


def save_or_update_current_comparison(selected_codes):
    if not selected_codes:
        return False, "Please select companies first."

    comparison_name = st.session_state.get("comparison_name_input", "").strip()
    if not comparison_name:
        comparison_name = auto_comparison_name(selected_codes)
        st.session_state["comparison_name_input"] = comparison_name

    comparison_file = get_current_user_comparisons_file()
    current_id = st.session_state.get("current_comparison_id")

    if current_id:
        comparison_lock = get_comparison_lock_info(comparison_file, current_id)
        if comparison_lock and not comparison_lock_owned_by_current_session(comparison_lock):
            holder = comparison_lock.get("owner_name") or comparison_lock.get("owner_email") or "another user"
            return False, f"This comparison is currently locked by {holder}."

        ok = update_comparison(
            comparison_file,
            comparison_id=current_id,
            owner_sub=get_current_user_id(),
            owner_email=get_current_user_email(),
            name=comparison_name,
            companies=[get_company_label(code) for code in selected_codes],
            source_files=build_source_files_map(selected_codes),
            state=collect_comparison_state_payload(selected_codes),
        )
        if ok:
            return True, "Comparison updated successfully."
        return False, "This comparison no longer exists. Save it again as new."

    comparison_id = save_new_comparison(
        comparison_file,
        owner_sub=get_current_user_id(),
        owner_email=get_current_user_email(),
        name=comparison_name,
        companies=[get_company_label(code) for code in selected_codes],
        source_files=build_source_files_map(selected_codes),
        state=collect_comparison_state_payload(selected_codes),
    )
    st.session_state["current_comparison_id"] = comparison_id
    _register_payload_to_central_table(payload_state, selected_codes, comparison_id=str(comparison_id or ""), source_files_map=payload_sources)
    return True, "Comparison saved successfully."


def save_current_comparison_from_state(force_new: bool = False, override_name: str | None = None):
    selected_codes = get_current_selected_codes_from_state()

    for row_id in st.session_state.get("row_ids", []):
        for code in selected_codes:
            sync_product_widget_to_data(row_id, code)
            for disc_number in range(1, 6):
                sync_discount_widget_to_data(row_id, code, disc_number)
    if not selected_codes:
        return False, "Please select companies first."

    comparison_name = (override_name or st.session_state.get("comparison_name_input", "")).strip()
    if not comparison_name:
        comparison_name = auto_comparison_name(selected_codes)

    comparison_file = get_current_user_comparisons_file()
    current_id = None if force_new else st.session_state.get("current_comparison_id")
    payload_state = (
        collect_merged_comparison_state_payload(selected_codes)
        if st.session_state.get("comparison_loaded_from_record") or st.session_state.get("active_loaded_state_payload")
        else collect_comparison_state_payload(selected_codes)
    )
    payload_companies = [get_company_label(code) for code in selected_codes]
    payload_sources = build_source_files_map(selected_codes)

    if current_id:
        ok = update_comparison(
            comparison_file,
            comparison_id=current_id,
            owner_sub=get_current_user_id(),
            owner_email=get_current_user_email(),
            name=comparison_name,
            companies=payload_companies,
            source_files=payload_sources,
            state=payload_state,
        )
        if ok:
            acquire_comparison_lock(comparison_file, current_id, comparison_name)
            st.session_state["pending_comparison_name_input"] = comparison_name
            st.session_state["active_comparison_label"] = comparison_name
            st.session_state["active_loaded_state_payload"] = dict(payload_state)
            st.session_state["comparison_loaded_from_record"] = True
            st.session_state["comparison_dirty"] = False
            st.session_state["comparison_user_modified"] = False
            st.session_state["comparison_clean_generation"] = st.session_state.get("comparison_edit_generation", 0)
            rebuild_match_history_from_scratch()
            _register_payload_to_central_table(payload_state, selected_codes, comparison_id=str(current_id or ""), source_files_map=payload_sources)
            return True, "Comparison updated successfully."
        return False, "This comparison no longer exists. Save it again as new."

    comparison_id = save_new_comparison(
        comparison_file,
        owner_sub=get_current_user_id(),
        owner_email=get_current_user_email(),
        name=comparison_name,
        companies=payload_companies,
        source_files=payload_sources,
        state=payload_state,
    )
    _register_payload_to_central_table(payload_state, selected_codes, comparison_id=str(comparison_id or ""), source_files_map=payload_sources)
    acquire_comparison_lock(comparison_file, comparison_id, comparison_name)
    st.session_state["current_comparison_id"] = comparison_id
    st.session_state["pending_comparison_name_input"] = comparison_name
    st.session_state["active_comparison_label"] = comparison_name
    st.session_state["active_loaded_state_payload"] = dict(payload_state)
    st.session_state["comparison_loaded_from_record"] = True
    st.session_state["comparison_dirty"] = False
    st.session_state["comparison_user_modified"] = False
    st.session_state["comparison_clean_generation"] = st.session_state.get("comparison_edit_generation", 0)
    rebuild_match_history_from_scratch()
    return True, "Comparison saved successfully."


def get_previous_row_discounts(current_row_id, code):
    row_ids = st.session_state.get("row_ids", [])
    try:
        idx = row_ids.index(current_row_id)
    except ValueError:
        return [0.0] * 5

    if idx <= 0:
        return [0.0] * 5

    prev_row_id = row_ids[idx - 1]
    values = []
    for j in range(1, 6):
        try:
            values.append(float(st.session_state.get(f"row_{prev_row_id}_{code}_disc_{j}", 0.0) or 0.0))
        except Exception:
            values.append(0.0)
    return values


def row_discounts_are_blank(row_id, code):
    values = []
    for j in range(1, 6):
        disc_key = f"row_{row_id}_{code}_disc_{j}"
        raw = st.session_state.get(disc_key, None)
        if raw is None:
            values.append(None)
            continue
        try:
            values.append(float(raw))
        except Exception:
            values.append(None)
    present_values = [v for v in values if v is not None]
    if not present_values:
        return True
    return all(abs(v) < 1e-9 for v in present_values)


def ensure_discount_defaults_for_row(row_id, selected_codes):
    for code in selected_codes:
        carry_enabled = bool(st.session_state.get(f"carry_forward_{code}", False))

        # ALWAYS ensure keys exist first (critical fix)
        for j in range(1, 6):
            disc_key = f"row_{row_id}_{code}_disc_{j}"
            if disc_key not in st.session_state or st.session_state[disc_key] is None:
                st.session_state[disc_key] = 0.0

        manual_final_key = get_manual_final_price_data_key(row_id, code)
        if manual_final_key not in st.session_state or st.session_state[manual_final_key] is None:
            st.session_state[manual_final_key] = ""

        # THEN apply carry forward only if needed
        if carry_enabled and row_id != st.session_state.row_ids[0] and row_discounts_are_blank(row_id, code):
            base_values = get_previous_row_discounts(row_id, code)
            for j, value in enumerate(base_values, start=1):
                st.session_state[f"row_{row_id}_{code}_disc_{j}"] = float(value)


def add_comparison_row(selected_codes, insert_after_row_id=None):
    new_row_id = st.session_state.next_row_id
    st.session_state.next_row_id += 1

    current_rows = list(st.session_state.get("row_ids", []))
    if insert_after_row_id is not None and insert_after_row_id in current_rows:
        insert_index = current_rows.index(insert_after_row_id) + 1
        current_rows.insert(insert_index, new_row_id)
    else:
        current_rows.append(new_row_id)

    st.session_state.row_ids = current_rows
    st.session_state["pending_focus_row_id"] = new_row_id
    st.session_state["active_row_id"] = new_row_id

    for code in selected_codes:
        product_key = f"row_{new_row_id}_{code}_product"
        if product_key not in st.session_state:
            st.session_state[product_key] = ""

        st.session_state[get_manual_final_price_data_key(new_row_id, code)] = ""

        carry_enabled = bool(st.session_state.get(f"carry_forward_{code}", False))
        base_values = get_previous_row_discounts(new_row_id, code) if carry_enabled else [0.0] * 5

        for j, value in enumerate(base_values, start=1):
            disc_key = f"row_{new_row_id}_{code}_disc_{j}"
            st.session_state[disc_key] = float(value)


def focus_existing_row(target_row_id):
    if target_row_id in st.session_state.get("row_ids", []):
        st.session_state["pending_focus_row_id"] = target_row_id
        st.session_state["active_row_id"] = target_row_id


def snapshot_current_comparison_state():
    selected_codes = get_current_selected_codes_from_state()
    return {
        "payload": collect_merged_comparison_state_payload(selected_codes),
        "comparison_mode": st.session_state.get("comparison_mode", "menu"),
        "show_saved_comparisons": st.session_state.get("show_saved_comparisons", False),
        "show_inline_save_options": st.session_state.get("show_inline_save_options", False),
        "inline_save_mode": st.session_state.get("inline_save_mode", "menu"),
        "active_save_row_id": st.session_state.get("active_save_row_id"),
        "pending_inline_save_as_name": st.session_state.get("pending_inline_save_as_name", ""),
        "pending_save_as_exit_name": st.session_state.get("pending_save_as_exit_name", ""),
        "current_comparison_id": st.session_state.get("current_comparison_id"),
        "comparison_name_input": st.session_state.get("comparison_name_input", ""),
        "active_comparison_label": st.session_state.get("active_comparison_label", ""),
        "comparison_loaded_from_record": st.session_state.get("comparison_loaded_from_record", False),
        "active_loaded_state_payload": dict(st.session_state.get("active_loaded_state_payload", {}) or {}),
        "comparison_dirty": st.session_state.get("comparison_dirty", False),
        "comparison_user_modified": st.session_state.get("comparison_user_modified", False),
        "comparison_edit_generation": st.session_state.get("comparison_edit_generation", 0),
        "comparison_clean_generation": st.session_state.get("comparison_clean_generation", 0),
    }


def open_leave_prompt(action_type, target_view=None, payload=None):
    st.session_state["leave_prompt_snapshot"] = snapshot_current_comparison_state()
    st.session_state["show_leave_prompt"] = True
    st.session_state["leave_prompt_step"] = ""
    st.session_state["pending_action_type"] = action_type
    st.session_state["pending_target_view"] = target_view
    st.session_state["pending_action_payload"] = payload


def restore_leave_prompt_snapshot():
    snapshot = st.session_state.get("leave_prompt_snapshot") or {}
    payload = snapshot.get("payload")
    if payload is not None:
        restore_comparison_state_payload(payload)
    for key in [
        "comparison_mode",
        "show_saved_comparisons",
        "show_inline_save_options",
        "inline_save_mode",
        "active_save_row_id",
        "pending_inline_save_as_name",
        "pending_save_as_exit_name",
        "current_comparison_id",
        "comparison_name_input",
        "active_comparison_label",
        "comparison_loaded_from_record",
        "active_loaded_state_payload",
        "comparison_dirty",
        "comparison_user_modified",
        "comparison_edit_generation",
        "comparison_clean_generation",
    ]:
        if key in snapshot:
            st.session_state[key] = snapshot[key]
    st.session_state["show_leave_prompt"] = False
    st.session_state["leave_prompt_step"] = ""
    st.session_state["pending_action_type"] = None
    st.session_state["pending_target_view"] = None
    st.session_state["pending_action_payload"] = None
    st.session_state["leave_prompt_snapshot"] = None


def execute_pending_leave_action():
    action_type = st.session_state.get("pending_action_type")
    target_view = st.session_state.get("pending_target_view")
    payload = st.session_state.get("pending_action_payload")

    st.session_state["show_leave_prompt"] = False
    st.session_state["leave_prompt_step"] = ""
    st.session_state["pending_action_type"] = None
    st.session_state["pending_target_view"] = None
    st.session_state["pending_action_payload"] = None
    st.session_state["pending_save_as_exit_name"] = ""
    st.session_state["leave_prompt_snapshot"] = None

    if action_type == "switch_view" and target_view:
        release_comparison_lock()
        st.session_state["committed_view"] = target_view
        if target_view == "Comparisons":
            st.session_state["comparison_mode"] = "menu"
            st.session_state["show_saved_comparisons"] = False
            st.session_state["show_inline_save_options"] = False
            st.session_state["inline_save_mode"] = "menu"
            st.session_state["active_save_row_id"] = None
            st.session_state["pending_inline_save_as_name"] = ""
            st.session_state["pending_save_as_exit_name"] = ""
    elif action_type == "logout":
        release_comparison_lock()
        logout_current_user()
    elif action_type == "clear_comparison":
        release_comparison_lock()
        st.session_state["show_new_comparison_confirm"] = False
        st.session_state["pending_clear_comparison"] = True
    elif action_type == "load_comparison" and payload:
        release_comparison_lock()
        st.session_state["pending_load_payload"] = payload.get("state_payload")
        st.session_state["pending_loaded_comparison_id"] = payload.get("comparison_id")
        st.session_state["pending_loaded_comparison_name"] = payload.get("comparison_name", "")
        st.session_state["show_saved_comparisons"] = False


def render_row_navigation_buttons(current_row_id, visible_index):
    row_ids = st.session_state.get("row_ids", [])
    if not row_ids:
        return

    first_row_id = row_ids[0]
    last_row_id = row_ids[-1]
    prev_row_id = row_ids[visible_index - 1] if visible_index > 0 else None
    next_row_id = row_ids[visible_index + 1] if visible_index < len(row_ids) - 1 else None

    def make_btn(symbol, target, disabled=False):
        if disabled or not target:
            return f'<span style="font-size:30px;opacity:0.3;margin:6px;">{symbol}</span>'
        return f'<a href="#row-anchor-{target}" style="font-size:30px;margin:6px;text-decoration:none;">{symbol}</a>'

    nav_html = f"""
    <div style="display:flex;gap:12px;align-items:center;">
        {make_btn("⏮", first_row_id, visible_index==0)}
        {make_btn("◀", prev_row_id, prev_row_id is None)}
        {make_btn("▶", next_row_id, next_row_id is None)}
        {make_btn("⏭", last_row_id, visible_index==len(row_ids)-1)}
    </div>
    """
    st.markdown(nav_html, unsafe_allow_html=True)


def apply_specific_discount_to_all_rows(selected_codes, target_code, disc_index, disc_value):
    if target_code not in selected_codes:
        return
    try:
        disc_number = int(disc_index)
    except Exception:
        return
    if disc_number < 1 or disc_number > 5:
        return

    for row_id in st.session_state.get("row_ids", []):
        current_values = {}
        for j in range(1, 6):
            disc_key = f"row_{row_id}_{target_code}_disc_{j}"
            if disc_key in st.session_state:
                try:
                    current_values[j] = float(st.session_state.get(disc_key, 0.0) or 0.0)
                except Exception:
                    current_values[j] = 0.0
            else:
                current_values[j] = 0.0

        current_values[disc_number] = float(disc_value)

        for j in range(1, 6):
            st.session_state[f"row_{row_id}_{target_code}_disc_{j}"] = float(current_values[j])


def get_discount_widget_key(row_id, code, disc_number):
    return f"widget_row_{row_id}_{code}_disc_{disc_number}"


def sync_discount_widget_to_data(row_id, code, disc_number):
    data_key = f"row_{row_id}_{code}_disc_{disc_number}"
    widget_key = get_discount_widget_key(row_id, code, disc_number)
    if widget_key not in st.session_state:
        return
    try:
        st.session_state[data_key] = float(st.session_state.get(widget_key, 0.0) or 0.0)
    except Exception:
        st.session_state[data_key] = 0.0


def mirror_discount_data_to_widget(row_id, code, disc_number):
    data_key = f"row_{row_id}_{code}_disc_{disc_number}"
    widget_key = get_discount_widget_key(row_id, code, disc_number)
    try:
        value = float(st.session_state.get(data_key, 0.0) or 0.0)
    except Exception:
        value = 0.0
    st.session_state[widget_key] = value



def _parse_manual_final_price(value):
    text = str(value or "").strip()
    if text == "":
        return None
    text = text.replace("€", "").replace(",", ".").strip()
    text = re.sub(r"[^0-9.\-]", "", text)
    if text.count(".") > 1:
        parts = text.split(".")
        text = "".join(parts[:-1]) + "." + parts[-1]
    try:
        return round(float(text), 2)
    except Exception:
        return None


def get_manual_final_price_data_key(row_id, code):
    return f"row_{row_id}_{code}_manual_final_price"


def get_manual_final_price_widget_key(row_id, code):
    return f"widget_row_{row_id}_{code}_manual_final_price"


def sync_manual_final_price_widget_to_data(row_id, code):
    data_key = get_manual_final_price_data_key(row_id, code)
    widget_key = get_manual_final_price_widget_key(row_id, code)
    if widget_key not in st.session_state:
        return
    st.session_state[data_key] = str(st.session_state.get(widget_key, "") or "").strip()


def mirror_manual_final_price_data_to_widget(row_id, code):
    data_key = get_manual_final_price_data_key(row_id, code)
    widget_key = get_manual_final_price_widget_key(row_id, code)
    data_value = str(st.session_state.get(data_key, "") or "").strip()
    if st.session_state.get(widget_key, "") != data_value:
        st.session_state[widget_key] = data_value


def get_effective_final_price(row_id, code, base_price, discounts):
    manual_value = _parse_manual_final_price(st.session_state.get(get_manual_final_price_data_key(row_id, code), ""))
    if manual_value is not None:
        return manual_value, True
    return apply_discounts(base_price, discounts), False

def get_product_widget_key(row_id, code):
    return f"widget_row_{row_id}_{code}_product"


def sync_product_widget_to_data(row_id, code):
    data_key = f"row_{row_id}_{code}_product"
    widget_key = get_product_widget_key(row_id, code)
    if widget_key not in st.session_state:
        return
    st.session_state[data_key] = st.session_state.get(widget_key, "")


def mirror_product_data_to_widget(row_id, code):
    data_key = f"row_{row_id}_{code}_product"
    widget_key = get_product_widget_key(row_id, code)
    data_value = st.session_state.get(data_key, "")
    if st.session_state.get(widget_key) != data_value:
        st.session_state[widget_key] = data_value



# -------------------------------------------------
# SESSION STATE
# -------------------------------------------------
if "row_ids" not in st.session_state:
    st.session_state.row_ids = [1]

if "next_row_id" not in st.session_state:
    st.session_state.next_row_id = 2

if "current_comparison_id" not in st.session_state:
    st.session_state.current_comparison_id = None

if "comparison_name_input" not in st.session_state:
    st.session_state.comparison_name_input = ""

if "pending_comparison_name_input" not in st.session_state:
    st.session_state["pending_comparison_name_input"] = None

if "post_save_success_message" not in st.session_state:
    st.session_state["post_save_success_message"] = ""

if "show_saved_comparisons" not in st.session_state:
    st.session_state.show_saved_comparisons = False

if "show_new_comparison_confirm" not in st.session_state:
    st.session_state.show_new_comparison_confirm = False

if "pending_load_payload" not in st.session_state:
    st.session_state.pending_load_payload = None

if "pending_clear_comparison" not in st.session_state:
    st.session_state.pending_clear_comparison = False

if "show_export_preview" not in st.session_state:
    st.session_state["show_export_preview"] = False

if "pending_loaded_comparison_id" not in st.session_state:
    st.session_state.pending_loaded_comparison_id = None

if "pending_loaded_comparison_name" not in st.session_state:
    st.session_state.pending_loaded_comparison_name = ""

if "selected_export_fields" not in st.session_state:
    st.session_state["selected_export_fields"] = [
        "Product",
        "Total Discounts",
        "Final Price",
        "Comparison %",
        "Best Price",
    ]

if "pending_company_delete_code" not in st.session_state:
    st.session_state["pending_company_delete_code"] = None

if "pending_company_delete_display" not in st.session_state:
    st.session_state["pending_company_delete_display"] = ""

if "pending_focus_row_id" not in st.session_state:
    st.session_state["pending_focus_row_id"] = None

if "active_row_id" not in st.session_state:
    st.session_state["active_row_id"] = None

if "bulk_discount_success_message" not in st.session_state:
    st.session_state["bulk_discount_success_message"] = ""

if "comparison_loaded_success_message" not in st.session_state:
    st.session_state["comparison_loaded_success_message"] = ""

if "comparison_saved_signature" not in st.session_state:
    st.session_state["comparison_saved_signature"] = ""

if "comparison_dirty" not in st.session_state:
    st.session_state["comparison_dirty"] = False

if "committed_view" not in st.session_state:
    st.session_state["committed_view"] = "Comparisons"

if "show_leave_prompt" not in st.session_state:
    st.session_state["show_leave_prompt"] = False

if "leave_prompt_step" not in st.session_state:
    st.session_state["leave_prompt_step"] = ""

if "pending_target_view" not in st.session_state:
    st.session_state["pending_target_view"] = None

if "pending_action_type" not in st.session_state:
    st.session_state["pending_action_type"] = None

if "pending_action_payload" not in st.session_state:
    st.session_state["pending_action_payload"] = None

if "save_as_exit_name" not in st.session_state:
    st.session_state["save_as_exit_name"] = ""

if "pending_save_as_exit_name" not in st.session_state:
    st.session_state["pending_save_as_exit_name"] = None

if "active_comparison_label" not in st.session_state:
    st.session_state["active_comparison_label"] = ""

if "active_loaded_state_payload" not in st.session_state:
    st.session_state["active_loaded_state_payload"] = {}

if "comparison_edit_generation" not in st.session_state:
    st.session_state["comparison_edit_generation"] = 0

if "comparison_clean_generation" not in st.session_state:
    st.session_state["comparison_clean_generation"] = 0

if "comparison_user_modified" not in st.session_state:
    st.session_state["comparison_user_modified"] = False

if "comparison_mode" not in st.session_state:
    st.session_state["comparison_mode"] = "menu"

if "show_inline_save_options" not in st.session_state:
    st.session_state["show_inline_save_options"] = False

if "inline_save_mode" not in st.session_state:
    st.session_state["inline_save_mode"] = "menu"

if "inline_save_as_name" not in st.session_state:
    st.session_state["inline_save_as_name"] = ""

if "pending_inline_save_as_name" not in st.session_state:
    st.session_state["pending_inline_save_as_name"] = None

if "comparison_baseline_state_json" not in st.session_state:
    st.session_state["comparison_baseline_state_json"] = ""

if "comparison_loaded_from_record" not in st.session_state:
    st.session_state["comparison_loaded_from_record"] = False

if "skip_export_preview_once" not in st.session_state:
    st.session_state["skip_export_preview_once"] = False


# -------------------------------------------------
# APP FLOW
# -------------------------------------------------
if not is_logged_in():
    show_login_screen()
    st.stop()

ensure_current_user_in_registry()
touch_current_user()

company_result = sync_company_assignment_from_domain()

if company_result["status"] == "full":
    company = company_result["company"]
    st.markdown(
        """
        <div class="locked-wrap">
            <div class="locked-badge">Company seat limit reached</div>
            <div class="locked-title">No available company seats</div>
            <div class="locked-subtitle">
                Your company account has reached the maximum number of active users allowed for this plan.
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    c1, c2, c3 = st.columns([1, 1.6, 1])
    with c2:
        st.warning(
            f"{company.get('name', 'This company')} is using all seats ({format_company_seats(company)})."
        )
        st.button(
            "Logout",
            on_click=logout_current_user,
            use_container_width=True,
            key="company_full_logout",
        )
    st.stop()

if current_user_is_blocked():
    st.error("Access denied. Your account has been blocked.")
    st.button(
        "Logout",
        on_click=logout_current_user,
        use_container_width=True,
        key="blocked_logout",
    )
    st.stop()

if not is_admin_user():
    session_allowed, active_count = register_current_session()
    dynamic_session_limit = get_user_max_active_sessions(get_current_user_registry_row()[1])

    if not session_allowed:
        st.markdown(
            """
            <div class="locked-wrap">
                <div class="locked-badge">Device limit reached</div>
                <div class="locked-title">Too many active devices</div>
                <div class="locked-subtitle">
                    Your account is already active on the maximum allowed number of devices/browsers for your plan.
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        st.info("You already have an active session on another device.")
        st.info(f"Active sessions: {active_count}/{dynamic_session_limit}.")
        if st.button("🔄 Replace previous session and continue", use_container_width=True, key="replace_session_limit_btn"):
            st.session_state["force_replace_session"] = True
            st.rerun()
        st.button(
            "Logout",
            on_click=logout_current_user,
            use_container_width=True,
            key="device_limit_logout",
        )
        st.stop()
    else:
        touch_current_session()

user_email = get_current_user_email()
current_company = get_current_user_company()
user_idx, user_row, _users = get_current_user_registry_row()

if (not is_admin_user()) and (not current_user_has_access()):
    checkout_url = get_checkout_url(user_email)

    st.markdown(
        """
        <div class="locked-wrap">
            <div class="locked-badge">Access required</div>
            <div class="locked-title">Your access is currently locked</div>
            <div class="locked-subtitle">
                Activate an individual plan to continue using all features.
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    days_left = trial_days_left(user_row.get("trial_end")) if user_row else 0
    if days_left > 0:
        st.info(f"Your free trial is still active. Days left: {days_left}")
    else:
        st.warning("Your 2-day free trial has expired.")

    if checkout_url:
        st.link_button("Subscribe €10/month", checkout_url, use_container_width=True)

    st.button(
        "Logout",
        on_click=logout_current_user,
        use_container_width=True,
        key="locked_logout",
    )
    st.stop()


# -------------------------------------------------
# APPLY PENDING COMPARISON LOAD BEFORE WIDGETS
# -------------------------------------------------
if st.session_state.get("pending_load_payload") is not None:
    restore_comparison_state_payload(st.session_state["pending_load_payload"])
    if "selected_export_fields" not in st.session_state or not st.session_state.get("selected_export_fields"):
        st.session_state["selected_export_fields"] = [
            "Product",
            "Total Discounts",
            "Final Price",
            "Comparison %",
            "Best Price",
        ]
    st.session_state["current_comparison_id"] = st.session_state.get("pending_loaded_comparison_id")
    st.session_state["comparison_name_input"] = st.session_state.get("pending_loaded_comparison_name", "")
    st.session_state["active_comparison_label"] = st.session_state.get("pending_loaded_comparison_name", "")
    st.session_state["comparison_loaded_from_record"] = True
    st.session_state["pending_load_payload"] = None
    st.session_state["pending_loaded_comparison_id"] = None
    st.session_state["pending_loaded_comparison_name"] = ""
    st.session_state["comparison_loaded_success_message"] = "Comparison loaded successfully."
    st.session_state["active_row_id"] = None
    st.session_state["comparison_mode"] = "edit"
    st.session_state["comparison_dirty"] = False
    st.session_state["comparison_clean_generation"] = st.session_state.get("comparison_edit_generation", 0)
    st.session_state["active_loaded_state_payload"] = collect_comparison_state_payload(get_current_selected_codes_from_state())
    mark_comparison_clean()

if st.session_state.get("pending_clear_comparison"):
    clear_current_comparison_state()
    st.session_state["pending_clear_comparison"] = False
    mark_comparison_clean()


if not st.session_state.get("comparison_saved_signature"):
    mark_comparison_clean()

# -------------------------------------------------
# SIDEBAR
# -------------------------------------------------
current_view = st.session_state.get("committed_view", "Comparisons")
current_view_ui = current_view

with st.sidebar:
    render_logo_if_available(FULL_LOGO_PATH, width=180)
    st.markdown("### Account")
    st.success("Logged in")
    st.write(f"User: {get_current_user_email() or get_current_user_id()}")

    if is_admin_user():
        st.success("Admin: Full Access")
    else:
        company = get_current_user_company()

        if company:
            st.success(f"Workspace: {company.get('name')}")
            st.info(f"Seats: {format_company_seats(company)}")
        elif user_row and user_row.get("billing_status") == "active":
            st.success("Plan: Premium")
        else:
            trial_left = trial_days_left(user_row.get("trial_end")) if user_row else 0
            if trial_left > 0:
                st.info(f"Trial: {trial_left} day(s) left")
            else:
                st.warning("Plan: Locked")

    st.markdown("---")
    st.subheader("🧭 Navigation")

    nav_buttons = ["Company Manager", "Sources", "Comparisons"]
    if is_admin_user():
        nav_buttons.append("Admin Panel")

    for i, nav_label in enumerate(nav_buttons):
        button_label = f"• {nav_label}" if current_view == nav_label else nav_label
        if st.button(button_label, use_container_width=True, key=f"sidebar_nav_btn_{nav_label}_{i}"):
            current_view_ui = nav_label

            if current_view_ui == "Comparisons" and current_view == "Comparisons":
                if has_unsaved_comparison_changes() and not st.session_state.get("show_leave_prompt"):
                    st.session_state["show_leave_prompt"] = True
                    st.session_state["leave_prompt_step"] = ""
                    st.session_state["pending_target_view"] = "Comparisons"
                    st.session_state["pending_action_type"] = "switch_view"
                    st.rerun()
                else:
                    st.session_state["committed_view"] = "Comparisons"
                    st.session_state["comparison_mode"] = "menu"
                    st.session_state["show_saved_comparisons"] = False
                    st.session_state["show_inline_save_options"] = False
                    st.session_state["inline_save_mode"] = "menu"
                    st.session_state["active_save_row_id"] = None
                    st.session_state["pending_inline_save_as_name"] = ""
                    st.session_state["pending_save_as_exit_name"] = ""
                    st.rerun()

            if (
                current_view_ui != current_view
                and current_view == "Comparisons"
                and has_unsaved_comparison_changes()
                and not st.session_state.get("show_leave_prompt")
            ):
                st.session_state["show_leave_prompt"] = True
                st.session_state["leave_prompt_step"] = ""
                st.session_state["pending_target_view"] = current_view_ui
                st.session_state["pending_action_type"] = "switch_view"
                st.rerun()
            else:
                previous_committed_view = st.session_state.get("committed_view", current_view_ui)
                st.session_state["committed_view"] = current_view_ui
                if current_view_ui == "Comparisons" and previous_committed_view != "Comparisons":
                    st.session_state["comparison_mode"] = "menu"
                    st.session_state["show_saved_comparisons"] = False
                    st.session_state["show_inline_save_options"] = False
                    st.session_state["inline_save_mode"] = "menu"
                    st.session_state["active_save_row_id"] = None
                    st.session_state["pending_inline_save_as_name"] = ""
                    st.session_state["pending_save_as_exit_name"] = ""
                st.rerun()

    st.markdown("---")
    st.subheader("💳 Billing")

    if not is_admin_user():
        if current_company:
            if current_company.get("billing_status") == "active":
                st.success("Your company subscription is active.")
            else:
                st.info("Your access is managed through your company workspace.")
        elif user_row and user_row.get("billing_status") == "active":
            st.success("Your individual subscription is active.")
        else:
            checkout_url = get_checkout_url(user_email)
            if checkout_url:
                st.link_button(
                    "Individual €10/month",
                    checkout_url,
                    use_container_width=True,
                )

    st.markdown("---")

    if st.button(
        "Logout",
        use_container_width=True,
        key="logout_button",
    ):
        if current_view == "Comparisons" and has_unsaved_comparison_changes():
            st.session_state["show_leave_prompt"] = True
            st.session_state["leave_prompt_step"] = ""
            st.session_state["pending_action_type"] = "logout"
            st.session_state["pending_target_view"] = None
            st.rerun()
        else:
            logout_current_user()

    if not is_admin_user():
        active_sessions_count = len(user_row.get("active_sessions", [])) if user_row else 0
        allowed_sessions_count = get_user_max_active_sessions(user_row)
        remaining_sessions = max(0, allowed_sessions_count - active_sessions_count)
        st.warning("⚠️ Please logout before closing the app to free your session.")
        st.caption(
            f"You already have an active session on another device.\nActive sessions: {active_sessions_count}/{allowed_sessions_count} • Remaining available: {remaining_sessions}"
        )


# -------------------------------------------------
# MAIN UI
# -------------------------------------------------
hero_logo_html = ""
if FULL_LOGO_PATH.exists():
    hero_logo_b64 = base64.b64encode(FULL_LOGO_PATH.read_bytes()).decode("utf-8")
    hero_logo_html = f'<div class="hero-logo-wrap"><img src="data:image/svg+xml;base64,{hero_logo_b64}" /></div>'

st.markdown(
    f"""
    <div class="app-hero">
        {hero_logo_html}
        <div style="font-size:16px;color:#dbeafe;margin-top:8px;">
            Upload supplier sources, compare products and export polished Excel reports.
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)

if st.query_params.get("payment") == "success":
    st.success("Payment completed successfully. Refreshing access may take a few seconds.")
elif st.query_params.get("payment") == "cancel":
    st.info("Checkout was cancelled.")

if st.session_state.get("show_leave_prompt"):
    comparison_label = st.session_state.get("active_comparison_label", "").strip() or st.session_state.get("comparison_name_input", "").strip() or "Untitled Comparison"
    action_type = st.session_state.get("pending_action_type")
    target_view = st.session_state.get("pending_target_view")
    prompt_text = f'Do you want to save the changes you made to Comparison "{comparison_label}"?'
    note_map = {
        "switch_view": f"You are leaving Comparisons and moving to {target_view}." if target_view else "",
        "logout": "You are about to logout.",
        "clear_comparison": "You are about to start a new comparison.",
        "load_comparison": "You are about to load another saved comparison.",
    }

    components.html(
        """
        <script>
        const scrollTopNow = () => {
            try {
                window.parent.scrollTo({top: 0, behavior: 'auto'});
            } catch (e) {}
        };
        requestAnimationFrame(scrollTopNow);
        setTimeout(scrollTopNow, 20);
        </script>
        """,
        height=0,
    )

    st.warning(prompt_text)
    if note_map.get(action_type):
        st.caption(note_map[action_type])

    ask_top_left, ask_top_right = st.columns(2)
    with ask_top_left:
        if st.button("Yes", key="leave_prompt_yes", use_container_width=True):
            st.session_state["leave_prompt_step"] = "save"
            st.rerun()
    with ask_top_right:
        if st.button("No", key="leave_prompt_no", use_container_width=True):
            st.session_state["leave_prompt_step"] = ""
            st.session_state["comparison_dirty"] = False
            st.session_state["comparison_user_modified"] = False
            st.session_state["comparison_clean_generation"] = st.session_state.get("comparison_edit_generation", 0)
            execute_pending_leave_action()
            st.rerun()

    ask_bottom_left, ask_bottom_right = st.columns(2)
    with ask_bottom_left:
        if st.button("Cancel", key="leave_prompt_cancel", use_container_width=True):
            restore_leave_prompt_snapshot()
            st.rerun()

    if st.session_state.get("leave_prompt_step") == "save":
        selected_codes_for_save = get_current_selected_codes_from_state()

        if not selected_codes_for_save:
            if st.session_state.get("current_comparison_id"):
                st.info("This saved comparison has no new unsaved changes right now. Use No to continue to the next section.")
            else:
                st.info("This empty untitled comparison has nothing to save yet. Use No to continue to the next section.")
        else:
            save_c1, save_c2 = st.columns(2)
            with save_c1:
                if st.button("Save", key="leave_prompt_save", use_container_width=True):
                    ok, msg = save_current_comparison_from_state(force_new=False)
                    if ok:
                        st.session_state["comparison_dirty"] = False
                        st.session_state["comparison_user_modified"] = False
                        st.session_state["comparison_clean_generation"] = st.session_state.get("comparison_edit_generation", 0)
                        mark_comparison_clean()
                        execute_pending_leave_action()
                        st.rerun()
                    else:
                        st.warning(msg)
            with save_c2:
                st.text_input("New name for Save As", key="save_as_exit_name")
                if st.button("Save As", key="leave_prompt_save_as", use_container_width=True):
                    new_name = st.session_state.get("save_as_exit_name", "").strip()
                    if not new_name:
                        st.warning("Please enter a new name for Save As.")
                    else:
                        ok, msg = save_current_comparison_from_state(force_new=True, override_name=new_name)
                        if ok:
                            st.session_state["comparison_dirty"] = False
                            st.session_state["comparison_user_modified"] = False
                            st.session_state["comparison_clean_generation"] = st.session_state.get("comparison_edit_generation", 0)
                            mark_comparison_clean()
                            execute_pending_leave_action()
                            st.rerun()
                        else:
                            st.warning(msg)

    st.stop()

# -------------------------------------------------
# SECTION RENDERERS
# -------------------------------------------------


def comparison_has_missing_companies(state_payload, company_options):
    selected_displays = state_payload.get("comparison_company_selection", []) if isinstance(state_payload, dict) else []
    missing = [item for item in selected_displays if item not in company_options]
    return missing


def comparisons_reference_company(comparison_file, company_code, company_name):
    records = list_comparisons(comparison_file)
    for record in records:
        companies = record.get("companies", []) or []
        source_files = record.get("source_files", {}) or {}
        state = record.get("state", {}) or {}

        if company_name in companies:
            return True

        if company_name in source_files:
            return True

        selected_displays = state.get("comparison_company_selection", []) if isinstance(state, dict) else []
        for display in selected_displays:
            if f"({company_code})" in str(display):
                return True

    return False


def render_company_manager():
    st.markdown('<div class="app-card">', unsafe_allow_html=True)
    st.markdown("## 1. Company Manager")

    add_c1, add_c2, add_c3 = st.columns(3)
    with add_c1:
        new_code = st.text_input(
            "Code", key="new_company_code", placeholder="TECHNOGIPS"
        )
    with add_c2:
        new_name = st.text_input(
            "Name", key="new_company_name", placeholder="Technogips"
        )
    with add_c3:
        st.write("")
        st.write("")
        if st.button("Add Company", key="add_company_button", use_container_width=True):
            code = normalize_code(new_code)
            name = str(new_name).strip()

            if not code:
                st.error("Please enter a company code.")
            elif not name:
                st.error("Please enter a company name.")
            elif code in companies_df["code"].astype(str).str.upper().tolist():
                st.warning(f"Company {code} already exists.")
            else:
                updated_df = pd.concat(
                    [companies_df, pd.DataFrame([{"code": code, "name": name}])],
                    ignore_index=True,
                )
                save_companies(updated_df)
                get_company_folder(code)
                st.success(f"Company {name} was added successfully.")
                st.rerun()

    st.dataframe(companies_df, use_container_width=True, hide_index=True)

    st.markdown("### Delete Company")

    company_delete_options = {
        f"{row['name']} ({row['code']})": row["code"] for _, row in companies_df.iterrows()
    }

    del_c1, del_c2 = st.columns([2, 1])
    with del_c1:
        delete_company_display = st.selectbox(
            "Select Company to Delete",
            [""] + list(company_delete_options.keys()),
            key="delete_company_display",
        )
    with del_c2:
        st.write("")
        st.write("")
        if st.button(
            "Delete Company", key="delete_company_button", use_container_width=True
        ):
            if not delete_company_display:
                st.error("Please select a company.")
            else:
                delete_code = company_delete_options[delete_company_display]
                if delete_code in MAIN_CODES:
                    st.warning("Core companies cannot be deleted at this stage.")
                elif company_has_files(delete_code):
                    st.error("This company has source files. Delete the source files first.")
                else:
                    st.session_state["pending_company_delete_code"] = delete_code
                    st.session_state["pending_company_delete_display"] = delete_company_display
                    st.rerun()

    pending_company_delete_code = st.session_state.get("pending_company_delete_code")
    pending_company_delete_display = st.session_state.get("pending_company_delete_display", "")

    if pending_company_delete_code:
        pending_company_name = get_company_label(pending_company_delete_code)
        comparison_file = get_current_user_comparisons_file()
        has_linked_comparisons = comparisons_reference_company(
            comparison_file,
            pending_company_delete_code,
            pending_company_name,
        )

        warning_text = (
            f"Deleting {pending_company_name} ({pending_company_delete_code}) may make saved comparisons that include this company unloadable."
        )
        if has_linked_comparisons:
            warning_text += " Some of your saved comparisons reference this company."

        st.warning(warning_text)

        confirm_c1, confirm_c2 = st.columns(2)

        with confirm_c1:
            if st.button(
                "Confirm Delete Company",
                key="confirm_delete_company_button",
                use_container_width=True,
            ):
                updated_df = companies_df[companies_df["code"] != pending_company_delete_code].copy()
                save_companies(updated_df)
                folder = get_company_folder(pending_company_delete_code)
                try:
                    folder.rmdir()
                except Exception:
                    pass

                st.session_state["pending_company_delete_code"] = None
                st.session_state["pending_company_delete_display"] = ""
                st.success(f"Company {pending_company_delete_code} was deleted.")
                st.rerun()

        with confirm_c2:
            if st.button(
                "Cancel Delete",
                key="cancel_delete_company_button",
                use_container_width=True,
            ):
                st.session_state["pending_company_delete_code"] = None
                st.session_state["pending_company_delete_display"] = ""
                st.info("Company deletion cancelled.")
                st.rerun()
    st.markdown("</div>", unsafe_allow_html=True)


def render_source_library(show_title=True):
    if show_title:
        st.markdown("### 3. Source Library")

    saved_df = list_saved_sources()
    if not saved_df.empty:
        st.dataframe(
            saved_df.drop(columns=["Full Path"]),
            use_container_width=True,
            hide_index=True,
        )
    else:
        st.info("No saved source files yet.")

    st.markdown("### Delete Source")

    source_delete_options = {}
    if not saved_df.empty:
        for _, row in saved_df.iterrows():
            label = f"{row['Company Name']} | {row['Filename']}"
            source_delete_options[label] = row["Full Path"]

    src_d1, src_d2 = st.columns([3, 1])
    with src_d1:
        delete_source_display = st.selectbox(
            "Select Source to Delete",
            [""] + list(source_delete_options.keys()),
            key="delete_source_display",
        )
    with src_d2:
        st.write("")
        st.write("")
        if st.button(
            "Delete Source", key="delete_source_button", use_container_width=True
        ):
            if not delete_source_display:
                st.error("Please select a source.")
            else:
                st.session_state["pending_source_delete_display"] = delete_source_display

    pending_delete_display = st.session_state.get("pending_source_delete_display", "")
    if pending_delete_display:
        pending_path = Path(source_delete_options.get(pending_delete_display, "")) if pending_delete_display in source_delete_options else None
        if pending_path and pending_path.exists():
            impacted = _find_comparisons_using_source(pending_path.name)
            impacted_names = [item["record"].get("name", "Untitled comparison") for item in impacted]

            st.warning(
                f"Soft delete will move the selected source to Trash and also remove {len(impacted)} comparison(s) that use it from the active list."
            )
            if impacted_names:
                st.caption("Affected comparisons: " + " • ".join(impacted_names[:8]))

            confirm_delete = st.checkbox(
                "I understand that the source and its related comparisons will be moved out of the active workspace.",
                key="confirm_soft_delete_source",
            )

            confirm_c1, confirm_c2 = st.columns([1, 1])
            with confirm_c1:
                if st.button("Confirm Soft Delete", key="confirm_soft_delete_button", use_container_width=True):
                    if not confirm_delete:
                        st.error("Please confirm before continuing.")
                    else:
                        result = _soft_delete_source_and_related_comparisons(pending_path)
                        st.session_state.pop("pending_source_delete_display", None)
                        st.session_state.pop("confirm_soft_delete_source", None)
                        st.success(
                            f"Moved source to Trash: {pending_path.name}. "
                            f"Soft-deleted {result['deleted_comparisons_count']} related comparison(s)."
                        )
                        refresh_source_file_views()
                        st.rerun()

            with confirm_c2:
                if st.button("Cancel", key="cancel_soft_delete_button", use_container_width=True):
                    st.session_state.pop("pending_source_delete_display", None)
                    st.session_state.pop("confirm_soft_delete_source", None)
                    st.rerun()
        else:
            st.session_state.pop("pending_source_delete_display", None)

def _source_generator_output_columns():
    return ["SAP", "Product", "Base Price", "Price", "MM", "Package", "Category"]


def _source_dataframe_to_excel_bytes(df: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    export_df = df.copy()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df.to_excel(writer, sheet_name="PRICELIST", index=False)
    output.seek(0)
    return output.getvalue()


def _normalize_text_simple(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"", "nan", "none"}:
        return ""
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _is_all_capsish(text: str) -> bool:
    letters = [ch for ch in text if ch.isalpha()]
    if not letters:
        return False
    upper_ratio = sum(1 for ch in letters if ch.isupper()) / len(letters)
    return upper_ratio >= 0.65


def _looks_like_section_title(values) -> bool:
    cleaned = [_normalize_text_simple(v) for v in values]
    cleaned = [x for x in cleaned if x]
    if not cleaned:
        return False
    if len(cleaned) == 1:
        only = cleaned[0]
        if _to_float_or_none(only) is not None:
            return False
        return len(only) <= 140 and (_is_all_capsish(only) or len(only.split()) <= 8)
    numeric = sum(1 for x in cleaned if _to_float_or_none(x) is not None)
    return numeric == 0 and len(cleaned) <= 2


def _read_raw_sheet(file_bytes: bytes, sheet_name: str) -> pd.DataFrame:
    return pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, header=None, engine="openpyxl")


def _extract_category_map(file_bytes: bytes):
    category_map = {}
    try:
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name="Categorie", engine="openpyxl")
        if df is not None and not df.empty and len(df.columns) >= 1:
            cat_col = df.columns[0]
            prc_col = df.columns[1] if len(df.columns) > 1 else None
            for _, row in df.iterrows():
                key = _normalize_text_simple(row.get(cat_col))
                if not key:
                    continue
                category_map[key] = _to_increase_fraction(row.get(prc_col)) if prc_col is not None else 0.0
    except Exception:
        pass
    return category_map


def _is_siniat_april_workbook(sheet_names):
    normalized = {str(s).strip().lower() for s in sheet_names}
    needed = {"σανίδες nida (placa nida)", " αξεσουάρ (accesorii)", " προφίλ (profile)", "κονιάματα (adera)"}
    return any("placa nida" in s for s in normalized) and any("accesorii" in s for s in normalized)


def _safe_row_value(row, idx):
    if idx >= len(row):
        return None
    return row[idx]


def _append_if_present(base, extra, sep=" | "):
    base = _normalize_text_simple(base)
    extra = _normalize_text_simple(extra)
    if base and extra:
        return f"{base}{sep}{extra}"
    return base or extra


def _build_package_from_area(unit, package_desc, width_mm=None, length_mm=None):
    package_desc = _normalize_text_simple(package_desc)
    unit = _normalize_text_simple(unit)
    if not package_desc:
        return ""
    m = re.search(r"(\d+(?:[.,]\d+)?)", package_desc.replace(" ", ""))
    if unit.lower() == "m2" and m and width_mm and length_mm:
        try:
            count = float(m.group(1).replace(",", "."))
            area_each = (float(width_mm) / 1000.0) * (float(length_mm) / 1000.0)
            total = round(count * area_each, 2)
            return f"{total:g} m²/παλέτα"
        except Exception:
            return package_desc
    if unit in {"Μ", "m", "M"} and m and length_mm:
        try:
            count = float(m.group(1).replace(",", "."))
            total = round(count * (float(length_mm) / 1000.0), 2)
            total_str = str(int(total)) if abs(total - int(total)) < 1e-9 else str(total)
            return f"{total_str} m/δέμα"
        except Exception:
            return package_desc
    if "τεμ" in package_desc.lower() and "/" not in package_desc:
        return package_desc
    return package_desc


def _parse_siniat_boards(file_bytes, sheet_name):
    raw = _read_raw_sheet(file_bytes, sheet_name)
    rows = []
    current_category = ""
    current_product = ""
    for i in range(1, len(raw)):
        vals = raw.iloc[i].tolist()
        product = _normalize_text_simple(_safe_row_value(vals, 0))
        sap = _normalize_text_simple(_safe_row_value(vals, 1))
        width = _safe_row_value(vals, 2)
        length = _safe_row_value(vals, 3)
        mm = _normalize_text_simple(_safe_row_value(vals, 4))
        package_desc = _normalize_text_simple(_safe_row_value(vals, 6))
        base_price = _to_float_or_none(_safe_row_value(vals, 10))
        category = _normalize_text_simple(_safe_row_value(vals, 11))
        inc = _to_increase_fraction(_safe_row_value(vals, 12))

        if _looks_like_section_title(vals[:3]):
            title = product
            if title:
                current_category = title
            continue

        if product:
            current_product = product
        elif not sap and base_price is None:
            continue

        final_product = current_product
        if width and length and _to_float_or_none(width) is not None and _to_float_or_none(length) is not None:
            final_product = _append_if_present(final_product, f"{int(float(width))}x{int(float(length))}mm")

        if not final_product:
            continue

        final_category = category or current_category or "Placa Nida"
        rows.append({
            "SAP": sap,
            "Product": final_product,
            "Base Price": base_price,
            "Increase %": 0.0,
            "Price": base_price,
            "MM": mm,
            "Package": _build_package_from_area(mm, package_desc, width, length),
            "Category": final_category,
        })
    return rows


def _parse_siniat_accessories(file_bytes, sheet_name):
    raw = _read_raw_sheet(file_bytes, sheet_name)
    rows = []
    current_category = ""
    for i in range(1, len(raw)):
        vals = raw.iloc[i].tolist()
        product = _normalize_text_simple(_safe_row_value(vals, 0))
        sap = _normalize_text_simple(_safe_row_value(vals, 2))
        mm = _normalize_text_simple(_safe_row_value(vals, 3))
        package_desc = _normalize_text_simple(_safe_row_value(vals, 5))
        base_price = _to_float_or_none(_safe_row_value(vals, 9))
        category = _normalize_text_simple(_safe_row_value(vals, 10))
        inc = _to_increase_fraction(_safe_row_value(vals, 11))

        if _looks_like_section_title(vals[:3]):
            if product:
                current_category = product
            continue

        if not product and not sap and base_price is None:
            continue

        rows.append({
            "SAP": sap,
            "Product": product,
            "Base Price": base_price,
            "Increase %": 0.0,
            "Price": base_price,
            "MM": mm,
            "Package": package_desc,
            "Category": category or current_category or "Accesorii",
        })
    return rows


def _parse_siniat_profiles(file_bytes, sheet_name):
    raw = _read_raw_sheet(file_bytes, sheet_name)
    rows = []
    current_category = ""
    for i in range(1, len(raw)):
        vals = raw.iloc[i].tolist()
        product = _normalize_text_simple(_safe_row_value(vals, 0))
        sap = _normalize_text_simple(_safe_row_value(vals, 2))
        length = _safe_row_value(vals, 3)
        mm = _normalize_text_simple(_safe_row_value(vals, 4))
        package_desc = _normalize_text_simple(_safe_row_value(vals, 6))
        base_price = _to_float_or_none(_safe_row_value(vals, 10))
        category = _normalize_text_simple(_safe_row_value(vals, 11))
        inc = _to_increase_fraction(_safe_row_value(vals, 12))

        if _looks_like_section_title(vals[:3]):
            if product:
                current_category = product
            continue

        if not product and not sap and base_price is None:
            continue

        final_product = product
        if length and _to_float_or_none(length) is not None:
            final_product = _append_if_present(final_product, f"{int(float(length))}mm")
        rows.append({
            "SAP": sap,
            "Product": final_product,
            "Base Price": base_price,
            "Increase %": 0.0,
            "Price": base_price,
            "MM": mm,
            "Package": _build_package_from_area(mm, package_desc, None, length),
            "Category": category or current_category or "Profile Nida",
        })
    return rows


def _parse_siniat_mortars(file_bytes, sheet_name):
    raw = _read_raw_sheet(file_bytes, sheet_name)
    rows = []
    current_category = ""
    current_product = ""
    for i in range(1, len(raw)):
        vals = raw.iloc[i].tolist()
        product = _normalize_text_simple(_safe_row_value(vals, 0))
        sap = _normalize_text_simple(_safe_row_value(vals, 1))
        qty = _normalize_text_simple(_safe_row_value(vals, 2))
        delivery = _normalize_text_simple(_safe_row_value(vals, 3))
        mm = _safe_row_value(vals, 4)
        base_price = _to_float_or_none(_safe_row_value(vals, 6))
        category = _normalize_text_simple(_safe_row_value(vals, 8))

        if _looks_like_section_title(vals[:2]):
            if product:
                current_category = product
            continue

        if product:
            current_product = product
        elif not sap and base_price is None:
            continue

        final_product = _append_if_present(current_product, qty)
        rows.append({
            "SAP": sap,
            "Product": final_product,
            "Base Price": base_price if base_price is not None else 0.0,
            "Increase %": 0.0,
            "Price": base_price if base_price is not None else 0.0,
            "MM": mm,
            "Package": delivery,
            "Category": category or current_category or "Adera",
        })
    return rows


def _parse_siniat_trape(file_bytes, sheet_name):
    raw = _read_raw_sheet(file_bytes, sheet_name)
    rows = []
    current_category = ""
    current_product = ""
    for i in range(1, len(raw)):
        vals = raw.iloc[i].tolist()
        product = _normalize_text_simple(_safe_row_value(vals, 0))
        sap = _normalize_text_simple(_safe_row_value(vals, 1))
        thickness = _normalize_text_simple(_safe_row_value(vals, 2))
        dims = _normalize_text_simple(_safe_row_value(vals, 3))
        mm = _normalize_text_simple(_safe_row_value(vals, 4))
        base_price = _to_float_or_none(_safe_row_value(vals, 9))
        category = _normalize_text_simple(_safe_row_value(vals, 10))
        inc = _to_increase_fraction(_safe_row_value(vals, 11))

        if _looks_like_section_title(vals[:3]):
            if product:
                current_category = product
            continue

        if product:
            current_product = product
        elif not sap and base_price is None:
            continue

        final_product = current_product
        suffix = " ".join([x for x in [dims, thickness] if x]).strip()
        if suffix:
            final_product = _append_if_present(final_product, suffix)
        rows.append({
            "SAP": sap,
            "Product": final_product,
            "Base Price": base_price,
            "Increase %": 0.0,
            "Price": base_price,
            "MM": mm,
            "Package": "1 τεμ" if mm else "",
            "Category": category or current_category or "Trape",
        })
    return rows


def _parse_siniat_services(file_bytes, sheet_name):
    raw = _read_raw_sheet(file_bytes, sheet_name)
    rows = []
    for i in range(1, len(raw)):
        vals = raw.iloc[i].tolist()
        product = _normalize_text_simple(_safe_row_value(vals, 0))
        sap = _normalize_text_simple(_safe_row_value(vals, 1))
        mm = _normalize_text_simple(_safe_row_value(vals, 2))
        package = _normalize_text_simple(_safe_row_value(vals, 3))
        base_price = _to_float_or_none(_safe_row_value(vals, 5))
        category = _normalize_text_simple(_safe_row_value(vals, 6)) or "Servicii"
        inc = _to_increase_fraction(_safe_row_value(vals, 7))
        if not product:
            continue
        rows.append({
            "SAP": sap,
            "Product": product,
            "Base Price": base_price,
            "Increase %": 0.0,
            "Price": base_price,
            "MM": mm,
            "Package": package,
            "Category": category,
        })
    return rows


def _parse_siniat_workbook(xls, file_bytes):
    all_rows = []
    used_sheets = []
    skipped_sheets = []
    for sheet_name in xls.sheet_names:
        stripped = str(sheet_name).strip()
        low = stripped.lower()
        try:
            if low == "categorie" or low == "extra":
                skipped_sheets.append(f"{sheet_name} (helper)")
                continue
            if "placa nida" in low:
                rows = _parse_siniat_boards(file_bytes, sheet_name)
            elif "accesorii" in low:
                rows = _parse_siniat_accessories(file_bytes, sheet_name)
            elif "profile" in low:
                rows = _parse_siniat_profiles(file_bytes, sheet_name)
            elif "adera" in low:
                rows = _parse_siniat_mortars(file_bytes, sheet_name)
            elif "trape" in low:
                rows = _parse_siniat_trape(file_bytes, sheet_name)
            elif "υπηρεσίες" in low and "servicii" in low:
                rows = _parse_siniat_services(file_bytes, sheet_name)
            elif low == "servicii":
                skipped_sheets.append(f"{sheet_name} (legacy services skipped)")
                continue
            else:
                skipped_sheets.append(f"{sheet_name} (unmapped)")
                continue
        except Exception as e:
            skipped_sheets.append(f"{sheet_name} (parse error)")
            continue

        if rows:
            all_rows.extend(rows)
            used_sheets.append(sheet_name)
        else:
            skipped_sheets.append(f"{sheet_name} (no valid product rows)")

    if not all_rows:
        return None, {"used_sheets": used_sheets, "skipped_sheets": skipped_sheets, "missing_price_rows": 0, "missing_sap_rows": 0, "missing_product_rows": 0, "detected_section_titles": [], "total_rows": 0}

    source_df = pd.DataFrame(all_rows)
    source_df["Base Price"] = pd.to_numeric(source_df["Base Price"], errors="coerce")
    source_df["Price"] = source_df["Base Price"]
    source_df = source_df[_source_generator_output_columns()].reset_index(drop=True)

    stats = {
        "used_sheets": used_sheets,
        "skipped_sheets": skipped_sheets,
        "missing_price_rows": int(source_df["Base Price"].isna().sum()),
        "missing_sap_rows": int(source_df["SAP"].astype(str).str.strip().eq("").sum()),
        "missing_product_rows": int(source_df["Product"].astype(str).str.strip().eq("").sum()),
        "detected_section_titles": [],
        "total_rows": len(source_df),
    }
    return source_df, stats


def _detect_supplier_header_row(raw_df: pd.DataFrame) -> int:
    max_scan = min(len(raw_df), 25)
    best_row = 0
    best_score = -1

    code_tokens = ["sap", "code", "item code", "item no", "article", "article no", "sku", "reference", "ref", "material", "κωδικ", "κωδ"]
    product_tokens = ["description", "product", "name", "item", "material description", "περιγραφ", "προϊ", "προιο", "όνομα", "ονομα"]
    price_tokens = ["price", "list", "net price", "catalog", "unit price", "τιμή", "τιμο", "value", "αξία", "€/"]
    unit_tokens = ["unit", "uom", "mm", "μον", "μ.μ", "μ/μ"]
    package_tokens = ["pack", "package", "packing", "minimum", "pallet", "box", "συσκ"]

    for idx in range(max_scan):
        values = [str(v).strip().lower() for v in raw_df.iloc[idx].tolist() if str(v).strip() not in ["", "nan", "none"]]
        score = 0
        if any(any(token in v for token in code_tokens) for v in values):
            score += 2
        if any(any(token in v for token in product_tokens) for v in values):
            score += 3
        if any(any(token in v for token in price_tokens) for v in values):
            score += 4
        if any(any(token in v for token in unit_tokens) for v in values):
            score += 1
        if any(any(token in v for token in package_tokens) for v in values):
            score += 1
        if len(values) >= 3:
            score += 0.5
        if score > best_score:
            best_score = score
            best_row = idx

    return best_row


def _normalize_supplier_column_name(col_name: str) -> str:
    c = str(col_name).strip().lower()

    if any(token in c for token in [
        "sap", "item code", "item no", "product code", "article", "article no", "sku",
        "reference", "ref", "material code", "material no", "κωδικ", "κωδ.", "κωδ", "code"
    ]):
        return "SAP"
    if any(token in c for token in [
        "description", "product", "item description", "product description", "name",
        "material description", "material", "περιγραφ", "προϊ", "προιο", "όνομα", "ονομα"
    ]):
        return "Product"
    if any(token in c for token in ["increase %", "increase", "diff%", "markup", "adjustment %", "delta %", "ανατ", "αύξη", "αυξη"]):
        return "Increase %"
    if any(token in c for token in [
        "base price", "list price", "net price", "catalog price", "price", "τιμή", "τιμο",
        "pricelist", "unit price", "value", "αξία", "€/", "eur"
    ]):
        return "Base Price"
    if any(token in c for token in ["unit of measure", "uom", "unit", " mm", "mm ", "μον", "μ.μ", "μ/μ"]) or c == "mm":
        return "MM"
    if any(token in c for token in ["package", "pack", "packing", "minimum order", "min order", "pallet", "box", "συσκ"]):
        return "Package"
    if any(token in c for token in ["category", "group", "family", "range", "line", "segment", "κατηγο"]):
        return "Category"

    return str(col_name).strip()


def _guess_supplier_columns_by_values(df: pd.DataFrame) -> dict:
    guessed = {}
    used_columns = set()

    def _sample(series):
        vals = []
        if isinstance(series, pd.DataFrame):
            if series.shape[1] == 0:
                return vals
            series = series.iloc[:, 0]
        for val in series.astype(object).tolist():
            if val is None:
                continue
            sval = str(val).strip()
            if sval.lower() in {"", "nan", "none"}:
                continue
            vals.append(sval)
            if len(vals) >= 30:
                break
        return vals

    def _numeric_ratio(values):
        if not values:
            return 0.0
        ok = 0
        for v in values:
            if _to_float_or_none(v) is not None:
                ok += 1
        return ok / len(values)

    def _long_text_ratio(values):
        if not values:
            return 0.0
        ok = 0
        for v in values:
            if len(v) >= 8 and _to_float_or_none(v) is None:
                ok += 1
        return ok / len(values)

    def _code_like_ratio(values):
        if not values:
            return 0.0
        ok = 0
        for v in values:
            compact = v.replace(" ", "")
            if len(compact) <= 20 and any(ch.isdigit() for ch in compact):
                ok += 1
        return ok / len(values)

    profiles = {}
    for col in df.columns:
        vals = _sample(df[col])
        profiles[col] = {
            "numeric_ratio": _numeric_ratio(vals),
            "long_text_ratio": _long_text_ratio(vals),
            "code_like_ratio": _code_like_ratio(vals),
        }

    price_candidates = sorted(df.columns, key=lambda c: (profiles[c]["numeric_ratio"], "price" in str(c).lower()), reverse=True)
    for col in price_candidates:
        if profiles[col]["numeric_ratio"] >= 0.55:
            guessed["Base Price"] = col
            used_columns.add(col)
            break

    product_candidates = sorted(df.columns, key=lambda c: (profiles[c]["long_text_ratio"], len(str(c))), reverse=True)
    for col in product_candidates:
        if col in used_columns:
            continue
        if profiles[col]["long_text_ratio"] >= 0.35:
            guessed["Product"] = col
            used_columns.add(col)
            break

    code_candidates = sorted(df.columns, key=lambda c: (profiles[c]["code_like_ratio"], -profiles[c]["numeric_ratio"]), reverse=True)
    for col in code_candidates:
        if col in used_columns:
            continue
        if profiles[col]["code_like_ratio"] >= 0.35:
            guessed["SAP"] = col
            used_columns.add(col)
            break

    return guessed


def _to_float_or_none(value):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, (int, float)) and not pd.isna(value):
        return float(value)

    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "-", "--", "upon request", "contact us"}:
        return None

    text = text.replace("€", "").replace("EUR", "").replace("eur", "")
    text = text.replace(" ", "")

    if text.count(",") > 0 and text.count(".") > 0:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", ".")

    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None

    try:
        return float(match.group(0))
    except Exception:
        return None


def _to_increase_fraction(value):
    num = _to_float_or_none(value)
    if num is None:
        return 0.0
    return num / 100.0 if abs(num) > 1 else num



def _clean_header_part(value):
    if value is None:
        return ""
    s = str(value).strip()
    if s.lower() in {"", "nan", "none"}:
        return ""
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def _build_supplier_header_names(raw_df: pd.DataFrame, header_row: int, depth: int = 3):
    if raw_df is None or raw_df.empty:
        return []

    header_end = min(len(raw_df), header_row + depth)
    header_block = raw_df.iloc[header_row:header_end].copy()
    if header_block.empty:
        return []

    header_block = header_block.ffill(axis=1).ffill(axis=0)

    names = []
    for col_idx in range(header_block.shape[1]):
        parts = []
        seen = set()
        for row_idx in range(header_block.shape[0]):
            part = _clean_header_part(header_block.iat[row_idx, col_idx])
            if not part:
                continue
            low = part.lower()
            if low not in seen:
                seen.add(low)
                parts.append(part)
        joined = " | ".join(parts).strip()
        names.append(joined if joined else f"Column {col_idx + 1}")
    return names


def _normalize_supplier_dataframe_from_raw(raw_df: pd.DataFrame, header_row: int):
    header_names = _build_supplier_header_names(raw_df, header_row=header_row, depth=3)
    if not header_names:
        return pd.DataFrame()

    data_start = min(len(raw_df), header_row + 3)
    body = raw_df.iloc[data_start:].copy().reset_index(drop=True)
    if body.empty:
        body = raw_df.iloc[header_row + 1:].copy().reset_index(drop=True)
    if body.empty:
        return pd.DataFrame()

    if body.shape[1] > len(header_names):
        header_names += [f"Column {i + 1}" for i in range(len(header_names), body.shape[1])]
    body.columns = header_names[:body.shape[1]]
    body = body.dropna(how="all").reset_index(drop=True)
    return body


def _is_section_title_row(values):
    cleaned = []
    for v in values:
        if v is None:
            continue
        s = str(v).strip()
        if s.lower() in {"", "nan", "none"}:
            continue
        cleaned.append(s)

    if not cleaned:
        return False

    numeric_count = sum(1 for s in cleaned if _to_float_or_none(s) is not None)

    if len(cleaned) == 1:
        s = cleaned[0]
        if _to_float_or_none(s) is not None:
            return False
        alpha_chars = [ch for ch in s if ch.isalpha()]
        upper_ratio = (sum(1 for ch in alpha_chars if ch.isupper()) / len(alpha_chars)) if alpha_chars else 0.0
        return len(s) <= 100 and (upper_ratio >= 0.6 or len(s.split()) <= 6)

    if numeric_count == 0 and len(cleaned) <= 2:
        merged = " ".join(cleaned)
        alpha_count = sum(1 for ch in merged if ch.isalpha())
        if alpha_count >= 3 and len(merged) <= 100:
            return True

    return False


def _extract_section_title(values):
    cleaned = []
    for v in values:
        if v is None:
            continue
        s = str(v).strip()
        if s.lower() in {"", "nan", "none"}:
            continue
        cleaned.append(s)
    return " | ".join(cleaned[:2]).strip()




class _PdfParseState:
    def __init__(self):
        self.category = ''
        self.family = ''
        self.last_table_like = False

    def reset_family(self):
        self.family = ''

    def reset_all(self):
        self.category = ''
        self.family = ''
        self.last_table_like = False


def _looks_like_pdf_family_title(line: str) -> bool:
    line = _normalize_text_simple(line)
    if not line or len(line) > 100:
        return False
    low = line.lower()
    if any(k in low for k in ['κωδικ', 'τιμη', 'sap code', 'list price', 'περιεχομενα']):
        return False
    if _looks_like_pdf_product_code(line):
        return False
    if _to_float_or_none(line) is not None:
        return False
    if re.search(r'\d+[.,]?\d*\s*(kg|g|gr|lt|l|ml|mm)', low):
        return True
    if any(b in low for b in ['nida', 'sika', 'sikaflex', 'isomat', 'aquamat', 'aquasil', 'sikalastic', 'ladura', 'pregy', 'creason', 'createx', 'resistex']):
        return True
    words = line.split()
    return 1 <= len(words) <= 6 and any(ch.isalpha() for ch in line) and not any(ch in line for ch in ['€', '%'])


def _looks_like_money_token(token: str) -> bool:
    raw = _normalize_text_simple(token)
    if not raw:
        return False
    low = raw.lower()
    if 'x' in low:
        return False
    if any(u in low for u in ['mm', 'cm', 'kg/um', 'kg /u', 'boards/pallet', 'σανιδ', 'παλετ']):
        return False
    return bool(re.fullmatch(r'-?\d+[.,]\d{1,4}\s*€?', raw))


def _extract_last_valid_price_from_values(values):
    cleaned = [_normalize_text_simple(v) for v in values]
    candidates = []
    for idx, value in enumerate(cleaned):
        if not _looks_like_money_token(value):
            continue
        num = _to_float_or_none(_clean_pdf_price_text(value))
        if num is None or not (0 < num < 100000):
            continue
        candidates.append((idx, num))
    if not candidates:
        return None, None
    return candidates[-1]


def _looks_like_pdf_noise_page_relaxed(text: str) -> bool:
    s = _normalize_text_simple(text).lower()
    if not s:
        return True
    if 'sap code' in s or 'κωδικ' in s or 'list price' in s or 'τιμη πωλησης' in s or 'τιμη' in s:
        return False
    if _looks_like_pdf_contents_page(s):
        return True
    if _looks_like_pdf_marketing_page(s):
        return True
    paragraph_lines = [ln for ln in s.splitlines() if len(ln.split()) > 10]
    numeric_hits = len(re.findall(r'\d+[.,]\d+|\d{4,8}', s))
    unit_hits = len(re.findall(r'(?:m2|m²|kg|lt|ml|mm|cm|τεμ|pcs?)', s))
    return len(paragraph_lines) >= 8 and numeric_hits < 6 and unit_hits < 3


def _parse_pdf_line_stateful(line: str, state: _PdfParseState, company_hint: str = ''):
    txt = _normalize_text_simple(line)
    if not txt:
        return None
    low = txt.lower()
    if any(k in low for k in ['κωδικ', 'sap code', 'list price', 'τιμη πωλησης', 'packaging', 'width', 'length', 'μονάδα μέτρησης', 'delivery conditions', 'valid from', 'table of contents']):
        return None
    if _looks_like_pdf_section_title(txt):
        state.category = txt
        state.reset_family()
        return {'__section_title__': txt}
    if _looks_like_pdf_family_title(txt):
        state.family = txt
        return {'__family_title__': txt}

    tokens = txt.split()
    code = ''
    code_idx = None
    for i, tok in enumerate(tokens[:8]):
        if _looks_like_pdf_product_code(tok):
            code = tok
            code_idx = i
            break
    if code_idx is None:
        return None

    price_idx, price = _extract_last_valid_price_from_values(tokens)
    if price_idx is None:
        price_idx = len(tokens)

    before_code = tokens[:code_idx]
    after_code = tokens[code_idx + 1:price_idx]

    width = ''
    length = ''
    unit = ''
    packaging = ''
    weight = ''
    delivery = ''
    name_tokens = []

    numeric_dims = []
    for tok in after_code:
        tok_norm = tok.replace('*', '')
        if re.fullmatch(r'\d{3,4}', tok_norm):
            numeric_dims.append(tok_norm)
            continue
        low_tok = tok_norm.lower()
        if not unit and low_tok in {'m2', 'm²', 'm', 'pcs', 'pc', 'τεμ'}:
            unit = tok_norm.replace('m²', 'm2')
            continue
        if not packaging and any(x in low_tok for x in ['παλέτ', 'pallet', 'box', 'bucket', 'bag', 'δοχεί', 'σακ', 'boards/pallet', 'σανίδες/παλέτα']):
            packaging = tok_norm
            continue
        if not weight and re.fullmatch(r'\d+[.,]\d{1,2}', tok_norm) and unit:
            weight = tok_norm
            continue
        if not delivery and tok_norm in {'A', '-'}:
            delivery = tok_norm
            continue
        name_tokens.append(tok)

    if numeric_dims:
        width = numeric_dims[0]
    if len(numeric_dims) > 1:
        length = numeric_dims[1]

    product_name = _normalize_text_simple(' '.join(before_code + name_tokens))
    if not product_name:
        product_name = state.family or ''
    if not product_name:
        return None

    notes = []
    if width and length:
        notes.append(f'{width}x{length}')
    elif length:
        notes.append(f'Length {length}')
    if weight:
        notes.append(f'Weight {weight}')
    if delivery:
        notes.append(f'Delivery {delivery}')

    mm_text = unit or _extract_mm_from_text(txt)
    package_text = _extract_package_from_text(txt) or packaging
    category_text = state.category or 'PDF Catalog'
    return {
        'confidence': 0.82 if price is not None else 0.68,
        'SAP': code,
        'Product': product_name,
        'Base Price': price,
        'Increase %': 0.0,
        'Price': price,
        'MM': mm_text,
        'Package': package_text,
        'Category': category_text,
        'Notes': _normalize_text_simple(' | '.join([n for n in notes if n])),
        'Company': company_hint or '',
    }


def _looks_like_pdf_section_title(line: str) -> bool:
    line = _normalize_text_simple(line)
    if not line or len(line) > 140:
        return False
    low = line.lower()
    if _to_float_or_none(line) is not None:
        return False
    if re.search(r'\b\d+[.,]\d+\b', low):
        return False
    hard_markers = [
        'price list', 'catalog', 'table of contents', 'contents', 'delivery conditions',
        'valid from', 'recommended retail', 'technical data', 'product range', 'system solutions',
        'installation', 'classification', 'overview', 'description', 'application', 'packaging',
        'τιμοκαταλογος', 'περιεχομενα', 'ισχυς απο', 'προυποθεσεις', 'συσκευασια',
        'τεχνικα χαρακτηριστικα', 'εφαρμογη', 'περιγραφη', 'λυσεις συστηματων',
        'γυψοσανιδες για', 'γυψοπλακες για', 'οι τιμες ειναι', 'χωρις φπα'
    ]
    if any(m in low for m in hard_markers):
        return True
    letters = [ch for ch in line if ch.isalpha()]
    if len(letters) < 3:
        return False
    upper_ratio = sum(1 for ch in letters if ch.isupper()) / max(len(letters), 1)
    words = line.split()
    return upper_ratio >= 0.75 or (len(words) <= 10 and _is_all_capsish(line) and not any(ch.isdigit() for ch in line))

def _extract_package_from_text(text: str) -> str:
    s = _normalize_text_simple(text).lower()
    if not s:
        return ''
    m = re.search(r'(\d+(?:[.,]\d+)?)\s*(kg|gr|g|lt|l|ml)\b', s)
    if not m:
        return ''
    qty = m.group(1).replace(',', '.')
    unit = m.group(2)
    if unit == 'gr':
        unit = 'g'
    if unit == 'lt':
        unit = 'l'
    qty = qty.rstrip('0').rstrip('.') if '.' in qty else qty
    return f"{qty}{unit}"


def _extract_mm_from_text(text: str) -> str:
    s = _normalize_text_simple(text).lower()
    if not s:
        return ''
    patterns = [
        r'\b(m2|m²|m3|m³|mm|cm|τεμ|τεμ\.|pcs|pc|m)\b',
    ]
    for pattern in patterns:
        m = re.search(pattern, s)
        if m:
            val = m.group(1)
            return val.replace('m²', 'm2').replace('m³', 'm3')
    return ''




def _looks_like_pdf_contents_page(text: str) -> bool:
    s = _normalize_text_simple(text).lower()
    if not s:
        return False
    markers = [
        'περιεχομενα', 'price list 2021', 'price list 2025', 'price list 2026', 'delivery term', 'valid from',
        'table of contents', 'contents', 'environmental product declaration', 'building performance',
        'ισχυρη παρουσια', 'πιστοποιημενο συστημα', 'phototimokatalogos', 'φωτοτιμοκαταλογος',
        'timo katalogos', 'recommended retail prices', 'catalogue', 'index'
    ]
    if any(m in s for m in markers):
        if 'sap code' in s or 'κωδικ' in s or 'list price' in s or 'τιμη' in s:
            return False
        return True
    return False

def _looks_like_pdf_marketing_page(text: str) -> bool:
    s = _normalize_text_simple(text).lower()
    if not s:
        return False
    if 'sap code' in s or 'κωδικ' in s or 'list price' in s or 'τιμη' in s:
        return False
    marketing_markers = [
        'learn more', 'μαθετε περισσοτερα', 'environmental product declaration',
        'η πρωτη ελληνικη εταιρεια', 'παθος για καινοτομια', 'ισχυρη παρουσια',
        'trust in every bond', 'home beauty', 'smart clean paint', 'road to sustainability',
        'inspiring ways of living', 'our brands', 'our main brands', 'δρομος προς την αειφορια',
        'discover more', 'sustainability', 'corporate profile'
    ]
    return any(m in s for m in marketing_markers)

def _looks_like_pdf_product_code(value: str) -> bool:
    v = _normalize_text_simple(value).replace(' ', '')
    if not v or len(v) > 16 or v in {'-', '—'}:
        return False
    if any('Ͱ' <= ch <= 'Ͽ' for ch in v):
        return False
    if not any(ch.isdigit() for ch in v):
        return False
    # Prefer true SAP-like numeric codes, but allow short structured alphanumeric SKUs such as CW75/UA50.
    if re.fullmatch(r'\d{5,8}', v):
        return True
    if re.fullmatch(r'[A-Z]{1,4}\d{1,5}[A-Z0-9./_-]{0,4}', v):
        return True
    return False

def _parse_pdf_table_variants(table, current_category: str = '', company_hint: str = ''):
    rows = []
    active_name = ''
    active_category = current_category or 'PDF Catalog'
    active_width = ''
    active_unit = ''
    active_packaging_uom = ''
    active_packaging = ''
    active_weight_sale = ''
    active_price = None

    for raw_row in table or []:
        original = raw_row or []
        cleaned = [_normalize_text_simple(v) for v in original]
        cleaned = [v if v is not None else '' for v in cleaned]
        if not any(cleaned):
            continue
        joined = ' '.join([v for v in cleaned if v]).strip()
        low = joined.lower()
        if not joined:
            continue
        if any(k in low for k in ['sap code', 'list price', 'κωδικ', 'τιμη', 'packaging', 'length (mm)', 'width (mm)']):
            continue
        if _looks_like_pdf_section_title(joined) and not any(_looks_like_pdf_product_code(v) for v in cleaned):
            active_category = joined
            continue

        code = ''
        name = ''
        width = ''
        length = ''
        unit = ''
        packaging_uom = ''
        packaging = ''
        weight_sale = ''
        price = None

        for idx, val in enumerate(cleaned[:3]):
            if _looks_like_pdf_product_code(val):
                code = val
                if idx > 0 and cleaned[0]:
                    name = cleaned[0]
                break

        if len(cleaned) >= 9:
            name = name or cleaned[0]
            code = code or cleaned[1]
            width = cleaned[2]
            length = cleaned[3]
            unit = cleaned[4]
            packaging_uom = cleaned[5]
            packaging = cleaned[6]
            weight_sale = cleaned[7]
            price = _to_float_or_none(_clean_pdf_price_text(cleaned[8]))
        elif len(cleaned) >= 6 and code:
            pos = cleaned.index(code)
            tail = cleaned[pos+1:]
            for t in tail:
                if not width and re.fullmatch(r'\d{3,4}\*{0,2}', t.replace(' ', '')):
                    length = t
                if not packaging and _extract_package_from_text(t):
                    packaging = _extract_package_from_text(t)
            nums = [(_to_float_or_none(_clean_pdf_price_text(t)), t) for t in tail]
            nums = [(n, t) for n, t in nums if n is not None]
            if nums:
                price = nums[-1][0]

        if name:
            active_name = name
        else:
            name = active_name
        width = width or active_width
        if width:
            active_width = width
        unit = unit or active_unit
        if unit:
            active_unit = unit
        packaging_uom = packaging_uom or active_packaging_uom
        if packaging_uom:
            active_packaging_uom = packaging_uom
        packaging = packaging or active_packaging
        if packaging:
            active_packaging = packaging
        weight_sale = weight_sale or active_weight_sale
        if weight_sale:
            active_weight_sale = weight_sale
        if price is None:
            price = active_price
        elif price is not None:
            active_price = price

        product_text = _normalize_text_simple(' '.join([name, width]).strip())
        if not product_text or len(product_text) < 3:
            continue
        if not (code or price is not None or length or packaging):
            continue

        mm_text = _extract_mm_from_text(' '.join([str(unit), str(packaging_uom)])) or _extract_mm_from_text(product_text)
        package_text = _extract_package_from_text(packaging) or _extract_package_from_text(product_text)
        notes = []
        if length:
            notes.append(f'Length {length}')
        if packaging_uom:
            notes.append(f'U.M. Packaging {packaging_uom}')
        if packaging and packaging != package_text:
            notes.append(packaging)
        if weight_sale:
            notes.append(f'Weight/U.M. Sale {weight_sale}')

        rows.append({
            'SAP': code,
            'Product': product_text,
            'Base Price': price,
            'Increase %': 0.0,
            'Price': price,
            'MM': mm_text,
            'Package': package_text,
            'Category': active_category,
            'Notes': _normalize_text_simple(' | '.join([n for n in notes if n])),
            'Company': company_hint or '',
            'confidence': 0.88 if code and (price is not None) else 0.72,
        })
    return rows


def _parse_pdf_detail_page_text(text: str, current_category: str = '', company_hint: str = ''):
    rows = []
    txt = _normalize_text_simple(text)
    if not txt or _looks_like_pdf_contents_page(txt) or _looks_like_pdf_marketing_page(txt):
        return rows
    lines = [l.strip() for l in (text or '').splitlines() if _normalize_text_simple(l)]
    title = ''
    for line in lines[:12]:
        l = _normalize_text_simple(line)
        low = l.lower()
        if len(l.split()) <= 6 and not _looks_like_pdf_section_title(l) and not any(x in low for x in ['κωδικ', 'τιμη', 'sap code', 'περιεχομενα']):
            if any(ch.isalpha() for ch in l):
                title = l
                break
    saw_pricing = False
    current_category = current_category or 'PDF Catalog'
    for raw in lines:
        line = _normalize_text_simple(raw)
        low = line.lower()
        if any(k in low for k in ['κωδικ', 'sap code']) and any(k in low for k in ['τιμη', 'list price']):
            saw_pricing = True
            continue
        if not saw_pricing or _looks_like_pdf_section_title(line) or len(line) < 6:
            continue
        price_matches = re.findall(r'\d+[.,]\d{1,4}\s*€?', line)
        if not price_matches:
            continue
        code_match = re.search(r'\b[0-9][0-9A-Za-z./_-]{3,}\b', line)
        package_match = re.search(r'(\d+(?:[.,]\d+)?)\s*(kg|gr|g|lt|l|ml|m2|m²|m|τεμ\.?|pcs?)\b', line, re.I)
        price = _to_float_or_none(_clean_pdf_price_text(price_matches[-1]))
        if not code_match and not package_match:
            continue
        code = code_match.group(0) if code_match else ''
        package = ''
        if package_match:
            qty = package_match.group(1).replace(',', '.')
            qty = qty.rstrip('0').rstrip('.') if '.' in qty else qty
            package = f"{qty}{package_match.group(2).lower().replace('gr','g').replace('lt','l').replace('m²','m2')}"
        rows.append({
            'SAP': code,
            'Product': title or current_category or 'PDF Product',
            'Base Price': price,
            'Increase %': 0.0,
            'Price': price,
            'MM': _extract_mm_from_text(line),
            'Package': package,
            'Category': current_category,
            'Notes': line,
            'Company': company_hint or '',
            'confidence': 0.74,
        })
    return rows
def _clean_pdf_price_text(value) -> str:
    if value is None:
        return ''
    text = str(value).strip()
    text = text.replace('€', '').replace('EUR', '').replace('eur', '').strip()
    return text


def _extract_json_payload_from_ai_text(text: str):
    raw = str(text or '').strip()
    if not raw:
        return []
    try:
        payload = json.loads(raw)
    except Exception:
        m = re.search(r'```json\s*(.*?)```', raw, flags=re.S | re.I)
        if m:
            try:
                payload = json.loads(m.group(1))
            except Exception:
                return []
        else:
            return []
    if isinstance(payload, dict):
        rows = payload.get('rows', [])
    elif isinstance(payload, list):
        rows = payload
    else:
        rows = []
    return rows if isinstance(rows, list) else []


def _coerce_ai_pdf_rows(rows, current_category: str = '', company_hint: str = '', page_number=None):
    out = []
    for row in rows or []:
        if not isinstance(row, dict):
            continue
        product = _normalize_text_simple(row.get('product_name') or row.get('product') or row.get('description') or row.get('name') or '')
        if not product or len(product) < 2:
            continue
        category = _normalize_text_simple(row.get('category') or current_category or 'PDF Catalog')
        sap = _normalize_text_simple(row.get('sap_code') or row.get('sap') or row.get('code') or '')
        price = row.get('price')
        price = _to_float_or_none(price if price is not None else row.get('price_text'))
        width = row.get('width_mm')
        length = row.get('length_mm')
        thickness = row.get('thickness_mm')
        dims = []
        for label, val in [('T', thickness), ('W', width), ('L', length)]:
            vv = _to_float_or_none(val)
            if vv is not None:
                vv = int(vv) if float(vv).is_integer() else float(vv)
                dims.append(f'{label}:{vv}mm')
        mm_text = _normalize_text_simple(' | '.join(dims) or row.get('mm') or '')
        package = _normalize_text_simple(row.get('package') or '')
        unit = _normalize_text_simple(row.get('unit') or '')
        notes_bits = []
        if unit:
            notes_bits.append(f'Unit {unit}')
        if row.get('notes'):
            notes_bits.append(_normalize_text_simple(row.get('notes')))
        if page_number:
            notes_bits.append(f'Source Page {page_number}')
        confidence = _to_float_or_none(row.get('confidence'))
        if confidence is None:
            confidence = 0.86 if (sap and price is not None) else 0.76
        out.append({
            'confidence': max(0.0, min(1.0, float(confidence))),
            'SAP': sap,
            'Product': product,
            'Base Price': price,
            'Increase %': 0.0,
            'Price': price,
            'MM': mm_text or _extract_mm_from_text(product),
            'Package': package or _extract_package_from_text(product),
            'Category': category,
            'Notes': _normalize_text_simple(' | '.join([n for n in notes_bits if n])),
            'Company': company_hint or '',
        })
    return out


def _ai_extract_pdf_rows_from_text(text: str, current_category: str = '', company_hint: str = '', page_number=None):
    txt = _normalize_text_simple(text)
    if not txt or _looks_like_pdf_contents_page(txt) or _looks_like_pdf_marketing_page(txt):
        return [], {}
    client = get_openai_client_for_pdf_extraction()
    if client is None:
        return [], {}
    system_prompt = (
        'Είσαι σύστημα εξαγωγής δομημένων προϊόντων από PDF τιμοκαταλόγους δομικών υλικών, χημικών και ξηράς δόμησης. '
        'Στόχος: να εξάγεις ΟΛΑ τα εμπορικά variants. 1 row = 1 variant. '
        'Αγνόησε εξώφυλλα, marketing pages και πίνακες περιεχομένων. '
        'Αν ένα προϊόν έχει πολλά μήκη, πάχη, συσκευασίες ή SAP codes, επέστρεψε ξεχωριστό row για κάθε variant. '
        'Μην απορρίπτεις row επειδή λείπει ένα πεδίο. Προτίμησε over-extraction και βάλε confidence. '
        'Επέστρεψε μόνο JSON object με κλειδί rows.'
    )
    user_prompt = (
        'Εξήγαγε προϊόντα από το παρακάτω κείμενο PDF page.\n'
        f'Company hint: {company_hint or ""}\n'
        f'Current category hint: {current_category or ""}\n'
        f'Page: {page_number or ""}\n\n'
        'Επιστροφή σε JSON με μορφή {"rows": [...]} και για κάθε row fields: '
        'category, product_name, sap_code, code, thickness_mm, width_mm, length_mm, package, unit, price, price_text, notes, confidence.\n\n'
        'ΚΕΙΜΕΝΟ:\n' + text[:45000]
    )
    usage_meta = {}
    try:
        resp = client.chat.completions.create(
            model=os.getenv('OPENAI_PDF_MODEL', 'gpt-4o-mini'),
            temperature=0,
            response_format={"type": "json_object"},
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
        )
        content = ''
        if getattr(resp, 'choices', None):
            content = resp.choices[0].message.content or ''
        rows = _extract_json_payload_from_ai_text(content)
        usage = getattr(resp, 'usage', None)
        usage_meta = {
            'prompt_tokens': int(getattr(usage, 'prompt_tokens', 0) or 0),
            'completion_tokens': int(getattr(usage, 'completion_tokens', 0) or 0),
            'total_tokens': int(getattr(usage, 'total_tokens', 0) or 0),
            'model': os.getenv('OPENAI_PDF_MODEL', 'gpt-4o-mini'),
        }
        return _coerce_ai_pdf_rows(rows, current_category=current_category, company_hint=company_hint, page_number=page_number), usage_meta
    except Exception as e:
        usage_meta['error'] = str(e)
        return [], usage_meta


def _parse_pdf_candidate_row(values, current_category: str = '', company_hint: str = ''):
    cleaned = [_normalize_text_simple(v) for v in values]
    cleaned = [v for v in cleaned if v]
    if not cleaned:
        return None

    if len(cleaned) == 1 and _looks_like_pdf_section_title(cleaned[0]):
        return {'__section_title__': cleaned[0]}
    joined = ' '.join(cleaned)
    if _looks_like_pdf_contents_page(joined):
        return None

    price_idx, base_price = _extract_last_valid_price_from_values(cleaned)
    if price_idx is None:
        price_idx = len(cleaned)
    left = cleaned[:price_idx] or cleaned
    trailing = cleaned[price_idx + 1:] if price_idx < len(cleaned) else []

    sap_text = ''
    if left and _looks_like_pdf_product_code(left[0].replace(' ', '')):
        sap_text = left[0]
        left = left[1:]

    product_parts = []
    for tok in left:
        low = tok.lower()
        if _looks_like_money_token(tok):
            continue
        if re.fullmatch(r'\d{3,4}\*?', tok.replace(' ', '')):
            continue
        if low in {'m2', 'm²', 'm', 'kg', 'lt', 'l', 'ml', 'pcs', 'pc', 'τεμ'}:
            continue
        product_parts.append(tok)

    product_text = _normalize_text_simple(' '.join(product_parts))
    if not product_text or len(product_text) < 2:
        return None
    if _to_float_or_none(product_text) is not None:
        return None

    package_text = _extract_package_from_text(joined) or _extract_package_from_text(product_text)
    mm_text = _extract_mm_from_text(joined) or _extract_mm_from_text(product_text)
    notes_bits = []
    if trailing:
        notes_bits.append(_normalize_text_simple(' '.join(trailing)))
    m = re.search(r'\d{3,4}\s*[xX]\s*\d{3,4}', joined)
    if m:
        notes_bits.append(m.group(0))

    return {
        'confidence': 0.74 if (sap_text and base_price is not None) else 0.62,
        'SAP': sap_text,
        'Product': product_text,
        'Base Price': base_price,
        'Increase %': 0.0,
        'Price': base_price,
        'MM': mm_text,
        'Package': package_text,
        'Category': current_category or 'PDF Catalog',
        'Notes': _normalize_text_simple(' | '.join([n for n in notes_bits if n])),
        'Company': company_hint or '',
    }


def _extract_pdf_rows_from_text(text: str, current_category: str = '', company_hint: str = ''):
    rows = []
    detected_titles = []
    state = _PdfParseState()
    state.category = current_category or 'PDF Catalog'
    buffer = []

    def _flush_buffer():
        nonlocal buffer
        if not buffer:
            return
        merged = _normalize_text_simple(' '.join(buffer))
        buffer = []
        candidate = _parse_pdf_line_stateful(merged, state, company_hint=company_hint)
        if isinstance(candidate, dict) and candidate.get('__section_title__'):
            detected_titles.append(candidate['__section_title__'])
            return
        if isinstance(candidate, dict) and candidate.get('__family_title__'):
            detected_titles.append(candidate['__family_title__'])
            return
        if candidate:
            rows.append(candidate)

    for raw_line in (text or '').splitlines():
        line = _normalize_text_simple(raw_line)
        if not line:
            _flush_buffer()
            continue
        if _looks_like_pdf_contents_page(line) or _looks_like_pdf_marketing_page(line):
            continue
        if _looks_like_pdf_section_title(line):
            _flush_buffer()
            state.category = line
            state.reset_family()
            detected_titles.append(line)
            continue
        if _looks_like_pdf_family_title(line):
            _flush_buffer()
            state.family = line
            detected_titles.append(line)
            continue
        first = line.split()[0]
        if _looks_like_pdf_product_code(first) or _looks_like_money_token(line.split()[-1]):
            _flush_buffer()
            buffer = [line]
            if _extract_last_valid_price_from_values(line.split())[0] is not None:
                _flush_buffer()
            continue
        if buffer:
            buffer.append(line)
        else:
            candidate = _parse_pdf_line_stateful(line, state, company_hint=company_hint)
            if candidate and not candidate.get('__section_title__') and not candidate.get('__family_title__'):
                rows.append(candidate)
    _flush_buffer()
    return rows, detected_titles


def _pdf_processing_chunk_size() -> int:
    try:
        return max(1, min(20, int(os.getenv("PDF_PROCESSING_CHUNK_SIZE", "4"))))
    except Exception:
        return 4


def _pdf_processing_max_pages() -> int:
    try:
        raw = int(os.getenv("PDF_PROCESSING_MAX_PAGES", "0"))
        return max(0, raw)
    except Exception:
        return 0


def _recommended_pdf_chunk_size(total_pages: int) -> int:
    base = _pdf_processing_chunk_size()
    if total_pages >= 150:
        return 1
    if total_pages >= 80:
        return min(base, 2)
    if total_pages >= 40:
        return min(base, 3)
    return base

def _iter_pdf_chunks(file_bytes: bytes, chunk_size: int, max_pages: int = 0):
    reader = PdfReader(io.BytesIO(file_bytes))
    total_pages = len(reader.pages)
    if max_pages and max_pages > 0:
        total_pages = min(total_pages, max_pages)
    for start in range(0, total_pages, chunk_size):
        writer = PdfWriter()
        end = min(start + chunk_size, total_pages)
        for i in range(start, end):
            writer.add_page(reader.pages[i])
        buf = io.BytesIO()
        writer.write(buf)
        yield start, end, buf.getvalue(), total_pages
        buf.close()
        gc.collect()


def _flush_pdfplumber_page(page):
    try:
        page.flush_cache()
    except Exception:
        pass
    try:
        page.get_textmap.cache_clear()
    except Exception:
        pass



def convert_supplier_pdf_to_source(uploaded_file):
    file_bytes = uploaded_file.getvalue()
    all_rows = []
    used_pages = []
    skipped_pages = []
    rows_missing_price = 0
    rows_missing_sap = 0
    rows_missing_product = 0
    detected_section_titles = []
    company_hint = Path(uploaded_file.name).stem
    ai_usage_totals = {"prompt_tokens": 0, "completion_tokens": 0, "total_tokens": 0, "calls": 0}
    ai_enabled = current_user_can_use_pdf_catalog_extraction() and (get_openai_client_for_pdf_extraction() is not None)

    reader_preview = PdfReader(io.BytesIO(file_bytes))
    preview_total_pages = len(reader_preview.pages)
    del reader_preview
    gc.collect()
    chunk_size = _recommended_pdf_chunk_size(preview_total_pages)
    max_pages = _pdf_processing_max_pages()
    total_pages_seen = 0
    state = _PdfParseState()

    for chunk_start, chunk_end, chunk_bytes, total_pages in _iter_pdf_chunks(file_bytes, chunk_size=chunk_size, max_pages=max_pages):
        total_pages_seen = max(total_pages_seen, total_pages)
        with pdfplumber.open(io.BytesIO(chunk_bytes)) as pdf:
            for offset, page in enumerate(pdf.pages, start=0):
                page_idx = chunk_start + offset + 1
                page_label = f'Page {page_idx}'
                page_rows = []
                current_category = state.category or f'PDF Page {page_idx}'

                text = page.extract_text() or ''
                normalized_text = _normalize_text_simple(text)
                text_lines = [_normalize_text_simple(line) for line in text.splitlines() if _normalize_text_simple(line)]

                if _looks_like_pdf_noise_page_relaxed(normalized_text):
                    skipped_pages.append(f'{page_label} (contents/marketing page)')
                    _flush_pdfplumber_page(page)
                    del text, normalized_text, text_lines, page_rows
                    gc.collect()
                    continue

                for line in text_lines[:30]:
                    if _looks_like_pdf_section_title(line) and 'page' not in line.lower():
                        current_category = line
                        state.category = current_category
                        state.reset_family()
                        detected_section_titles.append(f'{page_label}: {line}')
                        break
                    if _looks_like_pdf_family_title(line):
                        state.family = line

                tables = page.extract_tables() or []
                for table in tables:
                    variant_rows = _parse_pdf_table_variants(table, current_category=current_category, company_hint=company_hint)
                    if variant_rows:
                        page_rows.extend(variant_rows)
                        continue
                    for raw_row in table:
                        candidate = _parse_pdf_candidate_row(raw_row or [], current_category=current_category, company_hint=company_hint)
                        if isinstance(candidate, dict) and candidate.get('__section_title__'):
                            current_category = candidate['__section_title__']
                            detected_section_titles.append(f'{page_label}: {current_category}')
                            continue
                        if candidate:
                            page_rows.append(candidate)

                detail_rows = _parse_pdf_detail_page_text(text, current_category=current_category, company_hint=company_hint)
                page_rows.extend(detail_rows)

                fallback_rows, fallback_titles = _extract_pdf_rows_from_text(text, current_category=current_category, company_hint=company_hint)
                page_rows.extend(fallback_rows)
                detected_section_titles.extend([f'{page_label}: {title}' for title in fallback_titles])

                if ai_enabled and len(page_rows) < 6 and len(normalized_text) < 24000:
                    ai_rows, usage_meta = _ai_extract_pdf_rows_from_text(text, current_category=current_category, company_hint=company_hint, page_number=page_idx)
                    ai_usage_totals['prompt_tokens'] += int(usage_meta.get('prompt_tokens', 0) or 0)
                    ai_usage_totals['completion_tokens'] += int(usage_meta.get('completion_tokens', 0) or 0)
                    ai_usage_totals['total_tokens'] += int(usage_meta.get('total_tokens', 0) or 0)
                    if usage_meta.get('prompt_tokens') or usage_meta.get('completion_tokens'):
                        ai_usage_totals['calls'] += 1
                    if ai_rows:
                        page_rows.extend(ai_rows)

                if not page_rows:
                    skipped_pages.append(f'{page_label} (no valid product rows)')
                    state.reset_family()
                    _flush_pdfplumber_page(page)
                    del text, normalized_text, text_lines, tables, detail_rows, fallback_rows, fallback_titles, page_rows
                    gc.collect()
                    continue

                used_pages.append(page_label)
                state.last_table_like = bool(tables) or len(page_rows) >= 3
                for row in page_rows:
                    if (not row.get('Product')) and state.family:
                        row['Product'] = state.family
                    if row.get('Category') in {'', 'PDF Catalog', f'PDF Page {page_idx}'} and state.category:
                        row['Category'] = state.category
                    notes = row.get('Notes', '') or ''
                    row['Notes'] = (notes + (f' | Source {page_label}' if notes else f'Source {page_label}')).strip()
                    row.setdefault('Category', current_category)
                    row.setdefault('Company', company_hint or '')
                    row.setdefault('confidence', 0.70)
                all_rows.extend(page_rows)

                _flush_pdfplumber_page(page)
                del text, normalized_text, text_lines, tables, detail_rows, fallback_rows, fallback_titles, page_rows
                gc.collect()

        del chunk_bytes
        gc.collect()

    if not all_rows:
        return None, {
            'used_sheets': used_pages,
            'skipped_sheets': skipped_pages,
            'missing_price_rows': 0,
            'missing_sap_rows': 0,
            'missing_product_rows': 0,
            'detected_section_titles': detected_section_titles,
            'total_rows': 0,
            'input_kind': 'pdf',
            'processed_pages': total_pages_seen,
            'chunk_size': chunk_size,
        }

    source_df = pd.DataFrame(all_rows)
    source_df['Base Price'] = pd.to_numeric(source_df['Base Price'], errors='coerce')
    source_df['Price'] = source_df['Base Price']

    for col, default in [('SAP',''),('Product',''),('MM',''),('Package',''),('Category','PDF Catalog'),('confidence',0.70)]:
        if col not in source_df.columns:
            source_df[col] = default

    source_df['Product'] = source_df['Product'].astype(str).str.replace(r'\s+', ' ', regex=True).str.strip()
    source_df['SAP'] = source_df['SAP'].astype(str).str.strip()
    source_df['Package'] = source_df['Package'].astype(str).str.strip()
    source_df['Category'] = source_df['Category'].astype(str).str.strip().replace('', 'PDF Catalog')
    source_df = source_df[~((source_df['Product'].astype(str).str.strip().eq('')) & source_df['Base Price'].isna())].copy()
    source_df = source_df[~(source_df['Product'].str.lower().isin(['περιεχομενα', 'contents']))].copy()
    source_df = source_df[~source_df['Base Price'].astype(str).str.contains(r'x|mm|cm', case=False, na=False)].copy()
    title_noise_re = r'(?:^|\b)(price list|catalog|table of contents|contents|delivery conditions|valid from|technical data|overview|description|application|packaging|recommended retail|τιμοκαταλογος|περιεχομενα|ισχυς απο|προυποθεσεις|τεχνικα χαρακτηριστικα|εφαρμογη|περιγραφη|γυψοσανιδες για|γυψοπλακες για|οι τιμες ειναι|χωρις φπα)(?:\b|$)'
    source_df = source_df[~source_df['Product'].str.lower().str.contains(title_noise_re, regex=True, na=False)].copy()
    source_df['SAP'] = source_df['SAP'].where(source_df['SAP'].astype(str).str.match(r'^(?:\d{5,8}|[A-Z]{1,4}\d{1,5}[A-Z0-9./_-]{0,4})$', na=False), '')
    dedupe_cols = [c for c in ['SAP','Product','Package','Base Price','Category'] if c in source_df.columns]
    if dedupe_cols:
        source_df = source_df.drop_duplicates(subset=dedupe_cols, keep='first')
    source_df = source_df[_source_generator_output_columns()].reset_index(drop=True)

    rows_missing_price = int(source_df['Base Price'].isna().sum())
    rows_missing_sap = int(source_df['SAP'].astype(str).str.strip().eq('').sum())
    rows_missing_product = int(source_df['Product'].astype(str).str.strip().eq('').sum())

    stats = {
        'used_sheets': used_pages,
        'skipped_sheets': skipped_pages,
        'missing_price_rows': rows_missing_price,
        'missing_sap_rows': rows_missing_sap,
        'missing_product_rows': rows_missing_product,
        'detected_section_titles': detected_section_titles,
        'total_rows': len(source_df),
        'input_kind': 'pdf',
        'ai_prompt_tokens': int(ai_usage_totals.get('prompt_tokens', 0) or 0),
        'ai_completion_tokens': int(ai_usage_totals.get('completion_tokens', 0) or 0),
        'ai_total_tokens': int(ai_usage_totals.get('total_tokens', 0) or 0),
        'ai_calls': int(ai_usage_totals.get('calls', 0) or 0),
        'processed_pages': total_pages_seen,
        'chunk_size': chunk_size,
    }
    if stats['ai_calls'] > 0 and is_admin_user():
        append_pdf_ai_usage_log({
            'timestamp': datetime.now().isoformat(timespec='seconds'),
            'user': get_current_user_email() or 'admin',
            'file_name': getattr(uploaded_file, 'name', 'pdf_upload'),
            'pages': len(used_pages),
            'rows': len(source_df),
            'prompt_tokens': stats['ai_prompt_tokens'],
            'completion_tokens': stats['ai_completion_tokens'],
            'total_tokens': stats['ai_total_tokens'],
            'ai_calls': stats['ai_calls'],
        })
    return source_df, stats


def convert_supplier_pricelist_to_source(uploaded_file):
    xls, file_bytes = load_excel_file_any(uploaded_file)

    if _is_siniat_april_workbook(xls.sheet_names):
        return _parse_siniat_workbook(xls, file_bytes)

    all_rows = []
    used_sheets = []
    skipped_sheets = []
    rows_missing_price = 0
    rows_missing_sap = 0
    rows_missing_product = 0
    detected_section_titles = []

    helper_tokens = ["cover", "legend", "notes", "summary", "readme", "categorie", "lookup", "contents", "index"]

    for sheet_name in xls.sheet_names:
        lowered_sheet = str(sheet_name).strip().lower()
        if any(token in lowered_sheet for token in helper_tokens):
            skipped_sheets.append(f"{sheet_name} (helper)")
            continue

        raw = read_excel_any(file_bytes, sheet_name=sheet_name, header=None)
        if raw is None or raw.empty:
            skipped_sheets.append(f"{sheet_name} (empty)")
            continue

        raw = raw.ffill(axis=1)
        header_row = _detect_supplier_header_row(raw)
        df = _normalize_supplier_dataframe_from_raw(raw, header_row=header_row)
        if df.empty:
            skipped_sheets.append(f"{sheet_name} (no rows)")
            continue

        original_columns = [str(c).strip() for c in df.columns]
        renamed_columns = {_col: _normalize_supplier_column_name(_col) for _col in original_columns}
        df = df.rename(columns=renamed_columns)
        df = df.loc[:, ~pd.Index(df.columns).duplicated(keep="first")]

        guessed_columns = _guess_supplier_columns_by_values(df)
        for canonical_name, original_col in guessed_columns.items():
            if canonical_name not in df.columns and original_col in df.columns:
                df = df.rename(columns={original_col: canonical_name})
        df = df.loc[:, ~pd.Index(df.columns).duplicated(keep="first")]

        if "Product" not in df.columns:
            skipped_sheets.append(f"{sheet_name} (missing Product)")
            continue

        if "Base Price" not in df.columns:
            skipped_sheets.append(f"{sheet_name} (missing Base Price)")
            continue

        if "SAP" not in df.columns:
            df["SAP"] = ""

        if "Category" not in df.columns:
            df["Category"] = ""

        normalized_rows = []
        current_category = str(sheet_name).strip()

        for _, source_row in df.iterrows():
            row_values = source_row.tolist()

            if _is_section_title_row(row_values):
                section_title = _extract_section_title(row_values)
                if section_title:
                    current_category = section_title
                    detected_section_titles.append(f"{sheet_name}: {section_title}")
                continue

            product_value = source_row["Product"] if "Product" in source_row.index else None
            price_value = source_row["Base Price"] if "Base Price" in source_row.index else None
            sap_value = source_row["SAP"] if "SAP" in source_row.index else ""
            inc_value = source_row["Increase %"] if "Increase %" in source_row.index else 0.0
            mm_value = source_row["MM"] if "MM" in source_row.index else ""
            pack_value = source_row["Package"] if "Package" in source_row.index else ""
            cat_value = source_row["Category"] if "Category" in source_row.index else ""

            product_text = "" if product_value is None else str(product_value).strip()
            base_price = _to_float_or_none(price_value)
            sap_text = "" if sap_value is None else str(sap_value).strip()

            if product_text.lower() in {"", "nan", "none"} and base_price is None:
                continue

            if _to_float_or_none(product_text) is not None and base_price is None:
                continue

            if not product_text or product_text.lower() in {"nan", "none"}:
                rows_missing_product += 1
                continue

            increase_fraction = _to_increase_fraction(inc_value)
            mm_text = "" if mm_value is None or (isinstance(mm_value, float) and pd.isna(mm_value)) else str(mm_value).strip()
            package_text = "" if pack_value is None or (isinstance(pack_value, float) and pd.isna(pack_value)) else str(pack_value).strip()
            category_text = "" if cat_value is None or (isinstance(cat_value, float) and pd.isna(cat_value)) else str(cat_value).strip()

            final_category = category_text if category_text and category_text.lower() not in {"nan", "none"} else current_category

            if base_price is None:
                rows_missing_price += 1
            if not sap_text:
                rows_missing_sap += 1

            normalized_rows.append({
                "SAP": sap_text,
                "Product": product_text,
                "Base Price": base_price,
                "Increase %": 0.0,
                "Price": base_price,
                "MM": mm_text,
                "Package": package_text,
                "Category": final_category,
            })

        if not normalized_rows:
            skipped_sheets.append(f"{sheet_name} (no valid product rows)")
            continue

        out = pd.DataFrame(normalized_rows)
        if out.empty:
            skipped_sheets.append(f"{sheet_name} (all rows invalid)")
            continue

        mask_valid = ~(
            out["Product"].astype(str).str.strip().str.lower().isin(["", "nan", "none"])
            & out["Base Price"].isna()
        )
        out = out[mask_valid].copy()
        out = out.reset_index(drop=True)

        if out.empty:
            skipped_sheets.append(f"{sheet_name} (all rows invalid)")
            continue

        used_sheets.append(sheet_name)
        all_rows.append(out)

    if not all_rows:
        return None, {
            "used_sheets": used_sheets,
            "skipped_sheets": skipped_sheets,
            "missing_price_rows": rows_missing_price,
            "missing_sap_rows": rows_missing_sap,
            "missing_product_rows": rows_missing_product,
            "detected_section_titles": detected_section_titles,
            "total_rows": 0,
        }

    source_df = pd.concat(all_rows, ignore_index=True)
    source_df["Base Price"] = pd.to_numeric(source_df["Base Price"], errors="coerce")
    source_df["Price"] = source_df["Base Price"]
    source_df = source_df[_source_generator_output_columns()].reset_index(drop=True)

    stats = {
        "used_sheets": used_sheets,
        "skipped_sheets": skipped_sheets,
        "missing_price_rows": rows_missing_price,
        "missing_sap_rows": rows_missing_sap,
        "missing_product_rows": rows_missing_product,
        "detected_section_titles": detected_section_titles,
        "total_rows": len(source_df),
    }
    return source_df, stats

def _list_supplier_sheet_names(uploaded_file):
    xls, _file_bytes = load_excel_file_any(uploaded_file)
    return list(xls.sheet_names)


def _load_supplier_sheet_for_mapping(uploaded_file, sheet_name: str, header_row: int):
    _xls, file_bytes = load_excel_file_any(uploaded_file)
    raw_preview = read_excel_any(file_bytes, sheet_name=sheet_name, header=None)
    mapped_df = read_excel_any(file_bytes, sheet_name=sheet_name, header=header_row)
    if mapped_df is None or mapped_df.empty:
        return raw_preview, pd.DataFrame()
    mapped_df.columns = [str(c).strip() for c in mapped_df.columns]
    return raw_preview, mapped_df


def _build_source_from_manual_mapping(mapped_df: pd.DataFrame, column_mapping: dict, default_category: str):
    out = pd.DataFrame()

    sap_col = column_mapping.get("SAP")
    product_col = column_mapping.get("Product")
    price_col = column_mapping.get("Base Price")
    inc_col = column_mapping.get("Increase %")
    mm_col = column_mapping.get("MM")
    pack_col = column_mapping.get("Package")
    cat_col = column_mapping.get("Category")

    out["SAP"] = mapped_df[sap_col].astype(str).str.strip() if sap_col else ""
    out["Product"] = mapped_df[product_col].astype(str).str.strip() if product_col else ""
    out["Base Price"] = mapped_df[price_col].apply(_to_float_or_none) if price_col else None
    out["Increase %"] = mapped_df[inc_col].apply(_to_increase_fraction) if inc_col else 0.0
    out["Price"] = out.apply(
        lambda r: round(r["Base Price"] * (1 + r["Increase %"]), 4) if pd.notna(r["Base Price"]) and r["Base Price"] is not None else None,
        axis=1,
    )
    out["MM"] = mapped_df[mm_col].astype(str).str.strip() if mm_col else ""
    out["Package"] = mapped_df[pack_col].astype(str).str.strip() if pack_col else ""
    out["Category"] = mapped_df[cat_col].astype(str).str.strip() if cat_col else str(default_category).strip()

    out = out[_source_generator_output_columns()]
    out["SAP"] = out["SAP"].fillna("").astype(str).replace({"nan": "", "None": ""})
    out["Product"] = out["Product"].fillna("").astype(str).replace({"nan": "", "None": ""})
    out["MM"] = out["MM"].fillna("").astype(str).replace({"nan": "", "None": ""})
    out["Package"] = out["Package"].fillna("").astype(str).replace({"nan": "", "None": ""})
    out["Category"] = out["Category"].fillna("").astype(str).replace({"nan": "", "None": ""})

    out = out[
        ~(
            out["Product"].eq("")
            & out["Base Price"].isna()
        )
    ].reset_index(drop=True)

    if not out.empty:
        out = out[out["Base Price"].notna()].reset_index(drop=True)

    return out


def _default_manual_mapping(columns):
    mapping = {key: None for key in _source_generator_output_columns() if key != "Price"}
    for col in columns:
        norm = _normalize_supplier_column_name(col)
        if norm in mapping and mapping[norm] is None:
            mapping[norm] = col
    guessed = _guess_supplier_columns_by_values(pd.DataFrame(columns=columns)) if False else {}
    return mapping





def current_user_can_use_pdf_catalog_extraction():
    return bool(is_admin_user())

def render_sources():
    if "pdf_processing" not in st.session_state:
        st.session_state["pdf_processing"] = False

    st.markdown('<div class="app-card">', unsafe_allow_html=True)
    st.markdown("## Sources")

    company_display_map = {
        f"{row['name']} ({row['code']})": row["code"] for _, row in companies_df.iterrows()
    }

    st.markdown("### 1. Create Source from Supplier Pricelist")
    st.caption("Upload a supplier pricelist file, review the converted rows, edit anything you want, and then save it as a ready PRICELIST source file.")

    pdf_ai_premium_enabled = current_user_can_use_pdf_catalog_extraction()
    st.caption("Excel upload: διαθέσιμο για όλους. Το AI PDF extraction εμφανίζεται ως ξεχωριστό εργαλείο μόνο για Admin.")

    gen_c1, gen_c2, gen_c3 = st.columns(3)
    with gen_c1:
        generator_company_display = st.selectbox(
            "Company for generated Source",
            list(company_display_map.keys()),
            key="generator_company_display",
        )
        generator_company_code = company_display_map[generator_company_display]

    with gen_c2:
        generator_date_val = st.date_input(
            "Generated Source Date",
            value=date.today(),
            key="generator_date",
        )

    with gen_c3:
        uploaded_supplier_file = st.file_uploader(
            "Upload supplier pricelist (Excel)",
            type=["xlsx", "xlsm"],
            key="supplier_pricelist_upload",
            help="Excel files are converted directly into a Source preview.",
        )

    uploaded_pdf_supplier_file = None
    if pdf_ai_premium_enabled:
        st.markdown("### 🤖 AI PDF Catalog Extraction (Admin Only)")
        st.info("Έξυπνη εξαγωγή τιμοκαταλόγων PDF σε δομημένο Excel")
        st.caption("Ξεχωριστό εργαλείο PDF μόνο για Admin. Το Save Generated Source παραμένει ενεργό μέσα στην εφαρμογή, ενώ το τοπικό download παραμένει μόνο για admin.")
        uploaded_pdf_supplier_file = st.file_uploader(
            "Upload supplier pricelist (PDF)",
            type=["pdf"],
            key="supplier_pricelist_pdf_upload",
            help="Enhanced extraction with better page filtering and variant expansion.",
        )

    uploaded_supplier_file = uploaded_pdf_supplier_file if uploaded_pdf_supplier_file is not None else uploaded_supplier_file

    if uploaded_supplier_file is not None:
        source_df, conversion_stats = None, None
        uploaded_name = str(getattr(uploaded_supplier_file, "name", "") or "").lower()
        is_pdf_upload = uploaded_name.endswith(".pdf")

        progress_container = st.container() if is_pdf_upload else None
        status_placeholder = progress_container.empty() if progress_container is not None else None
        progress_placeholder = progress_container.empty() if progress_container is not None else None

        try:
            if is_pdf_upload:
                run_pdf = st.button("🚀 Convert PDF with AI", key="convert_pdf_with_ai_button", disabled=st.session_state.get("pdf_processing", False), use_container_width=True)
                if not run_pdf:
                    st.stop()
                st.session_state["pdf_processing"] = True
                status_placeholder.info("🔍 Ανάλυση PDF τιμοκαταλόγου…")
                progress_bar = progress_placeholder.progress(10)
                time.sleep(0.05)

                status_placeholder.info("🧠 Εξαγωγή δεδομένων και αναγνώριση προϊόντων…")
                progress_bar.progress(35)
                time.sleep(0.05)

                source_df, conversion_stats = convert_supplier_pdf_to_source(uploaded_supplier_file)

                status_placeholder.info("📊 Δομή καταλόγου και έλεγχος πεδίων…")
                progress_bar.progress(75)
                time.sleep(0.05)

                status_placeholder.info("📁 Δημιουργία δομημένου Excel preview…")
                progress_bar.progress(95)
                time.sleep(0.05)
                progress_bar.progress(100)
                status_placeholder.success("✅ Η έξυπνη εξαγωγή ολοκληρώθηκε.")
            else:
                source_df, conversion_stats = convert_supplier_pricelist_to_source(uploaded_supplier_file)
        except Exception as e:
            if is_pdf_upload and status_placeholder is not None:
                status_placeholder.error(f"❌ Σφάλμα επεξεργασίας: {e}")
            else:
                st.error(str(e))
            source_df, conversion_stats = None, None
        finally:
            if is_pdf_upload:
                st.session_state["pdf_processing"] = False

        auto_ok = source_df is not None and not source_df.empty

        if source_df is None or source_df.empty:
            st.error("Could not convert this supplier file automatically.")
            skipped = conversion_stats.get("skipped_sheets", []) if isinstance(conversion_stats, dict) else []
            if skipped:
                st.warning("Skipped sheets: " + ", ".join(skipped))
        else:
            input_kind = str(conversion_stats.get('input_kind', 'excel')) if isinstance(conversion_stats, dict) else 'excel'
            unit_label = 'page(s)' if input_kind == 'pdf' else 'sheet(s)'
            st.success(f"Detected {len(source_df)} valid source rows from {len(conversion_stats.get('used_sheets', []))} {unit_label}.")

            m1, m2, m3, m4 = st.columns(4)
            with m1:
                st.metric("Rows", conversion_stats.get("total_rows", len(source_df)))
            with m2:
                st.metric("Used Pages" if input_kind == "pdf" else "Used Sheets", len(conversion_stats.get("used_sheets", [])))
            with m3:
                st.metric("Missing Prices", conversion_stats.get("missing_price_rows", 0))
            with m4:
                st.metric("Missing SAP/Product", conversion_stats.get("missing_sap_rows", 0) + conversion_stats.get("missing_product_rows", 0))

            if input_kind == "pdf" and is_admin_user():
                ai_calls = int(conversion_stats.get("ai_calls", 0) or 0)
                if ai_calls > 0:
                    a1, a2, a3 = st.columns(3)
                    with a1:
                        st.metric("AI Calls", ai_calls)
                    with a2:
                        st.metric("Prompt Tokens", int(conversion_stats.get("ai_prompt_tokens", 0) or 0))
                    with a3:
                        st.metric("Completion Tokens", int(conversion_stats.get("ai_completion_tokens", 0) or 0))
                    month_usage = current_month_pdf_ai_usage_summary()
                    st.caption(f"Admin AI usage this month: {month_usage['calls']} calls • {month_usage['pages']} pages • {month_usage['rows']} rows extracted")

            detected_titles = conversion_stats.get("detected_section_titles", [])
            if detected_titles:
                st.caption("Detected section/category titles: " + " • ".join(detected_titles[:8]))

            if conversion_stats.get("used_sheets"):
                st.caption(("Used pages: " if input_kind == "pdf" else "Used sheets: ") + ", ".join(conversion_stats["used_sheets"]))
            if conversion_stats.get("skipped_sheets"):
                st.warning("Skipped sheets: " + ", ".join(conversion_stats["skipped_sheets"][:12]))
            if input_kind == "pdf" and "confidence" in source_df.columns:
                low_conf_count = int((pd.to_numeric(source_df["confidence"], errors="coerce") < 0.70).fillna(False).sum())
                if low_conf_count > 0:
                    st.warning(f"⚠️ {low_conf_count} γραμμές χρειάζονται έλεγχο πριν την αποθήκευση.")

            st.markdown("#### Review and Edit Before Save")

            current_generated_origin = f"{uploaded_supplier_file.name}|{generator_company_code}|{len(source_df)}"
            if (
                "generated_source_working_df" not in st.session_state
                or st.session_state.get("generated_source_working_origin") != current_generated_origin
            ):
                st.session_state["generated_source_working_df"] = _ensure_preview_row_id(source_df)
                st.session_state["generated_source_working_origin"] = current_generated_origin

            editable_df = st.data_editor(
                _preview_display_df(st.session_state["generated_source_working_df"]),
                use_container_width=True,
                num_rows="dynamic",
                key="generated_source_editor",
                column_config={
                    "Row": st.column_config.NumberColumn("Row", disabled=True),
                    "SAP": st.column_config.TextColumn("SAP"),
                    "Product": st.column_config.TextColumn("Product", width="large"),
                    "Base Price": st.column_config.NumberColumn("Base Price", format="%.4f"),
                    "Price": st.column_config.NumberColumn("Price", format="%.4f", disabled=True),
                    "MM": st.column_config.TextColumn("MM"),
                    "Package": st.column_config.TextColumn("Package"),
                    "Category": st.column_config.TextColumn("Category"),
                },
            )

            edited_df = _normalize_preview_editor_output(pd.DataFrame(editable_df))
            st.session_state["generated_source_working_df"] = edited_df

            delete_options = edited_df["__row_id"].astype(int).tolist()
            del_c1, del_c2 = st.columns([3, 1])
            with del_c1:
                selected_delete_rows = st.multiselect(
                    "Select preview rows to delete",
                    delete_options,
                    format_func=lambda x: f"Row {x}",
                    key="generated_source_delete_rows",
                )
            with del_c2:
                st.write("")
                st.write("")
                if st.button("Delete selected rows", key="delete_generated_preview_rows", use_container_width=True):
                    deleted_count = _delete_selected_preview_rows_from_state("generated_source_working_df", selected_delete_rows)
                    st.session_state.pop("generated_source_delete_rows", None)
                    if deleted_count:
                        st.success(f"Deleted {deleted_count} row(s) from the preview.")
                    st.rerun()

            edited_df = st.session_state["generated_source_working_df"].copy()
            edited_df["Price"] = edited_df["Base Price"]
            export_edited_df = edited_df.drop(columns=["__row_id"], errors="ignore").copy()

            generated_source_original_name = uploaded_supplier_file.name if not str(uploaded_supplier_file.name).lower().endswith(".pdf") else f"{Path(uploaded_supplier_file.name).stem}.xlsx"
            generated_default_name = get_next_version_filename(generator_company_code, generator_date_val, generated_source_original_name)
            generated_signature = (generator_company_code, str(generator_date_val), uploaded_supplier_file.name, int(len(export_edited_df)))
            if st.session_state.get("save_generated_source_as_signature") != generated_signature:
                st.session_state["save_generated_source_as_signature"] = generated_signature
                st.session_state["save_generated_source_as_name"] = generated_default_name

            preview_c1, preview_c2, preview_c3 = st.columns([1, 1, 1])
            with preview_c1:
                if is_admin_user():
                    st.download_button(
                        "Download Generated Source",
                        data=_source_dataframe_to_excel_bytes(export_edited_df),
                        file_name=f"{generator_company_code}_generated_source_preview.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="download_generated_source_preview",
                        use_container_width=True,
                    )
                else:
                    st.button(
                        "Download Generated Source",
                        key="download_generated_source_preview_locked",
                        disabled=True,
                        use_container_width=True,
                    )
                    st.caption("Μόνο ο admin μπορεί να κατεβάσει το παραγόμενο Excel τοπικά.")
            with preview_c2:
                if st.button("Save Generated Source", key="save_generated_source_button", use_container_width=True):
                    if export_edited_df.empty:
                        st.error("There are no rows to save.")
                    else:
                        name = generated_default_name
                        path = get_company_folder(generator_company_code) / name
                        with open(path, "wb") as f:
                            f.write(_source_dataframe_to_excel_bytes(export_edited_df))
                        st.success(f"Generated source saved as: {name}")
                        refresh_source_file_views()
                        st.rerun()
            with preview_c3:
                with st.popover("Save As", use_container_width=True):
                    st.text_input(
                        "Source name",
                        key="save_generated_source_as_name",
                        help="Choose a custom name for this generated source. You can include or omit the Excel extension.",
                    )
                    if st.button("Save As", key="save_generated_source_as_button", use_container_width=True):
                        if export_edited_df.empty:
                            st.error("There are no rows to save.")
                        else:
                            custom_filename = build_custom_source_filename(
                                st.session_state.get("save_generated_source_as_name", ""),
                                generated_source_original_name,
                            )
                            if not custom_filename:
                                st.error("Please enter a valid source name.")
                            else:
                                path = get_company_folder(generator_company_code) / custom_filename
                                if path.exists():
                                    st.error("A source with this name already exists. Please choose a different name.")
                                else:
                                    with open(path, "wb") as f:
                                        f.write(_source_dataframe_to_excel_bytes(export_edited_df))
                                    st.success(f"Generated source saved as: {custom_filename}")
                                    refresh_source_file_views()
                                    st.rerun()

        is_pdf_upload = str(getattr(uploaded_supplier_file, "name", "") or "").lower().endswith(".pdf")
        with st.expander("Manual column override", expanded=(not auto_ok) and (not is_pdf_upload)):
            if is_pdf_upload:
                st.caption("Manual column override is only available for Excel uploads. For PDFs, edit the preview rows directly before saving.")
                manual_sheet_names = []
            else:
                st.caption("Use this only when the automatic detection is not ideal. Choose the correct sheet, header row, and columns, then rebuild the Source preview.")

                try:
                    manual_sheet_names = _list_supplier_sheet_names(uploaded_supplier_file)
                except Exception as e:
                    manual_sheet_names = []
                    st.error(str(e))

            if manual_sheet_names:
                man_c1, man_c2, man_c3 = st.columns([2, 1, 1])
                with man_c1:
                    manual_sheet = st.selectbox(
                        "Sheet to map",
                        manual_sheet_names,
                        key="manual_override_sheet",
                    )
                with man_c2:
                    manual_header_row = st.number_input(
                        "Header row (0-based)",
                        min_value=0,
                        max_value=50,
                        value=0,
                        step=1,
                        key="manual_override_header_row",
                    )
                with man_c3:
                    manual_default_category = st.text_input(
                        "Default category",
                        value=str(manual_sheet) if 'manual_sheet' in locals() else "",
                        key="manual_override_default_category",
                    )

                if st.button("Load sheet for manual mapping", key="manual_override_load_sheet", use_container_width=True):
                    st.session_state["manual_override_loaded"] = True

                if st.session_state.get("manual_override_loaded"):
                    raw_preview_df, mapping_df = _load_supplier_sheet_for_mapping(
                        uploaded_supplier_file,
                        manual_sheet,
                        int(manual_header_row),
                    )

                    st.markdown("##### Raw preview")
                    st.dataframe(raw_preview_df.head(10), use_container_width=True)

                    if mapping_df is None or mapping_df.empty:
                        st.warning("No rows found for this sheet/header combination.")
                    else:
                        st.markdown("##### Column mapping")
                        st.dataframe(mapping_df.head(10), use_container_width=True)

                        available_cols = list(mapping_df.columns)
                        none_option = ["— None —"] + available_cols

                        auto_defaults = {}
                        for col in available_cols:
                            norm = _normalize_supplier_column_name(col)
                            if norm not in auto_defaults:
                                auto_defaults[norm] = col

                        map_c1, map_c2, map_c3 = st.columns(3)
                        with map_c1:
                            manual_sap = st.selectbox("SAP column", none_option, index=(none_option.index(auto_defaults.get("SAP")) if auto_defaults.get("SAP") in none_option else 0), key="manual_sap_col")
                            manual_product = st.selectbox("Product column", available_cols, index=(available_cols.index(auto_defaults.get("Product")) if auto_defaults.get("Product") in available_cols else 0), key="manual_product_col")
                            manual_price = st.selectbox("Base Price column", available_cols, index=(available_cols.index(auto_defaults.get("Base Price")) if auto_defaults.get("Base Price") in available_cols else 0), key="manual_price_col")
                        with map_c2:
                            manual_inc = st.selectbox("Increase % column", none_option, index=(none_option.index(auto_defaults.get("Increase %")) if auto_defaults.get("Increase %") in none_option else 0), key="manual_inc_col")
                            manual_mm = st.selectbox("MM column", none_option, index=(none_option.index(auto_defaults.get("MM")) if auto_defaults.get("MM") in none_option else 0), key="manual_mm_col")
                        with map_c3:
                            manual_pack = st.selectbox("Package column", none_option, index=(none_option.index(auto_defaults.get("Package")) if auto_defaults.get("Package") in none_option else 0), key="manual_pack_col")
                            manual_cat = st.selectbox("Category column", none_option, index=(none_option.index(auto_defaults.get("Category")) if auto_defaults.get("Category") in none_option else 0), key="manual_cat_col")

                        if st.button("Apply manual mapping", key="manual_override_apply", use_container_width=True):
                            mapping = {
                                "SAP": None if manual_sap == "— None —" else manual_sap,
                                "Product": manual_product,
                                "Base Price": manual_price,
                                "MM": None if manual_mm == "— None —" else manual_mm,
                                "Package": None if manual_pack == "— None —" else manual_pack,
                                "Category": None if manual_cat == "— None —" else manual_cat,
                            }

                            manual_source_df = _build_source_from_manual_mapping(mapping_df, mapping, manual_default_category or manual_sheet)
                            if manual_source_df is None or manual_source_df.empty:
                                st.error("Manual mapping produced no valid rows.")
                            else:
                                st.success(f"Manual mapping generated {len(manual_source_df)} rows.")
                                st.session_state["manual_source_df"] = manual_source_df.to_dict(orient="records")
                                st.session_state["manual_source_working_df"] = _ensure_preview_row_id(manual_source_df)
                                st.session_state["manual_source_working_origin"] = f"{uploaded_supplier_file.name}|{generator_company_code}|manual|{len(manual_source_df)}"

                if st.session_state.get("manual_source_df"):
                    st.markdown("##### Manual mapping preview")
                    current_manual_origin = f"{uploaded_supplier_file.name}|{generator_company_code}|manual|{len(st.session_state.get('manual_source_df', []))}"
                    if (
                        "manual_source_working_df" not in st.session_state
                        or st.session_state.get("manual_source_working_origin") != current_manual_origin
                    ):
                        st.session_state["manual_source_working_df"] = _ensure_preview_row_id(pd.DataFrame(st.session_state["manual_source_df"]))
                        st.session_state["manual_source_working_origin"] = current_manual_origin
                    manual_preview_df = st.data_editor(
                        _preview_display_df(st.session_state["manual_source_working_df"]),
                        use_container_width=True,
                        num_rows="dynamic",
                        key="manual_generated_source_editor",
                        column_config={
                            "Row": st.column_config.NumberColumn("Row", disabled=True),
                            "SAP": st.column_config.TextColumn("SAP"),
                            "Product": st.column_config.TextColumn("Product", width="large"),
                            "Base Price": st.column_config.NumberColumn("Base Price", format="%.4f"),
                                    "Price": st.column_config.NumberColumn("Price", format="%.4f", disabled=True),
                            "MM": st.column_config.TextColumn("MM"),
                            "Package": st.column_config.TextColumn("Package"),
                            "Category": st.column_config.TextColumn("Category"),
                        },
                    )

                    manual_preview_df = _normalize_preview_editor_output(pd.DataFrame(manual_preview_df))
                    st.session_state["manual_source_working_df"] = manual_preview_df

                    man_del_c1, man_del_c2 = st.columns([3, 1])
                    with man_del_c1:
                        selected_manual_delete_rows = st.multiselect(
                            "Select manual preview rows to delete",
                            manual_preview_df["__row_id"].astype(int).tolist(),
                            format_func=lambda x: f"Row {x}",
                            key="manual_source_delete_rows",
                        )
                    with man_del_c2:
                        st.write("")
                        st.write("")
                        if st.button("Delete selected manual rows", key="delete_manual_preview_rows", use_container_width=True):
                            deleted_count = _delete_selected_preview_rows_from_state("manual_source_working_df", selected_manual_delete_rows)
                            st.session_state.pop("manual_source_delete_rows", None)
                            if deleted_count:
                                st.success(f"Deleted {deleted_count} row(s) from the manual preview.")
                            st.rerun()

                    manual_preview_df = st.session_state["manual_source_working_df"].copy()
                    manual_preview_df["Price"] = manual_preview_df["Base Price"]
                    export_manual_preview_df = manual_preview_df.drop(columns=["__row_id"], errors="ignore").copy()

                    manual_default_name = get_next_version_filename(generator_company_code, generator_date_val, uploaded_supplier_file.name)
                    manual_signature = (generator_company_code, str(generator_date_val), uploaded_supplier_file.name, int(len(export_manual_preview_df)), "manual")
                    if st.session_state.get("save_manual_source_as_signature") != manual_signature:
                        st.session_state["save_manual_source_as_signature"] = manual_signature
                        st.session_state["save_manual_source_as_name"] = manual_default_name

                    man_save_c1, man_save_c2, man_save_c3 = st.columns(3)
                    with man_save_c1:
                        st.download_button(
                            "Download Manual Source",
                            data=_source_dataframe_to_excel_bytes(export_manual_preview_df),
                            file_name=f"{generator_company_code}_manual_source_preview.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_manual_source_preview",
                            use_container_width=True,
                        )
                    with man_save_c2:
                        if st.button("Save Manual Source", key="save_manual_source_button", use_container_width=True):
                            if export_manual_preview_df.empty:
                                st.error("There are no rows to save.")
                            else:
                                manual_name = manual_default_name
                                manual_path = get_company_folder(generator_company_code) / manual_name
                                with open(manual_path, "wb") as f:
                                    f.write(_source_dataframe_to_excel_bytes(export_manual_preview_df))
                                st.success(f"Manual source saved as: {manual_name}")
                                refresh_source_file_views()
                                st.rerun()
                    with man_save_c3:
                        with st.popover("Save As", use_container_width=True):
                            st.text_input(
                                "Source name",
                                key="save_manual_source_as_name",
                                help="Choose a custom name for this manual source. You can include or omit the Excel extension.",
                            )
                            if st.button("Save As", key="save_manual_source_as_button", use_container_width=True):
                                if export_manual_preview_df.empty:
                                    st.error("There are no rows to save.")
                                else:
                                    custom_filename = build_custom_source_filename(
                                        st.session_state.get("save_manual_source_as_name", ""),
                                        uploaded_supplier_file.name,
                                    )
                                    if not custom_filename:
                                        st.error("Please enter a valid source name.")
                                    else:
                                        manual_path = get_company_folder(generator_company_code) / custom_filename
                                        if manual_path.exists():
                                            st.error("A source with this name already exists. Please choose a different name.")
                                        else:
                                            with open(manual_path, "wb") as f:
                                                f.write(_source_dataframe_to_excel_bytes(export_manual_preview_df))
                                            st.success(f"Manual source saved as: {custom_filename}")
                                            refresh_source_file_views()
                                            st.rerun()

    st.markdown("---")
    st.markdown("### 2. Save Ready Source")

    s1, s2, s3 = st.columns(3)
    with s1:
        selected_company_display = st.selectbox(
            "Company",
            list(company_display_map.keys()),
            key="save_company",
        )
        company_code = company_display_map[selected_company_display]

    with s2:
        date_val = st.date_input("Date", value=date.today(), key="save_date")

    with s3:
        file = st.file_uploader("Upload Ready Source", type=["xlsx", "xlsm"], key="save_file")

    default_ready_source_name = get_next_version_filename(company_code, date_val, file.name if file is not None else "source.xlsx")
    ready_source_signature = (company_code, str(date_val), file.name if file is not None else "")
    if st.session_state.get("save_source_as_signature") != ready_source_signature:
        st.session_state["save_source_as_signature"] = ready_source_signature
        st.session_state["save_source_as_name"] = default_ready_source_name

    save_c1, save_c2 = st.columns(2)
    with save_c1:
        if st.button("Save Ready Source", key="save_source_button", use_container_width=True):
            if file is None:
                st.error("Please upload a source file first.")
            else:
                name = default_ready_source_name
                path = get_company_folder(company_code) / name
                with open(path, "wb") as f:
                    f.write(file.getbuffer())
                st.success(f"Saved as: {name}")
                refresh_source_file_views()
                st.rerun()

    with save_c2:
        with st.popover("Save As", use_container_width=True):
            st.text_input(
                "Source name",
                key="save_source_as_name",
                help="Choose a custom name for this source. You can include or omit the Excel extension.",
            )
            if st.button("Save As", key="save_source_as_button", use_container_width=True):
                if file is None:
                    st.error("Please upload a source file first.")
                else:
                    custom_filename = build_custom_source_filename(
                        st.session_state.get("save_source_as_name", ""),
                        file.name,
                    )
                    if not custom_filename:
                        st.error("Please enter a valid source name.")
                    else:
                        path = get_company_folder(company_code) / custom_filename
                        if path.exists():
                            st.error("A source with this name already exists. Please choose a different name.")
                        else:
                            with open(path, "wb") as f:
                                f.write(file.getbuffer())
                            st.success(f"Saved as: {custom_filename}")
                            refresh_source_file_views()
                            st.rerun()

    st.info(
        "Download the source template, fill in your products, and upload it back to the platform."
    )

    if TEMPLATE_FILE.exists():
        with open(TEMPLATE_FILE, "rb") as f:
            st.download_button(
                "Download Source Template",
                data=f.read(),
                file_name="source_template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_source_template",
                use_container_width=True,
            )
    else:
        st.warning("Template file not found.")

    st.markdown("---")
    render_source_library(show_title=True)
    st.markdown("</div>", unsafe_allow_html=True)



def render_comparisons():
    pending_name = st.session_state.get("pending_comparison_name_input")
    if pending_name is not None:
        st.session_state["comparison_name_input"] = pending_name
        st.session_state["pending_comparison_name_input"] = None

    pending_leave_name = st.session_state.get("pending_save_as_exit_name")
    if pending_leave_name is not None:
        st.session_state["save_as_exit_name"] = pending_leave_name
        st.session_state["pending_save_as_exit_name"] = None

    pending_inline_name = st.session_state.get("pending_inline_save_as_name")
    if pending_inline_name is not None:
        st.session_state["inline_save_as_name"] = pending_inline_name
        st.session_state["pending_inline_save_as_name"] = None

    st.markdown('<div class="app-card">', unsafe_allow_html=True)
    st.markdown("## Comparisons")

    mode = st.session_state.get("comparison_mode", "menu")
    if mode not in {"menu", "load", "edit"}:
        mode = "menu"
        st.session_state["comparison_mode"] = "menu"

    if mode == "load":
        st.session_state["comparison_mode"] = "menu"
        st.rerun()

    if mode == "menu":
        st.markdown("### Start")
        menu_c1, menu_c2 = st.columns(2)

        with menu_c1:
            if st.button("➕ New Comparison", use_container_width=True, key="comparison_menu_new"):
                clear_current_comparison_state()
                st.session_state["show_inline_save_options"] = False
                st.session_state["inline_save_mode"] = "menu"
                st.session_state["active_save_row_id"] = None
                st.session_state["pending_scroll_to_save_panel"] = False
                st.session_state["pending_inline_save_as_name"] = ""
                st.session_state["pending_save_as_exit_name"] = ""
                st.session_state["comparison_mode"] = "edit"
                st.session_state["show_export_preview"] = False
                st.session_state["skip_export_preview_once"] = False
                st.rerun()

        with menu_c2:
            with st.popover("📂 Load Comparison", use_container_width=True):
                comparison_file = get_current_user_comparisons_file()
                saved_records = list_comparisons(comparison_file)

                if not saved_records:
                    st.info("You have no saved comparisons yet.")
                else:
                    company_options = {
                        f"{row['name']} ({row['code']})": row["code"] for _, row in companies_df.iterrows()
                    }
                    saved_options = {
                        build_display_label(record): record["id"]
                        for record in saved_records
                    }

                    selected_saved_label = st.selectbox(
                        "Your saved comparisons",
                        [""] + list(saved_options.keys()),
                        key="selected_saved_comparison_label_popover",
                    )

                    if selected_saved_label:
                        selected_saved_id = saved_options[selected_saved_label]
                        selected_record = get_comparison(comparison_file, selected_saved_id)

                        if selected_record:
                            source_line = ", ".join(
                                [f"{k}: {v}" for k, v in selected_record.get("source_files", {}).items() if v]
                            )
                            if source_line:
                                st.caption("Source files: " + source_line)

                            selected_lock = get_comparison_lock_info(comparison_file, selected_record.get("id"))
                            if selected_lock:
                                holder = selected_lock.get("owner_name") or selected_lock.get("owner_email") or "another user"
                                if comparison_lock_owned_by_current_session(selected_lock):
                                    st.info("Editing mode active for this comparison.")
                                else:
                                    st.warning(f"🔒 Locked by {holder}")

                            load_c1, load_c2, load_c3 = st.columns(3)
                            with load_c1:
                                if st.button("Load Selected", use_container_width=True, key="load_selected_comparison_btn_popover"):
                                    state_payload = selected_record.get("state", {}) or {}
                                    missing_companies = comparison_has_missing_companies(state_payload, company_options)

                                    if missing_companies:
                                        st.warning(
                                            "This comparison cannot be loaded because some companies were deleted: "
                                            + ", ".join(missing_companies)
                                        )
                                    else:
                                        if has_real_changes_against_loaded_baseline():
                                            st.session_state["show_leave_prompt"] = True
                                            st.session_state["leave_prompt_step"] = ""
                                            st.session_state["pending_action_type"] = "load_comparison"
                                            st.session_state["pending_action_payload"] = {
                                                "state_payload": selected_record.get("state", {}) or {},
                                                "comparison_id": selected_record.get("id"),
                                                "comparison_name": selected_record.get("name", ""),
                                            }
                                            st.rerun()
                                        else:
                                            ok, msg = load_selected_comparison_record(selected_record)
                                            if ok:
                                                st.session_state["show_inline_save_options"] = False
                                                st.session_state["inline_save_mode"] = "menu"
                                                st.session_state["active_save_row_id"] = None
                                                st.session_state["pending_inline_save_as_name"] = ""
                                                st.session_state["pending_save_as_exit_name"] = ""
                                                st.session_state["comparison_mode"] = "edit"
                                                st.session_state["show_export_preview"] = False
                                                mark_comparison_clean()
                                                st.rerun()
                                            else:
                                                st.warning(msg)

                            with load_c2:
                                if st.button("📄 Duplicate", use_container_width=True, key="duplicate_selected_comparison_btn_popover"):
                                    ok, msg, _ = duplicate_saved_comparison(comparison_file, selected_record.get("id"))
                                    if ok:
                                        st.success(msg)
                                        st.rerun()
                                    else:
                                        st.warning(msg)

                            with load_c3:
                                if st.button("🗑️ Delete", use_container_width=True, key="delete_selected_comparison_btn_popover"):
                                    delete_lock = get_comparison_lock_info(comparison_file, selected_record.get("id"))
                                    if delete_lock and not comparison_lock_owned_by_current_session(delete_lock):
                                        holder = delete_lock.get("owner_name") or delete_lock.get("owner_email") or "another user"
                                        st.warning(f"Cannot delete while locked by {holder}.")
                                    else:
                                        ok = delete_comparison(comparison_file, selected_record.get("id"))
                                        release_comparison_lock(comparison_file, selected_record.get("id"))
                                        if ok:
                                            if st.session_state.get("current_comparison_id") == selected_record.get("id"):
                                                st.session_state["current_comparison_id"] = None
                                            st.success("Comparison deleted successfully.")
                                            st.rerun()
                                        else:
                                            st.warning("Could not delete comparison.")

                            save_as_default_name = str(selected_record.get("name", "") or "").strip()
                            save_as_key = f"save_as_duplicate_name_{selected_record.get('id')}"
                            if save_as_key not in st.session_state:
                                st.session_state[save_as_key] = save_as_default_name

                            st.text_input(
                                "Save As name",
                                key=save_as_key,
                                help="Enter the name for the new copied comparison.",
                            )

                            if st.button("💾 Save As", use_container_width=True, key="save_as_selected_comparison_btn_popover"):
                                target_name = str(st.session_state.get(save_as_key, "") or "").strip()
                                ok, msg, _ = duplicate_saved_comparison_with_name(
                                    comparison_file,
                                    selected_record.get("id"),
                                    target_name,
                                )
                                if ok:
                                    st.success(msg)
                                    st.rerun()
                                else:
                                    st.warning(msg)

        st.markdown("</div>", unsafe_allow_html=True)
        return

    touch_current_comparison_lock()

    current_name = st.session_state.get("comparison_name_input", "").strip()
    if current_name:
        st.info(f"📊 Working on: {current_name}")

    current_lock_id = st.session_state.get("current_comparison_id")
    if current_lock_id:
        current_lock = get_comparison_lock_info(get_current_user_comparisons_file(), current_lock_id)
        if current_lock:
            holder = current_lock.get("owner_name") or current_lock.get("owner_email") or "another user"
            if comparison_lock_owned_by_current_session(current_lock):
                st.caption("Editing mode active.")
            else:
                st.warning(f"🔒 This comparison is currently locked by {holder}.")

    loaded_msg = st.session_state.get("comparison_loaded_success_message", "")
    if loaded_msg:
        st.success(loaded_msg)
        st.session_state["comparison_loaded_success_message"] = ""

    post_save_msg = st.session_state.get("post_save_success_message", "")
    if post_save_msg:
        st.success(post_save_msg)
        st.session_state["post_save_success_message"] = ""

    company_options = {
        f"{row['name']} ({row['code']})": row["code"] for _, row in companies_df.iterrows()
    }

    current_selection = st.session_state.get("comparison_company_selection", [])
    missing_current_companies = [item for item in current_selection if item not in company_options]
    if missing_current_companies:
        st.warning(
            "Some previously selected companies were deleted and have been removed from the current comparison: "
            + ", ".join(missing_current_companies)
        )
        st.session_state["comparison_company_selection"] = [
            item for item in current_selection if item in company_options
        ]

    current_selected_displays = st.session_state.get("comparison_company_selection", [])
    current_selected_codes = [company_options[x] for x in current_selected_displays if x in company_options]

    active_bar_c1, active_bar_c2 = st.columns([1, 5])
    with active_bar_c1:
        st.empty()

    with active_bar_c2:
        active_label = st.session_state.get("active_comparison_label", "").strip()
        if active_label:
            st.info(f'📄 Active comparison: {active_label}')
        else:
            st.info("📄 New comparison")

    st.markdown("---")
    st.markdown("### 4. Select Saved Sources for Comparison")

    selected_company_displays = st.multiselect(
        "Select up to 5 companies to compare",
        options=list(company_options.keys()),
        max_selections=5,
        key="comparison_company_selection",
        on_change=mark_comparison_dirty,
    )

    selected_codes = [company_options[x] for x in selected_company_displays if x in company_options]

    if not selected_codes:
        st.info("Please select at least 1 company for comparison.")

    selected = {}
    catalogs = {}

    if selected_codes:
        source_cols_per_row = 2 if len(selected_codes) >= 4 else len(selected_codes)

        for start_idx in range(0, len(selected_codes), source_cols_per_row):
            chunk = selected_codes[start_idx:start_idx + source_cols_per_row]
            cols = st.columns(len(chunk))

            for i, code in enumerate(chunk):
                with cols[i]:
                    files = get_company_files(code)
                    selected[code] = st.selectbox(
                        f"{code} source file",
                        [""] + files,
                        key=f"select_{code}",
                        on_change=lambda c=code: (_handle_source_selection_change(c), mark_comparison_dirty()),
                    )

        for code in selected_codes:
            if selected.get(code):
                source_path = get_company_folder(code) / selected[code]
                try:
                    catalogs[code] = load_prepared_catalog_from_file(str(source_path), source_path.stat().st_mtime)
                except Exception as e:
                    st.error(f"Error loading file: {e}")
                    catalogs[code] = None
            else:
                catalogs[code] = None

        _run_render_loop_suggestions(selected_codes, catalogs)


    if selected_codes:
        st.markdown("---")
        st.markdown("### Apply Specific Discount to All Products")
        st.caption(
            "Choose one discount slot and one value for each company. It will be applied to all current rows of that company only."
        )

        bulk_cols_per_row = 2 if len(selected_codes) >= 4 else len(selected_codes)
        for start_idx in range(0, len(selected_codes), bulk_cols_per_row):
            chunk = selected_codes[start_idx:start_idx + bulk_cols_per_row]
            bulk_cols = st.columns(len(chunk))

            for i, code in enumerate(chunk):
                with bulk_cols[i]:
                    label = get_company_label(code)

                    selector_key = f"bulk_discount_slot_{code}"
                    value_key = f"bulk_discount_value_{code}"

                    if selector_key not in st.session_state:
                        st.session_state[selector_key] = "Disc1"
                    if value_key not in st.session_state:
                        st.session_state[value_key] = 0.0

                    st.markdown(f"**{label}**")
                    st.selectbox(
                        f"{label} discount slot",
                        ["Disc1", "Disc2", "Disc3", "Disc4", "Disc5"],
                        key=selector_key,
                    )
                    st.number_input(
                        f"{label} discount value %",
                        min_value=0.0,
                        max_value=100.0,
                        step=0.1,
                        key=value_key,
                    )

                    if st.button(
                        f"Apply to all {label} rows",
                        key=f"apply_bulk_discount_{code}",
                        use_container_width=True,
                    ):
                        disc_index = int(str(st.session_state[selector_key]).replace("Disc", ""))
                        disc_value = float(st.session_state[value_key])
                        apply_specific_discount_to_all_rows(selected_codes, code, disc_index, disc_value)
                        mark_comparison_dirty()
                        st.session_state["bulk_discount_success_message"] = (
                            f"{label}: {st.session_state[selector_key]} set to {disc_value:.2f}% for all current rows."
                        )
                        st.rerun()

        success_message = st.session_state.get("bulk_discount_success_message", "")
        if success_message:
            st.success(success_message)
            st.session_state["bulk_discount_success_message"] = ""

    if selected_codes:
        st.markdown("---")
        st.markdown("### Carry Discounts Forward")
        st.caption(
            "When enabled, each new row starts with the previous row's discounts for that company. If a row is still blank, turning this on will also prefill it. You can still edit any discount normally."
        )
        carry_cols_per_row = 2 if len(selected_codes) >= 4 else len(selected_codes)
        for start_idx in range(0, len(selected_codes), carry_cols_per_row):
            chunk = selected_codes[start_idx:start_idx + carry_cols_per_row]
            carry_cols = st.columns(len(chunk))
            for i, code in enumerate(chunk):
                with carry_cols[i]:
                    label = get_company_label(code)
                    st.checkbox(
                        f"{label}: use previous row discounts",
                        key=f"carry_forward_{code}",
                        on_change=mark_comparison_dirty,
                    )

        st.markdown("")
        st.checkbox(
            "Smart Product Matching (Beta)",
            key="smart_matching_enabled",
            help="When enabled, selecting a product in the first company can suggest equivalent products for the other selected companies based on learned comparison history and supporting product signals. Suggestions remain fully editable.",
        )
        st.caption(
            "This feature is optional. Suggestions never replace the sales user's judgment and remain fully editable."
        )

    st.markdown("---")
    st.markdown("### 6. Multi-Line Comparison")

    if not selected_codes:
        st.info("Select companies first to start comparison.")
    else:
        _backfill_match_history_from_saved_comparisons()
        st.info(f"Current rows: {len(st.session_state.row_ids)}")

        row_ids_list = list(st.session_state.get("row_ids", []))
        if row_ids_list:
            if st.session_state.get("active_row_id") not in row_ids_list:
                st.session_state["active_row_id"] = row_ids_list[-1]

        for visible_index, row_id in enumerate(st.session_state.row_ids):
            ensure_discount_defaults_for_row(row_id, selected_codes)
            st.markdown(f'<div id="row-anchor-{row_id}"></div>', unsafe_allow_html=True)

            row_summary = get_row_summary_text(row_id, selected_codes, catalogs)
            row_ids_list = list(st.session_state.get("row_ids", []))
            last_row_id = row_ids_list[-1] if row_ids_list else row_id
            is_active = row_id == st.session_state.get("active_row_id")

            if not is_active:
                summary_c1, summary_c2, summary_c3, summary_c4 = st.columns([4.0, 1.1, 1.3, 1.1])

                with summary_c1:
                    st.markdown(f"#### Row {visible_index + 1} — {row_summary}")

                with summary_c2:
                    if st.button("Open Row", key=f"open_row_{row_id}", use_container_width=True):
                        st.session_state["active_row_id"] = row_id
                        st.session_state["pending_focus_row_id"] = row_id
                        st.rerun()

                with summary_c3:
                    if st.button("Add Row Below", key=f"add_row_after_summary_{row_id}", use_container_width=True):
                        add_comparison_row(selected_codes, insert_after_row_id=row_id)
                        st.session_state["active_row_id"] = st.session_state["row_ids"][-1]
                        st.session_state["comparison_dirty"] = True
                        st.session_state["comparison_user_modified"] = True
                        st.session_state["export_preview_cache_signature"] = ""
                        st.rerun()

                with summary_c4:
                    if st.button("Delete Row", key=f"delete_row_summary_{row_id}", use_container_width=True):
                        st.session_state.row_ids = [
                            r for r in st.session_state.row_ids if r != row_id
                        ]
                        remaining_rows = st.session_state.get("row_ids", [])
                        if remaining_rows:
                            current_active = st.session_state.get("active_row_id")
                            if current_active == row_id or current_active not in remaining_rows:
                                st.session_state["active_row_id"] = remaining_rows[-1]
                        else:
                            st.session_state["row_ids"] = [1]
                            st.session_state["next_row_id"] = max(int(st.session_state.get("next_row_id", 2)), 2)
                            st.session_state["active_row_id"] = None
                        st.session_state["comparison_dirty"] = True
                        st.session_state["comparison_user_modified"] = True
                        st.rerun()

                render_row_navigation_buttons(row_id, visible_index)
                st.markdown("---")
                continue

            st.markdown(f"#### Row {visible_index + 1} — {row_summary}")
            row_final_prices = {}
            comparison_cols_per_row = 2 if len(selected_codes) >= 4 else len(selected_codes)

            for start_idx in range(0, len(selected_codes), comparison_cols_per_row):
                chunk = selected_codes[start_idx:start_idx + comparison_cols_per_row]
                row_cols = st.columns(len(chunk))

                for col_idx, code in enumerate(chunk):
                    with row_cols[col_idx]:
                        label = get_company_label(code)

                        st.write(f"**{label}**")

                        df = catalogs.get(code)
                        if df is not None and not df.empty:
                            options = [""] + df["DISPLAY"].tolist()
                            product_widget_key = get_product_widget_key(row_id, code)
                            mirror_product_data_to_widget(row_id, code)
                            st.selectbox(
                                f"{label} product",
                                options,
                                key=product_widget_key,
                                on_change=lambda r=row_id, c=code, sc=selected_codes, cats=catalogs: (
                                    sync_product_widget_to_data(r, c),
                                    _apply_smart_product_suggestions_for_row(r, sc, cats) if (len(sc) > 0 and c == sc[0] and st.session_state.get("smart_matching_enabled", False)) else None,
                                    mark_comparison_dirty()
                                ),
                            )
                            selected_product = st.session_state.get(f"row_{row_id}_{code}_product", "")
                            row = get_catalog_row(df, selected_product)

                            smart_note = st.session_state.get("smart_match_notes", {}).get(f"{row_id}|{code}", "")
                            smart_score = st.session_state.get("smart_match_scores", {}).get(f"{row_id}|{code}", "")
                            smart_confidence = str(st.session_state.get("smart_match_confidence", {}).get(f"{row_id}|{code}", "") or "").upper()
                            smart_mode = str(st.session_state.get("smart_match_mode", {}).get(f"{row_id}|{code}", "") or "")
                            smart_target_display = st.session_state.get("smart_match_target_display", {}).get(f"{row_id}|{code}", "")
                            if smart_note:
                                if smart_note == "No match found":
                                    st.caption("No match found by Smart Matching")
                                else:
                                    target_hint = ""
                                    if smart_note == "Better match available" and smart_target_display:
                                        target_hint = f" → {smart_target_display}"
                                    mode_hint = f" [{smart_mode}]" if smart_mode else ""
                                    confidence_hint = f" [{smart_confidence}]" if smart_confidence else ""
                                    try:
                                        st.caption(f"{smart_note} by Smart Matching{confidence_hint}{mode_hint}{target_hint} (score: {float(smart_score):.1f})")
                                    except Exception:
                                        st.caption(f"{smart_note} by Smart Matching{confidence_hint}{mode_hint}{target_hint}")

                            if row is not None:
                                st.write("SAP:", row["SAP"])
                                st.write("MM:", row["MM"])
                                st.write("Package:", row["Package"])
                                st.write("Base Price:", round(float(row["Price"]), 2))

                                discs = []
                                for j in range(1, 6):
                                    data_key = f"row_{row_id}_{code}_disc_{j}"
                                    widget_key = get_discount_widget_key(row_id, code, j)
                                    mirror_discount_data_to_widget(row_id, code, j)
                                    disc_val = st.number_input(
                                        f"{label} Disc {j}",
                                        min_value=0.0,
                                        max_value=100.0,
                                        step=0.1,
                                        key=widget_key,
                                        on_change=lambda r=row_id, c=code, d=j: (sync_discount_widget_to_data(r, c, d), mark_comparison_dirty()),
                                    )
                                    st.session_state[data_key] = float(disc_val)
                                    discs.append(disc_val)

                                manual_final_widget_key = get_manual_final_price_widget_key(row_id, code)
                                mirror_manual_final_price_data_to_widget(row_id, code)
                                manual_final_text = st.text_input(
                                    f"{label} Manual Final Price",
                                    key=manual_final_widget_key,
                                    placeholder="Optional manual final price",
                                    on_change=lambda r=row_id, c=code: (sync_manual_final_price_widget_to_data(r, c), mark_comparison_dirty()),
                                )
                                st.session_state[get_manual_final_price_data_key(row_id, code)] = str(manual_final_text or "").strip()

                                final, final_is_manual = get_effective_final_price(row_id, code, row["Price"], discs)
                                row_final_prices[code] = final
                                if final_is_manual:
                                    st.success(f"Final Price (Manual): {final}")
                                else:
                                    st.success(f"Final Price: {final}")
                            else:
                                for j in range(1, 6):
                                    widget_key = get_discount_widget_key(row_id, code, j)
                                    mirror_discount_data_to_widget(row_id, code, j)
                                    disc_val = st.number_input(
                                        f"{label} Disc {j}",
                                        min_value=0.0,
                                        max_value=100.0,
                                        step=0.1,
                                        key=widget_key,
                                        on_change=lambda r=row_id, c=code, d=j: (sync_discount_widget_to_data(r, c, d), mark_comparison_dirty()),
                                    )
                                    st.session_state[f"row_{row_id}_{code}_disc_{j}"] = float(disc_val)
                                manual_final_widget_key = get_manual_final_price_widget_key(row_id, code)
                                mirror_manual_final_price_data_to_widget(row_id, code)
                                manual_final_text = st.text_input(
                                    f"{label} Manual Final Price",
                                    key=manual_final_widget_key,
                                    placeholder="Optional manual final price",
                                    on_change=lambda r=row_id, c=code: (sync_manual_final_price_widget_to_data(r, c), mark_comparison_dirty()),
                                )
                                st.session_state[get_manual_final_price_data_key(row_id, code)] = str(manual_final_text or "").strip()
                                row_final_prices[code] = None
                                st.info("No product selected")
                        else:
                            row_final_prices[code] = None
                            st.info("No data")

            if row_final_prices:
                valid = {k: v for k, v in row_final_prices.items() if v is not None}
                if valid:
                    best_code = min(valid, key=valid.get)
                    best_label = get_company_label(best_code)
                else:
                    best_label = "-"

                action_cols = st.columns([3.2, 1.2, 1.2])

                with action_cols[0]:
                    st.metric(f"Row {visible_index + 1} Best Price", best_label)

                with action_cols[1]:
                    st.write("")
                    st.write("")
                    if st.button(
                        "Add Row Below",
                        key=f"add_row_after_{row_id}",
                        use_container_width=True,
                    ):
                        add_comparison_row(selected_codes, insert_after_row_id=row_id)
                        st.session_state["active_row_id"] = st.session_state["row_ids"][-1]
                        st.session_state["comparison_dirty"] = True
                        st.session_state["comparison_user_modified"] = True
                        st.session_state["skip_export_preview_once"] = True
                        st.rerun()

                    render_row_navigation_buttons(row_id, visible_index)

                with action_cols[2]:
                    st.write("")
                    st.write("")
                    if st.button(
                        "Delete This Row",
                        key=f"delete_row_{row_id}",
                        use_container_width=True,
                    ):
                        st.session_state.row_ids = [
                            r for r in st.session_state.row_ids if r != row_id
                        ]
                        st.session_state["comparison_dirty"] = True
                        st.session_state["comparison_user_modified"] = True
                        st.rerun()

                if visible_index == len(st.session_state.get("row_ids", [])) - 1:
                    row_save_cols = st.columns([3.2, 1.2, 1.2])

                    with row_save_cols[1]:
                        if st.button(
                            "⬅ Menu",
                            key=f"row_menu_back_{row_id}",
                            use_container_width=True,
                        ):
                            if has_unsaved_comparison_changes():
                                open_leave_prompt("switch_view", target_view="Comparisons")
                                st.rerun()
                            else:
                                st.session_state["comparison_mode"] = "menu"
                                st.session_state["show_saved_comparisons"] = False
                                st.rerun()

                    with row_save_cols[2]:
                        with st.popover("Save", use_container_width=True):
                            current_name = st.session_state.get("active_comparison_label", "").strip() or st.session_state.get("comparison_name_input", "").strip()

                            save_pop_c1, save_pop_c2 = st.columns(2)
                            with save_pop_c1:
                                if st.button("Save", use_container_width=True, key=f"save_existing_popover_{row_id}"):
                                    if not st.session_state.get("current_comparison_id"):
                                        st.warning("This comparison has not been saved before. Use Save As.")
                                    elif not current_name:
                                        st.warning("No active comparison name found. Use Save As.")
                                    else:
                                        ok, msg = save_current_comparison_from_state(force_new=False, override_name=current_name)
                                        if ok:
                                            st.session_state["comparison_dirty"] = False
                                            st.session_state["comparison_user_modified"] = False
                                            st.session_state["post_save_success_message"] = "Comparison saved successfully."
                                            mark_comparison_clean()
                                            st.rerun()
                                        else:
                                            st.warning(msg)

                            with save_pop_c2:
                                st.text_input("New name for Save As", key=f"save_as_name_popover_{row_id}")
                                if st.button("Save As", use_container_width=True, key=f"save_as_popover_{row_id}"):
                                    new_name = st.session_state.get(f"save_as_name_popover_{row_id}", "").strip()
                                    if not new_name:
                                        st.warning("Please enter a new name.")
                                    else:
                                        ok, msg = save_current_comparison_from_state(force_new=True, override_name=new_name)
                                        if ok:
                                            st.session_state["comparison_dirty"] = False
                                            st.session_state["comparison_user_modified"] = False
                                            st.session_state["post_save_success_message"] = "Saved as new comparison."
                                            mark_comparison_clean()
                                            st.rerun()
                                        else:
                                            st.warning(msg)

                comparison_summaries = build_comparison_summary(
                    row_final_prices, selected_codes
                )

                has_comparisons = any(v for v in comparison_summaries.values())
                if has_comparisons:
                    st.markdown("**Price Difference %**")
                    for code in selected_codes:
                        label = get_company_label(code)
                        summary = comparison_summaries.get(code, "")
                        if summary:
                            st.write(f"**{label}:** {summary}")

            st.markdown("---")
        active_save_row_id = st.session_state.get("active_save_row_id")
        if False and st.session_state.get("show_inline_save_options") and active_save_row_id in st.session_state.get("row_ids", []):
            st.markdown("---")
            st.markdown(f'<div id="save-panel-anchor-{active_save_row_id}"></div>', unsafe_allow_html=True)
            st.markdown('<div class="save-panel-card">', unsafe_allow_html=True)
            st.markdown("### Save Comparison")

            save_mode = st.session_state.get("inline_save_mode", "menu")

            if save_mode == "menu":
                menu_save_c1, menu_save_c2 = st.columns(2)

                with menu_save_c1:
                    if st.button("Save", use_container_width=True, key=f"inline_save_menu_save_{active_save_row_id}"):
                        current_name = st.session_state.get("active_comparison_label", "").strip() or st.session_state.get("comparison_name_input", "").strip()
                        if not st.session_state.get("current_comparison_id"):
                            st.warning("This comparison has not been saved before. Use Save As.")
                        elif not current_name:
                            st.warning("No active comparison name found. Use Save As.")
                        else:
                            ok, msg = save_current_comparison_from_state(force_new=False, override_name=current_name)
                            if ok:
                                st.session_state["show_inline_save_options"] = False
                                st.session_state["inline_save_mode"] = "menu"
                                st.session_state["active_save_row_id"] = None
                                st.session_state["comparison_dirty"] = False
                                st.session_state["comparison_user_modified"] = False
                                mark_comparison_clean()
                                st.success("Comparison saved successfully.")
                            else:
                                st.warning(msg)

                with menu_save_c2:
                    if st.button("Save As", use_container_width=True, key=f"inline_save_menu_save_as_{active_save_row_id}"):
                        st.session_state["inline_save_mode"] = "save_as"
                        st.rerun()

            else:
                back_c1, back_c2 = st.columns([1, 5])
                with back_c1:
                    if st.button("⬅ Back", use_container_width=True, key=f"inline_save_as_back_{active_save_row_id}"):
                        st.session_state["inline_save_mode"] = "menu"
                        st.rerun()
                with back_c2:
                    st.empty()

                st.text_input("New name for Save As", key="inline_save_as_name")
                if st.button("Save As", use_container_width=True, key=f"inline_save_as_btn_{active_save_row_id}"):
                    new_name = st.session_state.get("inline_save_as_name", "").strip()
                    if not new_name:
                        st.warning("Please enter a new name.")
                    else:
                        ok, msg = save_current_comparison_from_state(force_new=True, override_name=new_name)
                        if ok:
                            st.session_state["show_inline_save_options"] = False
                            st.session_state["inline_save_mode"] = "menu"
                            st.session_state["active_save_row_id"] = None
                            st.session_state["pending_scroll_to_save_panel"] = False
                            st.session_state["comparison_dirty"] = False
                            st.session_state["comparison_user_modified"] = False
                            st.session_state["pending_inline_save_as_name"] = ""
                            mark_comparison_clean()
                            st.success("Saved as new comparison.")
                        else:
                            st.warning(msg)

            st.markdown("</div>", unsafe_allow_html=True)

    target_row_id = st.session_state.get("pending_focus_row_id")
    if target_row_id is not None:
        components.html(
            f"""
            <script>
            const scrollToNewRow = () => {{
                const el = window.parent.document.getElementById("row-anchor-{target_row_id}");
                if (el) {{
                    el.scrollIntoView({{behavior: "smooth", block: "start"}});
                }}
            }};
            requestAnimationFrame(scrollToNewRow);
            setTimeout(scrollToNewRow, 60);
            </script>
            """,
            height=0,
        )
        st.session_state["pending_focus_row_id"] = None

    render_export_inside_comparisons(catalogs, selected_codes)
    st.markdown("</div>", unsafe_allow_html=True)

def render_export_inside_comparisons(catalogs, selected_codes):
    st.markdown("---")
    st.markdown("### 7. Export Excel Report")

    full_export_df = build_export_dataframe(
        st.session_state.row_ids, catalogs, selected_codes
    )

    selected_export_fields = st.multiselect(
        "Choose columns for Excel export",
        options=EXPORT_FIELD_OPTIONS,
        key="selected_export_fields",
    )

    if not selected_export_fields:
        st.warning("Please select at least one export field.")
        return

    export_df = filter_export_dataframe(
        full_export_df,
        selected_codes,
        selected_export_fields,
        companies_df,
    )

    if not export_df.empty:
        st.dataframe(export_df, use_container_width=True, hide_index=True)

        excel_bytes = to_excel_bytes(export_df)
        st.download_button(
            "Download Excel Report",
            data=excel_bytes,
            file_name="comparison_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_excel_report",
            use_container_width=True,
        )
    else:
        st.info("No data available for export yet.")


def _dataframe_to_excel_bytes(df: pd.DataFrame) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Learned Matches")
    return buffer.getvalue()


def _load_seed_rows_from_uploaded_file(uploaded_file) -> pd.DataFrame:
    if uploaded_file is None:
        raise ValueError("No seed file uploaded.")

    name = str(getattr(uploaded_file, "name", "") or "").lower()
    if name.endswith(".csv"):
        df = pd.read_csv(uploaded_file)
    else:
        workbook = pd.ExcelFile(uploaded_file)
        preferred_sheet = "Seed_Long_Format" if "Seed_Long_Format" in workbook.sheet_names else workbook.sheet_names[0]
        df = pd.read_excel(workbook, sheet_name=preferred_sheet)

    df.columns = [str(c).strip() for c in df.columns]
    required = {"group_id", "company", "product_name"}
    missing = sorted(required - set(df.columns))
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    for col in ["category", "subcategory", "standard_class", "confidence", "notes", "match_status"]:
        if col not in df.columns:
            df[col] = ""

    df = df.fillna("")
    df["group_id"] = df["group_id"].astype(str).str.strip()
    df["company"] = df["company"].astype(str).str.strip().str.upper()
    df["product_name"] = df["product_name"].astype(str).str.strip()
    df = df[(df["group_id"] != "") & (df["company"] != "") & (df["product_name"] != "")].copy()
    return df


def import_admin_seed_matches_from_uploaded_file(uploaded_file) -> tuple[bool, str]:
    try:
        df = _load_seed_rows_from_uploaded_file(uploaded_file)
    except Exception as exc:
        return False, str(exc)

    if df.empty:
        return False, "The uploaded seed file does not contain any usable rows."

    normalized_rows = []
    for _, row in df.iterrows():
        match_status = str(row.get("match_status", "") or "").strip().upper()
        confidence = str(row.get("confidence", "high") or "high").strip().lower() or "high"
        if match_status and match_status not in {"MATCH", "PROXY", "UNIQUE"}:
            match_status = ""
        normalized_rows.append({
            "group_id": str(row.get("group_id", "") or "").strip(),
            "category": str(row.get("category", "") or row.get("subcategory", "") or "").strip(),
            "subcategory": str(row.get("subcategory", "") or "").strip(),
            "company": str(row.get("company", "") or "").strip().upper(),
            "product_name": str(row.get("product_name", "") or "").strip(),
            "confidence": confidence,
            "notes": str(row.get("notes", "") or "").strip(),
            "match_status": match_status,
        })

    CENTRAL_ENGINE.import_seed_rows(normalized_rows)
    return True, f"Imported {len(normalized_rows)} seed rows across {df['group_id'].nunique()} groups."


def render_admin_seed_match_panel():
    st.markdown("### 🌱 Admin Seed Match Knowledge")
    st.caption("Upload a long-format seed mapping file to inject high-confidence baseline matches for all users. This does not require full price lists.")

    uploaded_seed_file = st.file_uploader(
        "Upload seed match file (.xlsx or .csv)",
        type=["xlsx", "xlsm", "csv"],
        key="admin_seed_match_file_uploader",
        help="Required columns: group_id, company, product_name. Recommended sheet name: Seed_Long_Format.",
    )

    action_c1, action_c2, action_c3 = st.columns([1, 1, 3])
    with action_c1:
        if st.button("Import Seed Matches", use_container_width=True, key="admin_import_seed_matches"):
            if uploaded_seed_file is None:
                st.warning("Upload a seed match file first.")
            else:
                with st.spinner("Importing admin seed matches..."):
                    ok, msg = import_admin_seed_matches_from_uploaded_file(uploaded_seed_file)
                if ok:
                    st.success(msg)
                else:
                    st.error(msg)
    with action_c2:
        current_data = CENTRAL_ENGINE.load()
        seed_rows = current_data.get("seed", {}) if isinstance(current_data, dict) else {}
        has_seed = bool(seed_rows)
        if st.button("Clear Seed Matches", use_container_width=True, key="admin_clear_seed_matches"):
            data = CENTRAL_ENGINE.load()
            data["seed"] = {}
            CENTRAL_ENGINE.save(data)
            st.success("Admin seed matches cleared successfully.")
    with action_c3:
        seed_data = CENTRAL_ENGINE.load().get("seed", {})
        seed_pair_count = 0
        if isinstance(seed_data, dict):
            for _source_key, company_map in seed_data.items():
                if not isinstance(company_map, dict):
                    continue
                for _company, target_map in company_map.items():
                    if isinstance(target_map, dict):
                        seed_pair_count += len(target_map)
        st.caption(f"Current admin seed pairs: {seed_pair_count}")

def render_admin_match_table_viewer():
    show_key = "admin_show_learned_matches"
    if show_key not in st.session_state:
        st.session_state[show_key] = False

    c1, c2, c3 = st.columns([1, 1, 4])
    with c1:
        if st.button("Show Learned Matches", use_container_width=True, key="admin_show_learned_matches_button"):
            st.session_state[show_key] = True
    with c2:
        if st.session_state.get(show_key):
            if st.button("Hide Learned Matches", use_container_width=True, key="admin_hide_learned_matches_button"):
                st.session_state[show_key] = False

    if not st.session_state.get(show_key):
        return

    rows = CENTRAL_ENGINE.export_rows()
    if not rows:
        st.caption("No learned matches available yet.")
        return

    df = pd.DataFrame(rows)
    display_df = df.rename(columns={
        "source_company": "Source Company",
        "source_product": "Source Product",
        "target_company": "Target Company",
        "matched_product": "Matched Product",
        "hits": "Hits",
        "confidence": "Confidence",
        "table_source": "Source",
    })

    f1, f2, f3, f4 = st.columns([2, 1, 1, 1])
    with f1:
        search_value = st.text_input("Search learned matches", key="admin_match_table_search", placeholder="Search company or product...")
    with f2:
        source_company_options = ["All"] + sorted([v for v in display_df["Source Company"].dropna().astype(str).unique().tolist() if v])
        selected_source_company = st.selectbox("Source company", source_company_options, key="admin_match_source_company_filter")
    with f3:
        target_company_options = ["All"] + sorted([v for v in display_df["Target Company"].dropna().astype(str).unique().tolist() if v])
        selected_target_company = st.selectbox("Target company", target_company_options, key="admin_match_target_company_filter")
    with f4:
        source_type_options = ["All", "stable", "register"]
        selected_source_type = st.selectbox("Table source", source_type_options, key="admin_match_table_source_filter")

    f5, f6, f7 = st.columns([1, 1, 2])
    with f5:
        confidence_options = ["All", "high", "medium", "low"]
        selected_confidence = st.selectbox("Confidence", confidence_options, key="admin_match_confidence_filter")
    with f6:
        sort_options = {
            "Hits ↓": ["Hits", False],
            "Hits ↑": ["Hits", True],
            "Source Company A→Z": ["Source Company", True],
            "Target Company A→Z": ["Target Company", True],
            "Confidence": ["Confidence", True],
        }
        selected_sort_label = st.selectbox("Sort by", list(sort_options.keys()), key="admin_match_sort_by")
    with f7:
        st.caption(f"Rows: {len(display_df)}")

    filtered_df = display_df.copy()
    if search_value:
        q = str(search_value).strip().lower()
        mask = (
            filtered_df["Source Company"].astype(str).str.lower().str.contains(q, na=False)
            | filtered_df["Source Product"].astype(str).str.lower().str.contains(q, na=False)
            | filtered_df["Target Company"].astype(str).str.lower().str.contains(q, na=False)
            | filtered_df["Matched Product"].astype(str).str.lower().str.contains(q, na=False)
        )
        filtered_df = filtered_df[mask]

    if selected_source_company != "All":
        filtered_df = filtered_df[filtered_df["Source Company"] == selected_source_company]
    if selected_target_company != "All":
        filtered_df = filtered_df[filtered_df["Target Company"] == selected_target_company]
    if selected_source_type != "All":
        filtered_df = filtered_df[filtered_df["Source"] == selected_source_type]
    if selected_confidence != "All":
        filtered_df = filtered_df[filtered_df["Confidence"] == selected_confidence]

    sort_col, ascending = sort_options[selected_sort_label]
    if sort_col == "Confidence":
        order = {"high": 0, "medium": 1, "low": 2}
        filtered_df = filtered_df.assign(__confidence_rank=filtered_df["Confidence"].map(order).fillna(9))
        filtered_df = filtered_df.sort_values(["__confidence_rank", "Hits"], ascending=[True, False]).drop(columns=["__confidence_rank"])
    else:
        filtered_df = filtered_df.sort_values(sort_col, ascending=ascending)

    export_df = filtered_df[["Source Company", "Source Product", "Target Company", "Matched Product", "Hits", "Confidence", "Source"]].copy()

    st.download_button(
        "Download Learned Matches as Excel",
        data=_dataframe_to_excel_bytes(export_df),
        file_name=f"learned_matches_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="admin_download_learned_matches_excel",
        use_container_width=False,
    )

    st.dataframe(export_df, use_container_width=True, hide_index=True)


def render_admin_panel():
    if not is_admin_user():
        return

    st.markdown('<div class="app-card">', unsafe_allow_html=True)
    st.markdown("## 8. Admin Panel")

    active_lock_rows = get_all_active_comparison_locks()
    if active_lock_rows:
        st.markdown("### Active Comparison Locks")
        st.dataframe(pd.DataFrame(active_lock_rows), use_container_width=True)
    else:
        st.caption("No active comparison locks.")

    st.markdown("### 🧠 Smart Matching")
    smart_c1, smart_c2, smart_c3 = st.columns([1, 1, 2])
    with smart_c1:
        if st.button("Rebuild Smart Matching History", use_container_width=True, key="admin_rebuild_smart_matching_history"):
            with st.spinner("Rebuilding history from saved comparisons..."):
                rebuild_match_history_from_scratch()
            st.success("Smart Matching history rebuilt successfully.")
    with smart_c2:
        if st.button("Rebuild Central Match Table", use_container_width=True, key="admin_rebuild_central_match_table"):
            with st.spinner("Rebuilding central match table from saved comparisons..."):
                rebuild_central_match_table_from_comparisons()
            st.success("Central match table rebuilt successfully.")
    with smart_c3:
        st.caption("Runtime path: 1) admin seed direct lookup, 2) central table direct lookup, 3) simple clean fallback. Re-evaluation happens every 10 saved events and restrictions apply there.")

    render_admin_seed_match_panel()
    render_admin_match_table_viewer()

    st.markdown("### 📁 Source Files Per User")

    admin_source_rows = []
    for workspace_dir in sorted(PERSIST_ROOT.iterdir()):
        if not workspace_dir.is_dir() or workspace_dir.name == "_admin":
            continue

        uploads_dir = workspace_dir / "uploads"
        if not uploads_dir.exists():
            continue

        for company_dir in sorted(uploads_dir.iterdir()):
            if not company_dir.is_dir():
                continue

            for f in sorted(company_dir.glob("*.*"), reverse=True):
                if f.suffix.lower() in [".xlsx", ".xlsm"]:
                    admin_source_rows.append(
                        {
                            "workspace": workspace_dir.name,
                            "company": company_dir.name,
                            "path": f,
                            "filename": f.name,
                            "modified": datetime.fromtimestamp(f.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
                        }
                    )

    if admin_source_rows:
        users_registry = load_users_registry()
        available_users = sorted({item["workspace"] for item in admin_source_rows})
        workspace_label_map = {
            workspace_name: resolve_workspace_label(workspace_name, users_registry)
            for workspace_name in available_users
        }

        selected_source_user = st.selectbox(
            "Source files per user",
            options=available_users,
            key="admin_selected_source_user",
            format_func=lambda x: workspace_label_map.get(x, x),
        )

        selected_rows = [item for item in admin_source_rows if item["workspace"] == selected_source_user]

        if selected_rows:
            selected_user_label = workspace_label_map.get(selected_source_user, selected_source_user)
            st.caption(f"User: {selected_user_label}")
            for item in selected_rows:
                with open(item["path"], "rb") as file_data:
                    st.download_button(
                        label=f"{item['company']} | {item['filename']} ({item['modified']})",
                        data=file_data.read(),
                        file_name=item["filename"],
                        key=f"admin_download_{item['workspace']}_{item['company']}_{item['filename']}",
                        use_container_width=True,
                    )
        else:
            st.caption("No source files found for this user.")
    else:
        st.caption("No source files found.")

    users_registry = load_users_registry()
    if users_registry:
        users_for_view = []
        for row in users_registry:
            sessions_count = len(row.get("active_sessions", []))
            users_for_view.append(
                {
                    "Email": row.get("email", ""),
                    "Name": row.get("name", ""),
                    "Status": row.get("status", "pending"),
                    "Billing": row.get("billing_status", "trialing"),
                    "Premium": row.get("is_premium", False),
                    "Company": row.get("company_name", "") or "",
                    "Role": row.get("role", "member"),
                    "Active Sessions": f"{sessions_count}/{get_user_max_active_sessions(row)}",
                    "Session Details": summarize_active_sessions(row.get("active_sessions", [])),
                        "Max Sessions": get_user_max_active_sessions(row),
                    "First Seen": row.get("first_seen", ""),
                    "Last Login": row.get("last_login", ""),
                    "Last Seen": row.get("last_seen", ""),
                    "Online": online_status_from_last_seen(row.get("last_seen", "")),
                    "Sub": row.get("sub", ""),
                }
            )

        users_df = pd.DataFrame(users_for_view)
        st.dataframe(users_df, use_container_width=True, hide_index=True)

        user_options = {}
        for row in users_registry:
            label = f"{row.get('email', '')} | {row.get('status', '')}"
            user_options[label] = row

        selected_user_label = st.selectbox(
            "Select user",
            [""] + list(user_options.keys()),
            key="admin_selected_user_to_manage",
        )

        selected_user_row = user_options.get(selected_user_label) if selected_user_label else None

        session_override_value = 0
        if selected_user_row:
            raw_override = selected_user_row.get("max_active_sessions_override")
            try:
                session_override_value = int(raw_override) if raw_override is not None else 0
            except Exception:
                session_override_value = 0

        st.markdown("#### User Session Limit")
        session_limit_cols = st.columns([2, 1, 1])
        with session_limit_cols[0]:
            admin_user_max_sessions = st.number_input(
                "User max sessions override (0 = default 2)",
                min_value=0,
                max_value=20,
                value=session_override_value,
                step=1,
                key="admin_user_max_sessions_override",
            )
        with session_limit_cols[1]:
            st.write("")
            st.write("")
            if st.button("Save Sessions Limit", key="save_user_sessions_limit_button", use_container_width=True):
                if selected_user_row:
                    set_user_max_active_sessions_override(
                        selected_user_row.get("email", ""),
                        selected_user_row.get("sub", ""),
                        int(admin_user_max_sessions),
                    )
                    st.success(f"Session limit updated for: {selected_user_row.get('email', '')}")
                    st.rerun()
        with session_limit_cols[2]:
            st.write("")
            st.write("")
            if st.button("Reset to Default", key="reset_user_sessions_limit_button", use_container_width=True):
                if selected_user_row:
                    set_user_max_active_sessions_override(
                        selected_user_row.get("email", ""),
                        selected_user_row.get("sub", ""),
                        0,
                    )
                    st.success(f"Session limit reset to default for: {selected_user_row.get('email', '')}")
                    st.rerun()

        c1, c2, c3, c4, c5, c6, c7, c8 = st.columns(8)

        with c1:
            if st.button(
                "Unblock User", key="unblock_user_button", use_container_width=True
            ):
                if selected_user_label:
                    row = user_options[selected_user_label]
                    set_user_status(
                        row.get("email", ""),
                        row.get("sub", ""),
                        "approved",
                    )
                    st.success(f"User restored: {row.get('email', '')}")
                    st.rerun()

        with c2:
            if st.button(
                "Block User", key="block_user_button", use_container_width=True
            ):
                if selected_user_label:
                    row = user_options[selected_user_label]
                    set_user_status(
                        row.get("email", ""),
                        row.get("sub", ""),
                        "blocked",
                    )
                    st.success(f"Blocked: {row.get('email', '')}")
                    st.rerun()

        with c3:
            if st.button(
                "Reset to Trial", key="reset_to_trial_button", use_container_width=True
            ):
                if selected_user_label:
                    row = user_options[selected_user_label]
                    reset_user_to_trial(
                        row.get("email", ""),
                        row.get("sub", ""),
                    )
                    st.success(f"Trial reset for: {row.get('email', '')}")
                    st.rerun()

        with c4:
            if st.button(
                "Give Premium", key="give_premium_button", use_container_width=True
            ):
                if selected_user_label:
                    row = user_options[selected_user_label]
                    set_user_premium(
                        row.get("email", ""),
                        row.get("sub", ""),
                        True,
                    )
                    st.success(f"Premium granted to: {row.get('email', '')}")
                    st.rerun()

        with c5:
            if st.button(
                "Remove Premium",
                key="remove_premium_button",
                use_container_width=True,
            ):
                if selected_user_label:
                    row = user_options[selected_user_label]
                    set_user_premium(
                        row.get("email", ""),
                        row.get("sub", ""),
                        False,
                    )
                    st.success(f"Premium removed from: {row.get('email', '')}")
                    st.rerun()

        with c6:
            if st.button(
                "Reset Sessions",
                key="reset_sessions_button",
                use_container_width=True,
            ):
                if selected_user_label:
                    row = user_options[selected_user_label]
                    reset_user_sessions(row.get("email", ""), row.get("sub", ""))
                    st.success(f"Sessions reset for: {row.get('email', '')}")
                    st.rerun()

        with c8:
            if st.button(
                "Remove From Company",
                key="remove_from_company_button",
                use_container_width=True,
            ):
                if selected_user_label:
                    row = user_options[selected_user_label]
                    remove_user_from_company(row.get("email", ""), row.get("sub", ""))
                    st.success(f"User removed from company: {row.get('email', '')}")
                    st.rerun()

    st.markdown("---")
    st.markdown("### Company Workspaces")

    companies_registry = load_companies_registry()
    companies_for_view = []
    for company in companies_registry:
        company = ensure_company_fields(company)
        companies_for_view.append(
            {
                "Key": company.get("key", ""),
                "Name": company.get("name", ""),
                "Domain": company.get("domain", ""),
                "Billing": company.get("billing_status", ""),
                "Seats": format_company_seats(company),
                "Max Seats": company.get("max_seats", 0),
                "Owner Email": company.get("owner_email", ""),
                "Shared Workspace": company.get("shared_workspace_enabled", False),
                "Plan Start": company.get("plan_start", ""),
                "Plan End": company.get("plan_end", ""),
                "Active": company.get("is_active", True),
                "Trial End": company.get("trial_end", ""),
            }
        )

    if companies_for_view:
        st.dataframe(
            pd.DataFrame(companies_for_view),
            use_container_width=True,
            hide_index=True,
        )
    else:
        st.info("No company workspaces found yet.")

    st.markdown("#### Create / Update Company Workspace")

    cc1, cc2, cc3, cc4 = st.columns(4)
    with cc1:
        company_key_input = st.text_input(
            "Company Key",
            key="company_key_input",
            placeholder="knauf_team",
        )
    with cc2:
        company_name_input = st.text_input(
            "Company Name",
            key="company_name_input",
            placeholder="Knauf",
        )
    with cc3:
        company_domain_input = st.text_input(
            "Company Domain",
            key="company_domain_input",
            placeholder="knauf.com",
        )
    with cc4:
        company_max_seats_input = st.number_input(
            "Max Seats",
            min_value=0,
            max_value=1000,
            value=3,
            step=1,
            key="company_max_seats_input",
        )

    company_owner_email = st.text_input(
        "Owner Email",
        key="company_owner_email",
        placeholder="owner@company.com",
    )
    company_active_input = st.checkbox(
        "Company Workspace Active",
        value=True,
        key="company_active_input",
    )
    company_shared_workspace_input = st.checkbox(
        "Enable Shared Company Workspace",
        value=False,
        key="company_shared_workspace_input",
        help="When enabled, users of this company share the same saved Comparisons.",
    )

    cp1, cp2 = st.columns(2)
    with cp1:
        company_plan_start_input = st.date_input(
            "Company Plan Start",
            value=None,
            key="company_plan_start_input",
        )
    with cp2:
        company_plan_end_input = st.date_input(
            "Company Plan End",
            value=None,
            key="company_plan_end_input",
        )

    ccu1, ccu2 = st.columns(2)

    with ccu1:
        if st.button(
            "Save Company Workspace",
            key="save_company_plan_button",
            use_container_width=True,
        ):
            if not company_key_input.strip():
                st.warning("Please enter a company key.")
            elif not company_name_input.strip():
                st.warning("Please enter a company name.")
            elif not company_domain_input.strip():
                st.warning("Please enter a company domain.")
            else:
                upsert_company(
                    company_key_input,
                    company_name_input,
                    company_domain_input,
                    company_max_seats_input,
                    company_active_input,
                    "trialing",
                    company_owner_email,
                    company_shared_workspace_input,
                    company_plan_start_input,
                    company_plan_end_input,
                )
                st.success(f"Company workspace saved: {company_name_input}")
                st.rerun()

    company_options_for_delete = {
        f"{c.get('name', '')} ({c.get('domain', '')})": c.get("key")
        for c in companies_registry
    }

    with ccu2:
        selected_company_to_delete = st.selectbox(
            "Delete Company Workspace",
            [""] + list(company_options_for_delete.keys()),
            key="selected_company_to_delete",
        )
        if st.button(
            "Delete Company Workspace",
            key="delete_company_plan_button",
            use_container_width=True,
        ):
            if selected_company_to_delete:
                remove_company(company_options_for_delete[selected_company_to_delete])
                st.success(
                    f"Company workspace deleted: {selected_company_to_delete}"
                )
                st.rerun()
    st.markdown("</div>", unsafe_allow_html=True)


# -------------------------------------------------
# RENDER CURRENT VIEW
# -------------------------------------------------
if current_view == "Company Manager":
    render_company_manager()
elif current_view == "Sources":
    render_sources()
elif current_view == "Comparisons":
    render_comparisons()
elif current_view == "Admin Panel":
    render_admin_panel()